"""
Loader condiviso per le mappature carte carburante (Edenred UTA, Enilive).

Espone XlsxBackedCardMap: una proiezione dict-like (read-only) su un file XLSX
in input/ che si AUTO-RICARICA quando il file su disco viene modificato
(check su mtime ad ogni accesso).

Schema XLSX atteso (header su riga 1):
    numero_carta | targa | nickname | classificazione_fiscale | stato

Solo righe con stato='ATTIVO' vengono incluse.
"""
from pathlib import Path
from typing import Optional
import threading

import openpyxl


CLASSIFICAZIONI_AMMESSE = {"POOL", "uso_promiscuo", "super_lusso", "SERVIZIO"}
STATI_AMMESSI = {"ATTIVO", "DISMESSO"}


def _norm_numero(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _norm_targa(raw):
    return str(raw).upper().strip() if raw else ""


def _norm_nickname(raw):
    return str(raw).strip() if raw else ""


def _norm_classif(raw):
    return str(raw).strip() if raw else ""


def _norm_stato(raw):
    if not raw:
        return "ATTIVO"
    s = str(raw).strip().upper()
    return s if s in STATI_AMMESSI else "ATTIVO"


_KNOWN_SHEETS = ("CARTE UTA", "Carte UTA", "CARTE ENILIVE", "Carte Enilive",
                 "Foglio1", "Sheet1")


def parse_xlsx_to_dict(xlsx_path: Path) -> dict:
    """Parsing puro: XLSX → dict[numero_carta] = info. Idempotente, no side effects."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = next((s for s in _KNOWN_SHEETS if s in wb.sheetnames),
                      wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {}
    for i, name in enumerate(header):
        n = name.lower()
        if n == "numero_carta":
            col["numero_carta"] = i
        elif n == "targa":
            col["targa"] = i
        elif n == "nickname":
            col["nickname"] = i
        elif "classificazione" in n:
            col["classificazione"] = i
        elif n == "stato":
            col["stato"] = i

    missing = {"numero_carta", "classificazione"} - set(col.keys())
    if missing:
        raise ValueError(
            f"Colonne obbligatorie mancanti in {xlsx_path.name} (foglio {sheet_name!r}): "
            f"{missing}. Header letto: {header}"
        )

    out = {}
    for row in rows[1:]:
        if not row or not any(row):
            continue
        numero = _norm_numero(row[col["numero_carta"]]
                              if col["numero_carta"] < len(row) else None)
        if not numero:
            continue
        classif = _norm_classif(row[col["classificazione"]]
                                if col["classificazione"] < len(row) else None)
        if not classif:
            continue
        stato = (_norm_stato(row[col["stato"]])
                 if "stato" in col and col["stato"] < len(row) else "ATTIVO")
        if stato != "ATTIVO":
            continue
        targa = (_norm_targa(row[col["targa"]])
                 if "targa" in col and col["targa"] < len(row) else "")
        nickname = (_norm_nickname(row[col["nickname"]])
                    if "nickname" in col and col["nickname"] < len(row) else "")
        out[numero] = {
            "numero_carta": numero,
            "targa": targa,
            "nickname": nickname,
            "classificazione": classif,
            "stato": stato,
        }
    return out


class XlsxBackedCardMap:
    """
    Mappa carte caricata da XLSX con auto-refresh su mtime.

    Si comporta come un dict read-only: supporta `[]`, `get`, `in`, `len`,
    `iter`, `keys`, `values`, `items`. Ad ogni accesso fa stat() sul file e
    ricarica se mtime è cambiato. Thread-safe (lock interno sulla ricarica).

    Se il file XLSX manca o l'ultimo parse è fallito, ritorna le ultime carte
    note (o {} se è il primo accesso).
    """

    def __init__(self, xlsx_path):
        self._path = Path(xlsx_path)
        self._mtime = None
        self._cache = {}
        self._lock = threading.Lock()

    def _refresh(self):
        if not self._path.exists():
            return
        try:
            cur = self._path.stat().st_mtime
        except OSError:
            return
        if cur == self._mtime and self._cache:
            return
        with self._lock:
            # Double-check sotto lock
            try:
                cur = self._path.stat().st_mtime
            except OSError:
                return
            if cur == self._mtime and self._cache:
                return
            try:
                self._cache = parse_xlsx_to_dict(self._path)
                self._mtime = cur
            except Exception:
                # Mantieni la cache precedente in caso di parsing rotto:
                # evita di rompere il writer in produzione se l'utente sta
                # editando l'XLSX in quel momento.
                pass

    # --- dict-like read-only API ---
    def __getitem__(self, k):
        self._refresh()
        return self._cache[k]

    def __contains__(self, k):
        self._refresh()
        return k in self._cache

    def __iter__(self):
        self._refresh()
        return iter(dict(self._cache))  # snapshot per iterazione safe

    def __len__(self):
        self._refresh()
        return len(self._cache)

    def get(self, k, default=None):
        self._refresh()
        return self._cache.get(k, default)

    def keys(self):
        self._refresh()
        return list(self._cache.keys())

    def values(self):
        self._refresh()
        return list(self._cache.values())

    def items(self):
        self._refresh()
        return list(self._cache.items())

    def __repr__(self):
        self._refresh()
        return (f"<XlsxBackedCardMap path={self._path.name!r} "
                f"carte={len(self._cache)} mtime={self._mtime}>")
