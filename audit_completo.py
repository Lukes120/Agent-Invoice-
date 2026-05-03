"""
Audit COMPLETO automazione fatture passive Ecotel.
- 10 mesi di storico
- TUTTI i fornitori delle fatture in attesa (no esclusioni)
- Analisi pattern per famiglia + opportunità tecniche
- SOLO LETTURA su Odoo

Output:
  output/audit_completo_<timestamp>/
    audit_completo.xlsx          # tabella completa per fornitore
    audit_completo.md            # sintesi categorizzazione + ranking
    roadmap_automazione.md       # roadmap dettagliata per fasi/famiglie
    raccomandazioni_codice.md    # proposte modifiche parser/classifier/writer
    audit_completo.json          # dump dati per uso futuro
    audit_completo.log
"""

import os
import sys
import re
import json
import base64
import logging
import datetime as dt
from pathlib import Path
from collections import Counter, defaultdict
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core.odoo_client import OdooReadOnlyClient
from core.fatturapa_parser import parse_fatturapa_xml

COMPANY_ID = 1
HISTORICAL_MONTHS = 10
SOGLIA_FATTURE_CANDIDATO = 4

# Già mappati o in fase di mappatura
ALREADY_MAPPED = {
    'IT05403151003', 'IT09247981005', 'IT00488410010',
    'IT13378520152', 'IT12874490159',
}


def setup_logging(out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    log_file = out_dir / 'audit_completo.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[logging.FileHandler(log_file, encoding='utf-8'),
                  logging.StreamHandler(sys.stdout)]
    )
    return logging.getLogger('audit')


# ============================================================
# RACCOLTA DATI
# ============================================================

def fetch_pending(client, log):
    domain = [('registered', '=', False), ('is_self_invoice', '=', False),
              ('company_id', '=', COMPANY_ID)]
    atts = client._call('fatturapa.attachment.in', 'search_read', domain,
        fields=['id', 'name', 'xml_supplier_id', 'datas',
                'invoices_total', 'invoices_date'])
    log.info(f"Fatture in attesa Ecotel: {len(atts)}")
    return atts


def parse_pending(atts, log):
    by_vat = defaultdict(list)
    for i, att in enumerate(atts, 1):
        try:
            raw = base64.b64decode(att['datas']).decode('utf-8', errors='replace')
            xd = parse_fatturapa_xml(raw)
            vat = (xd.cedente_partita_iva or '').strip().upper()
            if not vat:
                continue
            sup = att.get('xml_supplier_id')
            partner_id = sup[0] if isinstance(sup, list) else None
            supplier_name = xd.cedente_denominazione or (sup[1] if isinstance(sup, list) and len(sup) > 1 else '')

            # Estrazione campi extra utili per scoperta routing
            xml_extra = extract_routing_fields(raw)

            by_vat[vat].append({
                'attachment_id': att['id'],
                'partner_id': partner_id,
                'supplier_name': supplier_name,
                'numero': xd.numero,
                'data': xd.data,
                'tipo_doc': xd.tipo_documento,
                'imponibile': xd.imponibile_totale,
                'totale': xd.importo_totale or att.get('invoices_total', 0),
                'n_righe': len(xd.righe),
                'oda_riferimenti': list(xd.oda_riferimenti),
                'commessa_riferimenti': list(xd.commessa_riferimenti),
                'contratto_riferimenti': list(xd.contratto_riferimenti),
                'ricezione_riferimenti': list(xd.ricezione_riferimenti),
                'rif_amministrazione': xd.cedente_riferimento_amministrazione,
                'pod_riferimenti': list(xd.pod_riferimenti),
                'descrizioni': [r.descrizione for r in xd.righe if r.descrizione],
                'aliquote_uniche': sorted({r.aliquota_iva for r in xd.righe}),
                'xml_extra': xml_extra,
            })
        except Exception as e:
            log.warning(f"Parse err att {att.get('id')}: {e}")
        if i % 30 == 0:
            log.info(f"  parsing {i}/{len(atts)}")
    return by_vat


def extract_routing_fields(raw):
    """Estrae campi XML potenzialmente utili per routing che oggi non parsiamo."""
    out = {}
    patterns = {
        'CodiceCommessaConvenzione': r'<CodiceCommessaConvenzione>([^<]+)</CodiceCommessaConvenzione>',
        'CodiceCIG': r'<CodiceCIG>([^<]+)</CodiceCIG>',
        'CodiceCUP': r'<CodiceCUP>([^<]+)</CodiceCUP>',
        'RifAmministrazioneCessionario': r'<RiferimentoAmministrazione>([^<]+)</RiferimentoAmministrazione>',
    }
    for k, p in patterns.items():
        ms = re.findall(p, raw)
        if ms:
            out[k] = list(set(ms))
    # AltriDatiGestionali nelle DettaglioLinee
    altri_dati = re.findall(
        r'<TipoDato>([^<]+)</TipoDato>\s*<RiferimentoTesto>([^<]+)</RiferimentoTesto>', raw)
    if altri_dati:
        out['AltriDatiGestionali'] = list({(t.strip(), v.strip()) for t, v in altri_dati})[:10]
    return out


def fetch_history(client, partner_id, cutoff):
    if not partner_id:
        return []
    domain = [('partner_id', '=', partner_id),
              ('move_type', 'in', ['in_invoice', 'in_refund']),
              ('state', '=', 'posted'),
              ('invoice_date', '>=', cutoff),
              ('company_id', '=', COMPANY_ID)]
    return client._call('account.move', 'search_read', domain,
        fields=['id', 'name', 'ref', 'invoice_date', 'amount_total', 'amount_untaxed',
                'invoice_origin', 'invoice_line_ids', 'move_type'])


def fetch_lines(client, line_ids):
    if not line_ids:
        return []
    return client._call('account.move.line', 'read', line_ids,
        fields=['id', 'name', 'quantity', 'price_unit', 'price_subtotal',
                'account_id', 'tax_ids', 'product_id', 'purchase_line_id'])


def fetch_open_pos(client, partner_id):
    if not partner_id:
        return []
    domain = [('partner_id', '=', partner_id), ('state', '=', 'purchase'),
              ('company_id', '=', COMPANY_ID)]
    return client._call('purchase.order', 'search_read', domain,
        fields=['id', 'name', 'partner_ref', 'amount_total', 'amount_untaxed',
                'date_order', 'order_line', 'date_planned'])


def count_libere(client, line_ids):
    if not line_ids:
        return 0, 0, []
    lines = client._call('purchase.order.line', 'read', line_ids,
        fields=['id', 'name', 'product_qty', 'qty_invoiced', 'qty_received',
                'price_unit', 'date_planned'])
    libere = [l for l in lines
              if (l.get('qty_invoiced') or 0) == 0
              and (l.get('qty_received') or 0) == 0
              and (l.get('product_qty') or 0) >= 1]
    return len(libere), len(lines), libere


# ============================================================
# ANALISI FORNITORE
# ============================================================

def analyze_supplier(client, vat, supplier_name, partner_id, pending, cutoff, log):
    rec = {
        'vat': vat,
        'supplier_name': supplier_name,
        'partner_id': partner_id,
        'already_mapped': vat in ALREADY_MAPPED,
        'n_pending': len(pending),
        'pending_total': sum((p['totale'] or 0) for p in pending),
        'n_pending_nc': sum(1 for p in pending if p['tipo_doc'] == 'TD04'),
        'n_posted': 0,
        'posted_total': 0.0,
        'n_posted_nc': 0,
        'cadenza_giorni': None,
        'oda_pct': 0.0,
        'oda_distinct': [],
        'oda_aperti': [],
        'oda_aperti_con_libere': 0,
        'libere_totali': 0,
        'conti_distinct': [],
        'conti_top_pct': 0.0,
        'iva_distinct': [],
        'iva_top_pct': 0.0,
        'commesse_pending': [],
        'multi_routing_hint': False,
        'routing_keys_seen': {},
        'descrizioni_keywords': [],
        'has_bolli': False,
        'category': 'D',
        'category_reason': '',
        'family': 'altro',
        'mapping_strategy': '',
        'notes': [],
    }

    if not partner_id:
        rec['category_reason'] = 'partner_id mancante'
        return rec

    # Storico
    moves = fetch_history(client, partner_id, cutoff)
    rec['n_posted'] = len(moves)
    rec['posted_total'] = sum((m.get('amount_total') or 0) for m in moves)
    rec['n_posted_nc'] = sum(1 for m in moves if m.get('move_type') == 'in_refund')

    # Cadenza (tra prima e ultima posted, in giorni medi)
    dates = sorted([m['invoice_date'] for m in moves if m.get('invoice_date')])
    if len(dates) >= 2:
        d1 = dt.date.fromisoformat(dates[0])
        dN = dt.date.fromisoformat(dates[-1])
        delta = (dN - d1).days
        rec['cadenza_giorni'] = round(delta / max(1, len(dates) - 1), 1)

    # Righe move
    all_lines_ids = []
    for m in moves:
        all_lines_ids.extend(m.get('invoice_line_ids', []) or [])
    move_lines = fetch_lines(client, all_lines_ids) if all_lines_ids else []

    # OdA usage
    n_with_origin = sum(1 for m in moves if m.get('invoice_origin'))
    if moves:
        rec['oda_pct'] = n_with_origin / len(moves) * 100
    oda_counter = Counter()
    for m in moves:
        if m.get('invoice_origin'):
            oda_counter[m['invoice_origin'].strip()] += 1
    rec['oda_distinct'] = oda_counter.most_common(15)

    # Conti / IVA
    cnt_conti = Counter()
    cnt_iva = Counter()
    for ln in move_lines:
        acc = ln.get('account_id')
        if acc and isinstance(acc, list):
            cnt_conti[acc[1]] += 1
        tax = ln.get('tax_ids') or []
        if tax:
            cnt_iva[tuple(sorted(tax))] += 1
        # Bolli: tax_ids include 47 + descrizione "bollo"
        if 47 in tax and 'bollo' in (ln.get('name') or '').lower():
            rec['has_bolli'] = True
    rec['conti_distinct'] = cnt_conti.most_common(10)
    rec['iva_distinct'] = cnt_iva.most_common(10)
    n_lines = sum(c[1] for c in rec['conti_distinct'])
    n_lines_iva = sum(c[1] for c in rec['iva_distinct'])
    rec['conti_top_pct'] = (rec['conti_distinct'][0][1] / n_lines * 100) if n_lines else 0
    rec['iva_top_pct'] = (rec['iva_distinct'][0][1] / n_lines_iva * 100) if n_lines_iva else 0

    # Commesse pending
    for p in pending:
        for c in p.get('commessa_riferimenti', []):
            if c not in rec['commesse_pending']:
                rec['commesse_pending'].append(c)

    # Multi-routing hint: più valori di POD/RifAmm/contratto tra le fatture in attesa
    rks = defaultdict(set)
    for p in pending:
        for v in p.get('pod_riferimenti', []):
            rks['pod'].add(v)
        if p.get('rif_amministrazione'):
            rks['rif_amm'].add(p['rif_amministrazione'])
        for v in p.get('contratto_riferimenti', []):
            rks['contratto'].add(v)
        for v in p.get('ricezione_riferimenti', []):
            rks['ricezione'].add(v)
    rec['routing_keys_seen'] = {k: sorted(list(v)) for k, v in rks.items()}
    rec['multi_routing_hint'] = any(len(v) > 1 for v in rks.values())

    # Keywords descrizioni (analisi opportunità line_groups)
    all_descs = []
    for p in pending:
        all_descs.extend(p.get('descrizioni', []))
    rec['descrizioni_keywords'] = extract_keywords(all_descs)

    # OdA aperti
    pos = fetch_open_pos(client, partner_id)
    pos_summary = []
    for po in pos:
        nl, tl, libere = count_libere(client, po.get('order_line', []))
        pos_summary.append({
            'name': po['name'],
            'partner_ref': po.get('partner_ref') or '',
            'amount_total': po.get('amount_total') or 0,
            'libere_n': nl,
            'total_lines': tl,
            'libere_descrizioni': [l.get('name', '')[:80] for l in libere[:5]],
        })
        if nl > 0:
            rec['oda_aperti_con_libere'] += 1
            rec['libere_totali'] += nl
    rec['oda_aperti'] = pos_summary

    # Famiglia + categoria + strategy
    rec['family'] = guess_family(supplier_name, rec)
    cat, reason, strategy, notes = categorize(rec)
    rec['category'] = cat
    rec['category_reason'] = reason
    rec['mapping_strategy'] = strategy
    rec['notes'] = notes

    return rec


# ============================================================
# CLASSIFICAZIONE
# ============================================================

FAMILY_RULES = [
    ('leasing', re.compile(r'leasing|car\s*lease|leasys|athlon|arval|alphabet|alphabet|ald\b|automotive|noleggio.*lung', re.I)),
    ('factoring', re.compile(r'factoring|factor\b|mbfacta|sarda', re.I)),
    ('utility_energia', re.compile(r'sorgenia|enilive|enel|nwg|a2a|acea|edison|engie|eni\s|hera\b', re.I)),
    ('utility_gas_acqua', re.compile(r'\bgas\b|acqua|idriche|aqp\b|smat', re.I)),
    ('telco', re.compile(r'telecom|tim\b|fastweb|wind|vodafone|netreality|tiscali', re.I)),
    ('trasporti', re.compile(r'trenitalia|italo|telepass|frecciarossa|ntv', re.I)),
    ('hotel_ristorazione', re.compile(r'hotel|ristoran|albergo|b&b|catering|edenred|pellegrini|pam\b|conviviale', re.I)),
    ('autonoleggio_breve', re.compile(r'europcar|hertz|avis|sixt|locauto|maggiore|car2go', re.I)),
    ('carburanti', re.compile(r'eni\s*card|q8|ip\s*spa|tamoil|esso|edenred.*uta|fuel|carburant', re.I)),
    ('retail_materiali', re.compile(r'wuerth|sonepar|rexel|tarlazzi|mef\b|com[-\s]*cavi|hilti|abbattista|elettrica|lyreco|esprinet|comet', re.I)),
    ('servizi_professionali', re.compile(r'studio|consulen|formazione|avvocat|notaio|commercialist|aruba|infocert|tinexta', re.I)),
    ('manutenzione', re.compile(r'manuten|atecna|coel|tecnoalt|gam\s+engineering|electro\s*rent', re.I)),
    ('cleaning_safety', re.compile(r'cleanic|sebach|sir\s*safety|bio\s*clean|dot\s*impresa|formatori', re.I)),
    ('logistica', re.compile(r'fedex|tnt|bartolini|gls|brt|spedizion|tuscia|nlg\b', re.I)),
    ('immobiliare', re.compile(r'real\s*estate|immobil|locazion|affitto|pellegrini', re.I)),
]


def guess_family(name, rec):
    s = name.lower()
    for fam, rx in FAMILY_RULES:
        if rx.search(s):
            return fam
    return 'altro'


def categorize(rec):
    notes = []
    n_total = rec['n_pending'] + rec['n_posted']

    if rec['already_mapped']:
        return 'MAPPED', 'Già mappato', 'n/a', notes

    if rec['commesse_pending']:
        notes.append(f"Commesse rilevate: {rec['commesse_pending'][:3]}")
        return 'D', 'Commesse S##### → Fase 2', 'Skip; gestione manuale (fase 2 commesse)', notes

    if n_total < SOGLIA_FATTURE_CANDIDATO:
        return 'D', f'Solo {n_total} fatt./10m', 'Skip; manuale', notes

    has_oda = rec['oda_pct'] >= 80
    conti_stabili = rec['conti_top_pct'] >= 80
    iva_stabile = rec['iva_top_pct'] >= 80
    has_ledger = rec['oda_aperti_con_libere'] >= 1
    multi_routing = rec['multi_routing_hint']
    has_bolli = rec['has_bolli']

    # Caso A: tutto stabile + ledger pronto
    if has_oda and conti_stabili and iva_stabile and has_ledger:
        if multi_routing:
            notes.append(f"Multi-routing: {rec['routing_keys_seen']}")
            strategy = 'multi-contratto + libere_criterio=standard_qty_inv_rec'
        else:
            strategy = 'mapping diretto + libere_criterio=standard_qty_inv_rec'
        return 'A', 'pronto: ledger esistente + pattern stabile', strategy, notes

    # Caso B: stabile ma ledger da predisporre
    if conti_stabili and iva_stabile and not has_ledger:
        notes.append('OdA ledger da predisporre (chiedere a contabilità)')
        strategy = ('predisporre OdA-ledger ricorrente, poi mapping. '
                    + ('Multi-routing rilevato → 1 OdA per chiave.' if multi_routing else ''))
        return 'B', 'pattern stabile ma OdA non ledger', strategy, notes

    # Caso C-multi-conto (line_groups candidate)
    if has_oda and not conti_stabili and rec['conti_top_pct'] >= 30:
        notes.append('Più conti contabili → candidato line_groups (come Telecom P04516)')
        strategy = 'analisi descrizioni linee per definire line_groups'
        return 'C', f"multi-conto (top {rec['conti_top_pct']:.0f}%)", strategy, notes

    # Caso C-multi-IVA (es. canone+bollo+rate)
    if has_oda and not iva_stabile:
        notes.append('Più aliquote IVA → candidato line_groups + tax per gruppo')
        strategy = 'multi-IVA: line_groups con taxes_id specifico'
        return 'C', f"multi-IVA (top {rec['iva_top_pct']:.0f}%)", strategy, notes

    # Caso C-NoOda: pattern stabile ma fatture senza OdA
    if not has_oda and conti_stabili and iva_stabile:
        notes.append('Pattern stabile ma fatture spesso senza OdA → candidato fornitore fisso senza ricerca OdA')
        strategy = 'mapping fornitore fisso (no oda lookup) o predisporre ledger'
        return 'C', f"pattern stabile ma OdA solo nel {rec['oda_pct']:.0f}% delle fatture", strategy, notes

    # Default C eterogeneo
    return 'C', 'pattern eterogeneo', 'caso per caso', notes


def extract_keywords(descrizioni, top_n=10):
    if not descrizioni:
        return []
    # Estrae parole significative (>=4 char, non numeriche)
    word_counter = Counter()
    for d in descrizioni:
        words = re.findall(r'[A-Za-zÀ-ÿ]{4,}', d.lower())
        for w in words:
            if w not in {'della', 'delle', 'dello', 'dalla', 'sulla', 'come', 'sono',
                         'anche', 'questo', 'questa', 'molto', 'oltre', 'fattura',
                         'rifer', 'periodo', 'cliente', 'codice'}:
                word_counter[w] += 1
    return word_counter.most_common(top_n)


# ============================================================
# OUTPUT
# ============================================================

def write_excel(records, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"

    headers = [
        'Cat', 'Famiglia', 'Fornitore', 'P.IVA', 'partner_id',
        'N pend', 'NC pend', 'Tot pend €',
        'N posted', 'NC posted', 'Tot posted €',
        'Cadenza gg',
        '% OdA', 'N OdA distinti', 'OdA top',
        'Top conto', '% top conto',
        'Top IVA', '% top IVA',
        'OdA aperti', 'OdA con libere', 'Tot righe libere',
        'Multi-routing', 'Routing keys',
        'Bolli',
        'Motivo', 'Strategia', 'Note'
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='DDDDDD')

    cat_color = {'A': 'C6EFCE', 'B': 'FFEB9C', 'C': 'FFCC99', 'D': 'F2F2F2', 'MAPPED': 'B4C7E7'}

    sorted_recs = sorted(records, key=lambda r: (
        ['A', 'B', 'C', 'MAPPED', 'D'].index(r['category']),
        -(r['n_pending'] + r['n_posted'])
    ))

    for r in sorted_recs:
        oda_top = ", ".join(f"{n}({c}x)" for n, c in r['oda_distinct'][:3])
        top_conto = r['conti_distinct'][0][0] if r['conti_distinct'] else ''
        top_iva = list(r['iva_distinct'][0][0]) if r['iva_distinct'] else []
        oda_aperti_str = ", ".join(p['name'] for p in r['oda_aperti'][:8])
        routing_str = '; '.join(f"{k}={v}" for k, v in r['routing_keys_seen'].items() if v)

        row = [
            r['category'], r['family'], r['supplier_name'], r['vat'], r['partner_id'],
            r['n_pending'], r['n_pending_nc'], round(r['pending_total'], 2),
            r['n_posted'], r['n_posted_nc'], round(r['posted_total'], 2),
            r['cadenza_giorni'] or '',
            round(r['oda_pct'], 1),
            len(r['oda_distinct']),
            oda_top,
            top_conto, round(r['conti_top_pct'], 1),
            str(top_iva), round(r['iva_top_pct'], 1),
            oda_aperti_str, r['oda_aperti_con_libere'], r['libere_totali'],
            'sì' if r['multi_routing_hint'] else 'no',
            routing_str[:100],
            'sì' if r['has_bolli'] else 'no',
            r['category_reason'], r['mapping_strategy'],
            ' | '.join(r['notes']),
        ]
        ws.append(row)
        last_row = ws.max_row
        ws.cell(row=last_row, column=1).fill = PatternFill('solid',
            fgColor=cat_color.get(r['category'], 'F2F2F2'))

    widths = [8, 20, 35, 16, 10, 6, 6, 12, 8, 6, 14, 8, 7, 6, 25,
              30, 8, 12, 8, 30, 8, 8, 10, 30, 6, 35, 50, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'A2'

    # Foglio per famiglia (aggregati)
    ws2 = wb.create_sheet("Per famiglia")
    ws2.append(['Famiglia', 'N fornitori', 'Cat A', 'Cat B', 'Cat C', 'Cat D',
                'N fatt. pend.', 'Tot pend. €', 'N fatt. posted', 'Tot posted €'])
    fam_aggr = defaultdict(lambda: {'recs': [], 'A': 0, 'B': 0, 'C': 0, 'D': 0,
                                    'pend': 0, 'pend_eur': 0, 'post': 0, 'post_eur': 0})
    for r in records:
        if r['category'] == 'MAPPED':
            continue
        f = fam_aggr[r['family']]
        f['recs'].append(r)
        f[r['category']] = f.get(r['category'], 0) + 1
        f['pend'] += r['n_pending']
        f['pend_eur'] += r['pending_total']
        f['post'] += r['n_posted']
        f['post_eur'] += r['posted_total']
    for fam, data in sorted(fam_aggr.items(), key=lambda x: -x[1]['pend_eur']):
        ws2.append([fam, len(data['recs']), data['A'], data['B'], data['C'], data['D'],
                    data['pend'], round(data['pend_eur'], 2),
                    data['post'], round(data['post_eur'], 2)])

    wb.save(out_path)


def write_md_summary(records, out_dir, period_from, period_to):
    out_path = out_dir / 'audit_completo.md'
    lines = []
    lines.append(f"# Audit completo Ecotel — {dt.date.today()}")
    lines.append("")
    lines.append(f"**Periodo storico**: {period_from} → {period_to} ({HISTORICAL_MONTHS} mesi)")
    lines.append(f"**Ambito**: Ecotel Italia (company_id=1), tutti i fornitori delle fatture in attesa.")
    lines.append("")

    n = len(records)
    by_cat = Counter(r['category'] for r in records)
    by_fam = Counter(r['family'] for r in records if r['category'] != 'MAPPED')
    pend_count = sum(r['n_pending'] for r in records)
    pend_amt = sum(r['pending_total'] for r in records)
    post_count = sum(r['n_posted'] for r in records)

    lines.append("## Panoramica")
    lines.append("")
    lines.append(f"- Fornitori totali: **{n}** (di cui già mappati: {by_cat.get('MAPPED', 0)})")
    lines.append(f"- Fatture in attesa: **{pend_count}** (€{pend_amt:,.2f})")
    lines.append(f"- Fatture posted nei {HISTORICAL_MONTHS} mesi: **{post_count}**")
    lines.append("")
    lines.append("### Categorizzazione")
    lines.append("")
    lines.append(f"| Cat | N | Significato |")
    lines.append("| --- | ---: | --- |")
    lines.append(f"| **A** | {by_cat.get('A', 0)} | Pronti subito (ledger esistenti, conto/IVA stabili) |")
    lines.append(f"| **B** | {by_cat.get('B', 0)} | Pattern stabile, OdA-ledger da predisporre |")
    lines.append(f"| **C** | {by_cat.get('C', 0)} | Complessi (multi-conto, multi-IVA, eterogenei) |")
    lines.append(f"| **D** | {by_cat.get('D', 0)} | <{SOGLIA_FATTURE_CANDIDATO} fatt. o commesse Fase 2 |")
    lines.append(f"| MAPPED | {by_cat.get('MAPPED', 0)} | Già automatizzati |")
    lines.append("")

    lines.append("### Famiglie più presenti")
    lines.append("")
    lines.append("| Famiglia | N fornitori |")
    lines.append("| --- | ---: |")
    for fam, n in by_fam.most_common(15):
        lines.append(f"| {fam} | {n} |")
    lines.append("")

    # Cat A details
    a_recs = sorted([r for r in records if r['category'] == 'A'],
                    key=lambda x: -(x['n_pending'] + x['n_posted']))
    lines.append(f"## Cat A — pronti subito ({len(a_recs)})")
    lines.append("")
    for r in a_recs:
        n_total = r['n_pending'] + r['n_posted']
        top_conto = r['conti_distinct'][0][0] if r['conti_distinct'] else '?'
        top_iva = list(r['iva_distinct'][0][0]) if r['iva_distinct'] else []
        lines.append(f"### {r['supplier_name']} ({r['vat']}) — *{r['family']}*")
        lines.append(f"- Fatture: {r['n_pending']} pend + {r['n_posted']} posted = **{n_total}**")
        lines.append(f"- Conto: {top_conto} ({r['conti_top_pct']:.0f}%) — IVA: {top_iva}")
        lines.append(f"- OdA con righe libere: {r['oda_aperti_con_libere']} ({r['libere_totali']} righe libere totali)")
        if r['multi_routing_hint']:
            lines.append(f"- ⚠ Multi-routing: {r['routing_keys_seen']}")
        lines.append(f"- **Strategia**: {r['mapping_strategy']}")
        lines.append("")

    # Cat B details
    b_recs = sorted([r for r in records if r['category'] == 'B'],
                    key=lambda x: -(x['n_pending'] + x['n_posted']))
    lines.append(f"## Cat B — predisporre OdA-ledger ({len(b_recs)})")
    lines.append("")
    for r in b_recs:
        n_total = r['n_pending'] + r['n_posted']
        top_conto = r['conti_distinct'][0][0] if r['conti_distinct'] else '?'
        top_iva = list(r['iva_distinct'][0][0]) if r['iva_distinct'] else []
        lines.append(f"### {r['supplier_name']} ({r['vat']}) — *{r['family']}*")
        lines.append(f"- Fatture: {r['n_pending']} pend + {r['n_posted']} posted = **{n_total}**")
        lines.append(f"- Cadenza media: {r['cadenza_giorni']} giorni")
        lines.append(f"- Conto: {top_conto} — IVA: {top_iva}")
        if r['multi_routing_hint']:
            lines.append(f"- ⚠ Multi-routing rilevato (1 OdA per chiave): {r['routing_keys_seen']}")
        lines.append(f"- **Strategia**: {r['mapping_strategy']}")
        lines.append("")

    # Cat C breakdown
    c_recs = sorted([r for r in records if r['category'] == 'C'],
                    key=lambda x: -(x['n_pending'] + x['n_posted']))
    lines.append(f"## Cat C — complessi ({len(c_recs)})")
    lines.append("")
    lines.append("Aggregati per famiglia (vedi `roadmap_automazione.md` per strategie):")
    lines.append("")
    c_by_fam = defaultdict(list)
    for r in c_recs:
        c_by_fam[r['family']].append(r)
    for fam, items in sorted(c_by_fam.items(), key=lambda x: -sum(r['n_pending'] for r in x[1])):
        n_pend = sum(r['n_pending'] for r in items)
        n_post = sum(r['n_posted'] for r in items)
        lines.append(f"### {fam} ({len(items)} fornitori, {n_pend} pend + {n_post} posted)")
        for r in items:
            n_total = r['n_pending'] + r['n_posted']
            lines.append(f"- **{r['supplier_name']}** ({r['vat']}) — {n_total} fatt. — {r['category_reason']}")
        lines.append("")

    out_path.write_text('\n'.join(lines), encoding='utf-8')


def write_roadmap(records, out_dir):
    out_path = out_dir / 'roadmap_automazione.md'
    lines = []
    lines.append("# Roadmap automazione fatture passive")
    lines.append("")
    lines.append("Strategie di automazione organizzate per famiglie e ondate.")
    lines.append("")

    by_fam = defaultdict(list)
    for r in records:
        if r['category'] in ('A', 'B', 'C'):
            by_fam[r['family']].append(r)

    for fam in sorted(by_fam.keys(), key=lambda f: -sum(r['n_pending'] for r in by_fam[f])):
        items = by_fam[fam]
        n_pend = sum(r['n_pending'] for r in items)
        n_post = sum(r['n_posted'] for r in items)
        lines.append(f"## Famiglia: {fam}")
        lines.append("")
        lines.append(f"**Fornitori**: {len(items)} — **Fatture**: {n_pend} pend + {n_post} posted")
        lines.append("")

        # Specific strategies per family
        if fam == 'leasing':
            lines.append("### Pattern tipico leasing veicoli")
            lines.append("")
            lines.append("- Canone mensile/trimestrale per ogni veicolo")
            lines.append("- Bollo (IVA esente, tax 47) ricorrente")
            lines.append("- Quota interessi (IVA esente, tax 51)")
            lines.append("- Eventuali rate finali")
            lines.append("- Conti diversi per canone vs bollo (es. 420840 vs 410100)")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("1. **Predisporre OdA-ledger per ogni contratto leasing** (1 contratto = 1 veicolo = 1 OdA con righe canone+bollo+interessi mensili)")
            lines.append("2. **Estendere line_groups** del writer per gestire 3+ keyword: 'Canone', 'Bollo', 'Quota interessi'")
            lines.append("3. **Routing**: il numero contratto/targa/codice cliente deve essere estratto dall'XML — verificare se compare in DatiContratto, RiferimentoAmministrazione, AltriDatiGestionali, o nelle descrizioni linee")
            lines.append("")
        elif fam == 'utility_energia':
            lines.append("### Pattern tipico utility energia")
            lines.append("")
            lines.append("- Fattura mensile/bimestrale per ogni POD")
            lines.append("- Più righe per fattura: corrispettivo energia + oneri + reattiva")
            lines.append("- POD nelle descrizioni linee (formato IT###L########)")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("Replicare lo schema **Sorgenia** (già implementato):")
            lines.append("- 1 OdA-ledger per POD con righe libere mensili")
            lines.append("- Routing tramite `pod_riferimenti` (già nel parser)")
            lines.append("- `description_strategy: keep_original`, ramo aggregato di `create_bozza_multilinea`")
            lines.append("")
        elif fam == 'factoring':
            lines.append("### Pattern tipico factoring")
            lines.append("")
            lines.append("- Commissioni periodiche, IVA esente (tax 54)")
            lines.append("- Importi variabili in base al volume ceduto")
            lines.append("- Conto contabile dedicato per fornitore (525020 SARDA, 525030 MBFACTA)")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("1. Predisporre OdA-ledger annuale con 12 righe libere mensili per ogni società di factoring")
            lines.append("2. Mapping diretto (no multi-contratto, ogni società ha il suo conto)")
            lines.append("3. **Verificare con contabilità**: il conto è di costo (525xxx) corretto, o serve conto deposito/anticipo?")
            lines.append("")
        elif fam == 'autonoleggio_breve':
            lines.append("### Pattern tipico autonoleggio breve termine")
            lines.append("")
            lines.append("- Fatture sporadiche (per noleggio specifico)")
            lines.append("- Conti vari (carburanti, noleggi, autostrade)")
            lines.append("- Gestione spesso a commessa interna")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("- Match implicito su importo + ricerca OdA per importo / data")
            lines.append("- Non automatizzabile come fornitore fisso (importi variabili)")
            lines.append("")
        elif fam == 'carburanti':
            lines.append("### Pattern tipico carburanti")
            lines.append("")
            lines.append("- Fatture periodiche per più mezzi/card")
            lines.append("- Conto stabile (410300/410410)")
            lines.append("- Possibile multi-card / multi-veicolo")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("Predisporre OdA-ledger con righe libere mensili. Per multi-card: una riga per card.")
            lines.append("")
        elif fam == 'hotel_ristorazione':
            lines.append("### Pattern tipico hotel/ristorazione")
            lines.append("")
            lines.append("- Fatture per soggiorni/pasti dipendenti in trasferta")
            lines.append("- Importi variabili")
            lines.append("- Conti deducibilità trasferta (420170/420171/420172)")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("Per ricorrenti (es. AN HOTEL SAN DONATO con 36 fatture): predisporre OdA-ledger da consumare. Per saltuari: manuale.")
            lines.append("")
        elif fam == 'retail_materiali':
            lines.append("### Pattern tipico retail materiali (esclusi nell'audit precedente)")
            lines.append("")
            lines.append("- Fatture multi-riga su materiali vari")
            lines.append("- OdA-DDT (1 OdA per consegna) → match implicito già funzionante per molti casi")
            lines.append("- Pattern eterogenei → automazione full non praticabile")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("Restano gestiti dal classifier esistente (AUTO_VALIDABILE / DA_VERIFICARE / MATCH_PARZIALE). Nessuna mappatura fornitore fisso.")
            lines.append("")
        elif fam == 'manutenzione':
            lines.append("### Pattern tipico manutenzione")
            lines.append("")
            lines.append("- Spesso 1 OdA per intervento → AUTO_VALIDABILE già funzionante")
            lines.append("- Conto stabile (420180)")
            lines.append("")
            lines.append("### Strategia proposta")
            lines.append("")
            lines.append("Lasciare al classifier. Non serve mappatura fornitore fisso.")
            lines.append("")

        lines.append("**Fornitori della famiglia:**")
        lines.append("")
        for r in sorted(items, key=lambda x: -(x['n_pending'] + x['n_posted'])):
            tot = r['n_pending'] + r['n_posted']
            lines.append(f"- [{r['category']}] **{r['supplier_name']}** ({r['vat']}) — {tot} fatt. ({r['n_pending']} pend)")
        lines.append("")

    out_path.write_text('\n'.join(lines), encoding='utf-8')


def write_codice(records, out_dir):
    """Raccomandazioni di modifiche al codice."""
    out_path = out_dir / 'raccomandazioni_codice.md'
    lines = []
    lines.append("# Raccomandazioni modifiche codice")
    lines.append("")
    lines.append("Le seguenti modifiche al codice sono suggerite dall'analisi pattern. NON implementate; richiedono approvazione utente.")
    lines.append("")

    # Analisi opportunità line_groups multi-conto
    multi_conto = [r for r in records
                   if r['category'] == 'C' and r['conti_top_pct'] < 80
                   and r['oda_pct'] >= 50]
    lines.append(f"## 1. Estendere `line_groups` con multi-conto ({len(multi_conto)} fornitori candidati)")
    lines.append("")
    lines.append("Oggi `line_groups` permette N keyword → N PO line, ma tutte vanno sullo stesso conto contabile (preso dal mapping).")
    lines.append("")
    lines.append("**Proposta**: aggiungere campo opzionale `account_id` per gruppo. Esempio:")
    lines.append("```python")
    lines.append("'line_groups': [")
    lines.append("    {'match': 'Canone', 'account_id': 420840, 'taxes_id': [11]},")
    lines.append("    {'match': 'Bollo', 'account_id': 410100, 'taxes_id': [47]},")
    lines.append("    {'match': 'Quota interessi', 'account_id': 524000, 'taxes_id': [51]},")
    lines.append("]")
    lines.append("```")
    lines.append("")
    lines.append("Top fornitori che ne beneficerebbero:")
    for r in sorted(multi_conto, key=lambda x: -(x['n_pending'] + x['n_posted']))[:10]:
        lines.append(f"- {r['supplier_name']} ({r['vat']}) — top conto {r['conti_top_pct']:.0f}%, famiglia {r['family']}")
    lines.append("")

    # Casi multi-IVA
    multi_iva = [r for r in records
                 if r['category'] == 'C' and r['iva_top_pct'] < 80
                 and r['oda_pct'] >= 50]
    lines.append(f"## 2. Multi-IVA per fattura ({len(multi_iva)} fornitori)")
    lines.append("")
    lines.append("Già supportato in line_groups (campo taxes_id), ma serve gestire pattern misti come:")
    lines.append("- Canone con IVA 22%")
    lines.append("- Bollo con tax 47 (esente)")
    lines.append("- Quota interessi con tax 51 (esente)")
    lines.append("")
    lines.append("**Proposta**: validare che `_match_lines_to_groups` (in `core/odoo_writer.py`) gestisca correttamente tassazioni miste senza errori.")
    lines.append("")

    # Bolli ricorrenti
    con_bolli = [r for r in records if r['has_bolli']]
    lines.append(f"## 3. Pattern bolli (€1, tax 47) — {len(con_bolli)} fornitori")
    lines.append("")
    lines.append("Pattern frequente: fatture utility/leasing hanno una riga 'Bollo' da €1-€2 con tax 47.")
    lines.append("")
    lines.append("**Proposta**: estendere `KEYWORD_RULES` (config/rules.py) con:")
    lines.append("```python")
    lines.append("('bollo', 'CONTO_BOLLI', 'BOLLO'),")
    lines.append("```")
    lines.append("e configurare `CONTO_BOLLI` con conto reale (es. 410100 o dedicato).")
    lines.append("")
    lines.append("Alternativa: nei `line_groups`, gruppo dedicato bollo con tax_id=47.")
    lines.append("")

    # Note di credito ricorrenti
    fornitori_nc = [r for r in records if r['n_pending_nc'] > 0 or r['n_posted_nc'] >= 2]
    lines.append(f"## 4. Note di credito ricorrenti ({len(fornitori_nc)} fornitori)")
    lines.append("")
    lines.append("Fornitori con NC ricorrenti (≥2 NC nei 10 mesi o NC in attesa). Verifica supporto TD04 nel writer.")
    lines.append("")
    for r in sorted(fornitori_nc, key=lambda x: -(x['n_pending_nc'] + x['n_posted_nc']))[:15]:
        lines.append(f"- {r['supplier_name']} ({r['vat']}) — {r['n_posted_nc']} NC posted, {r['n_pending_nc']} NC in attesa")
    lines.append("")

    # Estrazione campi XML extra
    lines.append("## 5. Campi XML extra utili (da analizzare per routing)")
    lines.append("")
    lines.append("Dall'analisi delle fatture in attesa, alcuni fornitori usano campi XML che oggi NON parsiamo:")
    lines.append("")
    cci_seen = defaultdict(set)
    for r in records:
        for p in []:  # placeholder; il dettaglio è in JSON
            pass
    lines.append("Campi candidati da aggiungere in `parse_fatturapa_xml`:")
    lines.append("- `<CodiceCommessaConvenzione>`")
    lines.append("- `<CodiceCIG>` / `<CodiceCUP>` (settore pubblico, per Ecotel forse non rilevante)")
    lines.append("- `<NumeroDDT>` (per match con DDT ricezione)")
    lines.append("")
    lines.append("Esempi presenti nelle fatture in attesa: vedi `audit_completo.json` campo `xml_extra`.")
    lines.append("")

    # Performance / scalabilità
    lines.append("## 6. Performance / scalabilità (suggerimenti generali)")
    lines.append("")
    lines.append("- **Cache OdA**: già presente in `_po_cache`. Potrebbe estendersi a `partner_id → posted_count` per evitare ricerche ripetute durante run.")
    lines.append("- **Bulk fetch invoice_line_ids**: nel writer, oggi si legge una riga PO per volta. Si può fare batch su `purchase.order.line` per le righe libere.")
    lines.append("- **Webapp**: oggi salva analisi una alla volta in dashboard.db. Per run ≥200 attachment l'INSERT singolo è lento; passare a INSERT bulk.")
    lines.append("")

    # UX webapp
    lines.append("## 7. UX webapp")
    lines.append("")
    lines.append("- Filtro per famiglia di fornitore (oggi solo per categoria)")
    lines.append("- Pulsante 'Crea bozze' bulk per famiglia (es. tutti i leasing)")
    lines.append("- Anteprima bozza prima della scrittura (mostra le righe che andranno create)")
    lines.append("- Indicatore stato OdA (libere disponibili / esaurite)")
    lines.append("")

    out_path.write_text('\n'.join(lines), encoding='utf-8')


# ============================================================
# MAIN
# ============================================================

def main():
    root = Path(__file__).parent
    timestamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    out_dir = root / 'output' / f'audit_completo_{timestamp}'
    out_dir.mkdir(parents=True, exist_ok=True)
    log = setup_logging(out_dir)
    log.info("=== Audit completo Ecotel — solo lettura ===")
    log.info(f"Output: {out_dir}")

    load_dotenv(root / 'config' / 'credentials.env')
    client = OdooReadOnlyClient(
        os.environ['ODOO_URL'], os.environ['ODOO_DB'],
        os.environ['ODOO_USERNAME'], os.environ['ODOO_PASSWORD']
    )
    client.connect()

    pending = fetch_pending(client, log)
    log.info("Parsing XML...")
    by_vat = parse_pending(pending, log)
    log.info(f"Fornitori distinti in attesa: {len(by_vat)}")

    today = dt.date.today()
    cutoff = today - dt.timedelta(days=HISTORICAL_MONTHS * 30)
    cutoff_str = cutoff.strftime('%Y-%m-%d')
    log.info(f"Periodo storico da {cutoff_str}")

    records = []
    vats = list(by_vat.keys())
    for i, vat in enumerate(vats, 1):
        invs = by_vat[vat]
        first = invs[0]
        log.info(f"[{i}/{len(vats)}] {vat} — {first['supplier_name'][:40]}")
        try:
            rec = analyze_supplier(client, vat, first['supplier_name'],
                                    first['partner_id'], invs, cutoff_str, log)
            records.append(rec)
        except Exception as e:
            log.error(f"Errore analisi {vat}: {e}")
            import traceback
            log.error(traceback.format_exc())

    log.info(f"Analisi completata: {len(records)} fornitori")

    # Output
    write_excel(records, out_dir / 'audit_completo.xlsx')
    log.info(f"Excel: {out_dir / 'audit_completo.xlsx'}")

    write_md_summary(records, out_dir, cutoff_str, today.strftime('%Y-%m-%d'))
    log.info(f"Markdown sintesi: {out_dir / 'audit_completo.md'}")

    write_roadmap(records, out_dir)
    log.info(f"Roadmap: {out_dir / 'roadmap_automazione.md'}")

    write_codice(records, out_dir)
    log.info(f"Raccomandazioni codice: {out_dir / 'raccomandazioni_codice.md'}")

    # JSON dump
    json_path = out_dir / 'audit_completo.json'
    serializable = []
    for r in records:
        d = dict(r)
        d['oda_distinct'] = [list(t) for t in d['oda_distinct']]
        d['conti_distinct'] = [list(t) for t in d['conti_distinct']]
        d['iva_distinct'] = [[list(t[0]), t[1]] for t in d['iva_distinct']]
        d['descrizioni_keywords'] = [list(t) for t in d['descrizioni_keywords']]
        serializable.append(d)
    json_path.write_text(json.dumps(serializable, indent=2, default=str), encoding='utf-8')
    log.info(f"JSON: {json_path}")

    by_cat = Counter(r['category'] for r in records)
    log.info("\n=== Riepilogo finale ===")
    for c in ['A', 'B', 'C', 'D', 'MAPPED']:
        log.info(f"  {c}: {by_cat.get(c, 0)}")
    log.info("=== FINE ===")


if __name__ == '__main__':
    main()
