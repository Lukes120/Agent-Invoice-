"""
Audit fornitori Ecotel — analisi pattern di registrazione fatture passive
per identificare candidati a "fornitore fisso" (mappatura automatica).

SOLO LETTURA: usa OdooReadOnlyClient. Niente scritture su Odoo.

Output:
  output/audit_fornitori_<timestamp>/
    audit_fornitori.xlsx       # tabella per fornitore
    audit_fornitori.md         # sintesi + strategia
    audit_fornitori.log        # log dettagliato
"""

import os
import sys
import json
import base64
import logging
import datetime as dt
from pathlib import Path
from collections import Counter, defaultdict
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core.odoo_client import OdooReadOnlyClient
from core.fatturapa_parser import parse_fatturapa_xml

# ============================================================
# CONFIGURAZIONE
# ============================================================

COMPANY_ID = 1  # Ecotel Italia S.r.l. a socio unico
HISTORICAL_MONTHS = 6
SOGLIA_FATTURE_CANDIDATO = 4  # >= per cat A/B; sotto cat C/D

# P.IVA da escludere (già mappate o esplicitamente non automatizzabili)
EXCLUDED_VATS = {
    # Già mappate (MAPPATURA_FORNITORI_FISSI)
    'IT05403151003': 'Trenitalia (mappato)',
    'IT09247981005': 'Italo (mappato)',
    'IT00488410010': 'Telecom (mappato)',
    'IT13378520152': 'Wind Tre (mappato)',
    # Commerciali / retail (esclusi a richiesta utente)
    'IT01634070435': 'RemaTarlazzi (retail)',
    'IT01257400992': 'Wuerth Elektronik (retail)',
    'IT00125230219': 'Wuerth (retail)',
    'IT00825330285': 'Sonepar (retail)',
    'IT02931690966': 'Rexel (retail)',
    'IT00763300480': 'MEF (retail)',
    'IT07095920638': 'Com-Cavi Abruzzo (retail)',
}


# ============================================================
# LOGGING
# ============================================================

def setup_logging(out_dir: Path) -> logging.Logger:
    out_dir.mkdir(parents=True, exist_ok=True)
    log_file = out_dir / 'audit_fornitori.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout),
        ]
    )
    return logging.getLogger('audit')


# ============================================================
# RACCOLTA DATI
# ============================================================

def fetch_pending_invoices(client, log):
    """Recupera fatture fatturapa.attachment.in non registrate Ecotel."""
    domain = [
        ('registered', '=', False),
        ('is_self_invoice', '=', False),
        ('company_id', '=', COMPANY_ID),
    ]
    atts = client._call('fatturapa.attachment.in', 'search_read', domain,
        fields=['id', 'name', 'xml_supplier_id', 'datas',
                'invoices_total', 'invoices_date'])
    log.info(f"Trovate {len(atts)} fatturapa non registrate Ecotel")
    return atts


def parse_pending_to_supplier_map(atts, log):
    """Per ogni fattura in attesa, parsa l'XML ed estrae P.IVA + dati base."""
    by_vat = defaultdict(list)
    for i, att in enumerate(atts, 1):
        try:
            raw = base64.b64decode(att['datas']).decode('utf-8', errors='replace')
            xml_data = parse_fatturapa_xml(raw)
            vat = (xml_data.cedente_partita_iva or '').strip().upper()
            if not vat:
                continue
            supplier_name = xml_data.cedente_denominazione or ''
            sup = att.get('xml_supplier_id')
            if not supplier_name and isinstance(sup, list) and len(sup) > 1:
                supplier_name = sup[1]
            partner_id = sup[0] if isinstance(sup, list) else None

            by_vat[vat].append({
                'attachment_id': att['id'],
                'attachment_name': att.get('name', ''),
                'supplier_name': supplier_name,
                'partner_id': partner_id,
                'numero': xml_data.numero,
                'data': xml_data.data,
                'tipo_doc': xml_data.tipo_documento,
                'imponibile_totale': xml_data.imponibile_totale,
                'importo_totale': xml_data.importo_totale or att.get('invoices_total', 0),
                'n_righe_xml': len(xml_data.righe),
                'oda_riferimenti': list(xml_data.oda_riferimenti),
                'commessa_riferimenti': list(xml_data.commessa_riferimenti),
                'contratto_riferimenti': list(xml_data.contratto_riferimenti),
                'rif_amministrazione': xml_data.cedente_riferimento_amministrazione,
            })
        except Exception as e:
            log.warning(f"Errore parsing att {att.get('id')}: {e}")
        if i % 25 == 0:
            log.info(f"  parsed {i}/{len(atts)}...")
    return by_vat


def fetch_supplier_history(client, partner_id, cutoff_date, log):
    """Storico 6 mesi: account.move posted del fornitore."""
    if not partner_id:
        return []
    domain = [
        ('partner_id', '=', partner_id),
        ('move_type', 'in', ['in_invoice', 'in_refund']),
        ('state', '=', 'posted'),
        ('invoice_date', '>=', cutoff_date),
        ('company_id', '=', COMPANY_ID),
    ]
    moves = client._call('account.move', 'search_read', domain,
        fields=['id', 'name', 'ref', 'invoice_date', 'amount_total', 'amount_untaxed',
                'invoice_origin', 'invoice_line_ids', 'move_type'])
    return moves


def fetch_move_lines_summary(client, line_ids):
    """Recupera righe move e ne fa il sommario per conto+IVA+prodotto."""
    if not line_ids:
        return []
    return client._call('account.move.line', 'read', line_ids,
        fields=['id', 'name', 'quantity', 'price_unit', 'price_subtotal',
                'account_id', 'tax_ids', 'product_id', 'purchase_line_id'])


def fetch_supplier_open_pos(client, partner_id):
    """OdA aperti del fornitore (state='purchase')."""
    if not partner_id:
        return []
    domain = [('partner_id', '=', partner_id), ('state', '=', 'purchase')]
    pos = client._call('purchase.order', 'search_read', domain,
        fields=['id', 'name', 'partner_ref', 'amount_total', 'amount_untaxed',
                'date_order', 'order_line', 'date_planned'])
    return pos


def count_libere(client, po_line_ids):
    """Conta le righe 'libere' di un OdA (qty_inv=0, qty_rec=0, product_qty>=1)."""
    if not po_line_ids:
        return 0, 0
    lines = client._call('purchase.order.line', 'read', po_line_ids,
        fields=['id', 'name', 'product_qty', 'qty_invoiced', 'qty_received'])
    libere = [l for l in lines
              if (l.get('qty_invoiced') or 0) == 0
              and (l.get('qty_received') or 0) == 0
              and (l.get('product_qty') or 0) >= 1]
    return len(libere), len(lines)


# ============================================================
# ANALISI / CATEGORIZZAZIONE
# ============================================================

def analyze_supplier(client, vat, supplier_name, partner_id,
                     pending_invoices, cutoff_date, log):
    """Analizza un fornitore: ritorna dict con metriche e categoria."""
    record = {
        'vat': vat,
        'supplier_name': supplier_name,
        'partner_id': partner_id,
        'n_pending': len(pending_invoices),
        'pending_total': sum((p['importo_totale'] or 0) for p in pending_invoices),
        'n_posted_6m': 0,
        'posted_total_6m': 0.0,
        'oda_usage_pct': 0.0,
        'oda_distinct': [],
        'conti_distinct': [],
        'iva_distinct': [],
        'has_commesse': False,
        'multi_contratto_hint': False,
        'open_pos': [],
        'open_pos_with_libere': 0,
        'category': 'D',
        'category_reason': '',
        'mapping_proposal': '',
    }

    if not partner_id:
        record['category_reason'] = 'partner_id mancante'
        return record

    # Storico 6 mesi
    moves = fetch_supplier_history(client, partner_id, cutoff_date, log)
    record['n_posted_6m'] = len(moves)
    record['posted_total_6m'] = sum((m.get('amount_total') or 0) for m in moves)

    # Aggregazione righe move
    all_line_ids = []
    for m in moves:
        all_line_ids.extend(m.get('invoice_line_ids', []) or [])
    move_lines = fetch_move_lines_summary(client, all_line_ids) if all_line_ids else []

    # OdA via invoice_origin
    n_with_origin = sum(1 for m in moves if m.get('invoice_origin'))
    if moves:
        record['oda_usage_pct'] = n_with_origin / len(moves) * 100
    oda_counter = Counter()
    for m in moves:
        origin = m.get('invoice_origin')
        if origin:
            oda_counter[origin.strip()] += 1
    record['oda_distinct'] = oda_counter.most_common(10)

    # Conti contabili (top 5 per N righe)
    conti_counter = Counter()
    iva_counter = Counter()
    for ln in move_lines:
        acc = ln.get('account_id')
        if acc and isinstance(acc, list):
            conti_counter[acc[1]] += 1
        tax_ids = ln.get('tax_ids') or []
        if tax_ids:
            iva_counter[tuple(sorted(tax_ids))] += 1
    record['conti_distinct'] = conti_counter.most_common(5)
    record['iva_distinct'] = iva_counter.most_common(5)

    # Commesse rilevate (in attesa o storico)
    has_commesse_pending = any(p.get('commessa_riferimenti') for p in pending_invoices)
    record['has_commesse'] = has_commesse_pending

    # Multi-contratto hint: se rif_amministrazione o contratto_riferimenti variano
    rif_amms = set()
    contratti = set()
    for p in pending_invoices:
        if p.get('rif_amministrazione'):
            rif_amms.add(p['rif_amministrazione'])
        for c in p.get('contratto_riferimenti', []):
            contratti.add(c)
    if len(rif_amms) > 1 or len(contratti) > 1:
        record['multi_contratto_hint'] = True

    # OdA aperti del fornitore
    open_pos = fetch_supplier_open_pos(client, partner_id)
    pos_summary = []
    n_with_libere = 0
    for po in open_pos:
        libere_n, total_n = count_libere(client, po.get('order_line', []))
        if libere_n > 0:
            n_with_libere += 1
        pos_summary.append({
            'name': po['name'],
            'partner_ref': po.get('partner_ref') or '',
            'amount_total': po.get('amount_total') or 0,
            'libere_n': libere_n,
            'total_lines': total_n,
        })
    record['open_pos'] = pos_summary
    record['open_pos_with_libere'] = n_with_libere

    # === Categorizzazione ===
    n_total = record['n_pending'] + record['n_posted_6m']
    cat, reason, proposal = categorize(record, n_total)
    record['category'] = cat
    record['category_reason'] = reason
    record['mapping_proposal'] = proposal

    return record


def categorize(rec, n_total):
    """Decide cat A/B/C/D + motivo + proposta."""
    if rec['has_commesse']:
        return 'D', 'Commesse S##### → Fase 2', \
               'Skip per ora; Fase 2 (matching commesse)'
    if n_total < SOGLIA_FATTURE_CANDIDATO:
        return 'D', f'Solo {n_total} fatture in 6 mesi → non vale automatizzare', \
               'Skip; registrazione manuale'

    # Conti / IVA stabili?
    n_lines = sum(c[1] for c in rec['conti_distinct'])
    top_conto_pct = (rec['conti_distinct'][0][1] / n_lines * 100) if n_lines else 0
    n_lines_iva = sum(c[1] for c in rec['iva_distinct'])
    top_iva_pct = (rec['iva_distinct'][0][1] / n_lines_iva * 100) if n_lines_iva else 0

    conti_stabili = top_conto_pct >= 80
    iva_stabile = top_iva_pct >= 80

    # OdA usage
    has_oda = rec['oda_usage_pct'] >= 80

    # OdA ledger predisposti? Almeno 1 OdA aperto con righe libere
    has_ledger = rec['open_pos_with_libere'] >= 1

    # Multi-contratto?
    multi = rec['multi_contratto_hint']

    if has_oda and conti_stabili and iva_stabile and has_ledger:
        cat = 'A'
        reason = (f"{n_total} fatt., {rec['oda_usage_pct']:.0f}% con OdA, "
                  f"top conto {top_conto_pct:.0f}%, top IVA {top_iva_pct:.0f}%, "
                  f"{rec['open_pos_with_libere']} OdA con righe libere")
        proposal = build_proposal(rec, multi)
        return cat, reason, proposal

    if has_oda and conti_stabili and iva_stabile and not has_ledger:
        cat = 'B'
        reason = (f"{n_total} fatt. stabili (conto {top_conto_pct:.0f}%, "
                  f"IVA {top_iva_pct:.0f}%) ma OdA non ledger → predisporre")
        proposal = build_proposal(rec, multi, needs_oda_setup=True)
        return cat, reason, proposal

    if not has_oda or not conti_stabili or not iva_stabile:
        # complicato ma frequente
        cat = 'C'
        details = []
        if not has_oda:
            details.append(f'solo {rec["oda_usage_pct"]:.0f}% con OdA')
        if not conti_stabili:
            details.append(f'conti dispersi (top {top_conto_pct:.0f}%)')
        if not iva_stabile:
            details.append(f'IVA disperse (top {top_iva_pct:.0f}%)')
        reason = f"{n_total} fatt., complicazioni: {'; '.join(details)}"
        proposal = "Caso per caso; richiede analisi specifica"
        return cat, reason, proposal

    # Fallback (non dovrebbe mai capitare)
    return 'D', 'Pattern non riconosciuto', 'Skip'


def build_proposal(rec, multi, needs_oda_setup=False):
    parts = []
    top_conto = rec['conti_distinct'][0][0] if rec['conti_distinct'] else '?'
    top_iva = rec['iva_distinct'][0][0] if rec['iva_distinct'] else '?'
    if needs_oda_setup:
        parts.append("[1] Predisporre OdA-ledger ricorrente con righe libere")
    parts.append(f"[2] Mappare: conto={top_conto}, taxes={list(top_iva)}, "
                 f"libere_criterio=standard_qty_inv_rec")
    if multi:
        parts.append("[3] Multi-contratto: serve identificare il campo XML di routing "
                     "(es. RiferimentoAmministrazione, DatiContratto, IdDocumento)")
    if rec['open_pos']:
        oda_names = ", ".join(p['name'] for p in rec['open_pos'][:5])
        parts.append(f"OdA aperti attuali: {oda_names}")
    return " | ".join(parts)


# ============================================================
# OUTPUT
# ============================================================

def write_excel(records, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Fornitori"

    headers = [
        'Categoria', 'Fornitore', 'P.IVA', 'partner_id',
        'N attesa', 'Tot attesa €', 'N posted 6m', 'Tot posted 6m €',
        'Tot fatture 6m', '% con OdA', 'OdA distinti',
        'Top conto', '% top conto', 'Top IVA',
        'OdA aperti', 'OdA con righe libere',
        'Multi-contratto?', 'Commesse?',
        'Motivo', 'Proposta'
    ]
    ws.append(headers)
    for c, _h in enumerate(headers, 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='DDDDDD')

    cat_color = {'A': 'C6EFCE', 'B': 'FFEB9C', 'C': 'FFCC99', 'D': 'F2F2F2'}

    # Ordina: A, B, C, D, e dentro per N totale fatture
    sorted_recs = sorted(records, key=lambda r: (
        ['A', 'B', 'C', 'D'].index(r['category']),
        -(r['n_pending'] + r['n_posted_6m'])
    ))

    for r in sorted_recs:
        n_total = r['n_pending'] + r['n_posted_6m']
        n_lines = sum(c[1] for c in r['conti_distinct'])
        top_conto = r['conti_distinct'][0][0] if r['conti_distinct'] else ''
        top_conto_pct = (r['conti_distinct'][0][1] / n_lines * 100) if n_lines else 0
        n_lines_iva = sum(c[1] for c in r['iva_distinct'])
        top_iva = list(r['iva_distinct'][0][0]) if r['iva_distinct'] else []
        oda_distinct_str = ", ".join(f"{n}({c}x)" for n, c in r['oda_distinct'][:5])
        open_pos_str = ", ".join(p['name'] for p in r['open_pos'][:5])

        row = [
            r['category'], r['supplier_name'], r['vat'], r['partner_id'],
            r['n_pending'], round(r['pending_total'], 2),
            r['n_posted_6m'], round(r['posted_total_6m'], 2),
            n_total, round(r['oda_usage_pct'], 1),
            oda_distinct_str,
            top_conto, round(top_conto_pct, 1),
            str(top_iva),
            open_pos_str, r['open_pos_with_libere'],
            'sì' if r['multi_contratto_hint'] else 'no',
            'sì' if r['has_commesse'] else 'no',
            r['category_reason'], r['mapping_proposal']
        ]
        ws.append(row)
        last_row = ws.max_row
        fill = PatternFill('solid', fgColor=cat_color[r['category']])
        ws.cell(row=last_row, column=1).fill = fill

    # Larghezza colonne
    widths = [8, 35, 16, 10, 8, 12, 8, 12, 8, 8, 30, 12, 8, 12, 25, 8, 8, 8, 50, 60]
    for i, w in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

    wb.save(out_path)


def write_markdown(records, out_path, period_start, period_end):
    lines = []
    lines.append(f"# Audit fornitori Ecotel — {dt.date.today()}")
    lines.append("")
    lines.append(f"**Periodo storico**: {period_start} → {period_end} (6 mesi)")
    lines.append(f"**Ambito**: company Ecotel Italia (id 1), solo fatture/note credito.")
    lines.append("")

    # Statistiche generali
    n = len(records)
    by_cat = Counter(r['category'] for r in records)
    pending_total_count = sum(r['n_pending'] for r in records)
    pending_total_amount = sum(r['pending_total'] for r in records)
    lines.append("## Riepilogo")
    lines.append("")
    lines.append(f"- Fornitori analizzati: **{n}**")
    lines.append(f"- Fatture in attesa coperte: **{pending_total_count}** (€{pending_total_amount:,.2f})")
    lines.append(f"- Cat **A** (subito automatizzabili): **{by_cat.get('A', 0)}**")
    lines.append(f"- Cat **B** (predisporre OdA ledger): **{by_cat.get('B', 0)}**")
    lines.append(f"- Cat **C** (complessi): **{by_cat.get('C', 0)}**")
    lines.append(f"- Cat **D** (skip): **{by_cat.get('D', 0)}**")
    lines.append("")

    # Fornitori cat A
    a_recs = [r for r in records if r['category'] == 'A']
    lines.append(f"## Categoria A — Subito automatizzabili ({len(a_recs)})")
    lines.append("")
    lines.append("OdA già ledger predisposti, conto/IVA stabili, ≥80% fatture con OdA esplicito. Bastano 30 minuti per aggiungerli a `MAPPATURA_FORNITORI_FISSI`.")
    lines.append("")
    if a_recs:
        for r in sorted(a_recs, key=lambda x: -(x['n_pending'] + x['n_posted_6m'])):
            n_total = r['n_pending'] + r['n_posted_6m']
            top_conto = r['conti_distinct'][0][0] if r['conti_distinct'] else '?'
            top_iva = list(r['iva_distinct'][0][0]) if r['iva_distinct'] else []
            oda_str = ", ".join(n for n, _ in r['oda_distinct'][:5])
            lines.append(f"### {r['supplier_name']} ({r['vat']})")
            lines.append(f"- Fatture: {r['n_pending']} attesa + {r['n_posted_6m']} posted = **{n_total}** in 6 mesi")
            lines.append(f"- OdA: {oda_str}")
            lines.append(f"- Conto {top_conto}, IVA {top_iva}")
            lines.append(f"- Multi-contratto: {'sì' if r['multi_contratto_hint'] else 'no'}")
            lines.append(f"- **Proposta**: {r['mapping_proposal']}")
            lines.append("")
    else:
        lines.append("_Nessun fornitore in questa categoria._")
        lines.append("")

    # Fornitori cat B
    b_recs = [r for r in records if r['category'] == 'B']
    lines.append(f"## Categoria B — OdA da predisporre ({len(b_recs)})")
    lines.append("")
    lines.append("Pattern stabile (conto/IVA), ma OdA usati attualmente sono \"spot\" (singolo OdA per singola fattura, senza righe predisposte). **Richiede setup contabilità**: predisporre per ogni fornitore un OdA-ledger con righe libere, poi mappare.")
    lines.append("")
    if b_recs:
        for r in sorted(b_recs, key=lambda x: -(x['n_pending'] + x['n_posted_6m'])):
            n_total = r['n_pending'] + r['n_posted_6m']
            top_conto = r['conti_distinct'][0][0] if r['conti_distinct'] else '?'
            top_iva = list(r['iva_distinct'][0][0]) if r['iva_distinct'] else []
            lines.append(f"### {r['supplier_name']} ({r['vat']})")
            lines.append(f"- Fatture: {r['n_pending']} attesa + {r['n_posted_6m']} posted = **{n_total}** in 6 mesi")
            lines.append(f"- Conto stabile: {top_conto}, IVA stabile: {top_iva}")
            lines.append(f"- Multi-contratto: {'sì' if r['multi_contratto_hint'] else 'no'}")
            lines.append(f"- **Proposta**: {r['mapping_proposal']}")
            lines.append("")
    else:
        lines.append("_Nessun fornitore in questa categoria._")
        lines.append("")

    # Fornitori cat C
    c_recs = [r for r in records if r['category'] == 'C']
    lines.append(f"## Categoria C — Complessi ({len(c_recs)})")
    lines.append("")
    lines.append("Hanno volume sufficiente ma con complicazioni (multi-conto, multi-IVA, % bassa con OdA). Da valutare uno per uno; potrebbero richiedere logica custom.")
    lines.append("")
    if c_recs:
        for r in sorted(c_recs, key=lambda x: -(x['n_pending'] + x['n_posted_6m'])):
            n_total = r['n_pending'] + r['n_posted_6m']
            lines.append(f"- **{r['supplier_name']}** ({r['vat']}) — {n_total} fatt. — {r['category_reason']}")
        lines.append("")
    else:
        lines.append("_Nessun fornitore in questa categoria._")
        lines.append("")

    # Cat D solo conteggio
    d_recs = [r for r in records if r['category'] == 'D']
    lines.append(f"## Categoria D — Da non automatizzare ({len(d_recs)})")
    lines.append("")
    lines.append(f"{len(d_recs)} fornitori con <{SOGLIA_FATTURE_CANDIDATO} fatture in 6 mesi o commesse Fase 2. Lista completa in Excel.")
    lines.append("")

    # ===== STRATEGIA =====
    lines.append("## Strategia di automazione proposta")
    lines.append("")
    lines.append("### Fase 1 — Quick wins (Cat A)")
    lines.append("Aggiungere subito a `MAPPATURA_FORNITORI_FISSI` i fornitori in cat A. Sono già pronti per partire. ETA: ~1 giorno di sviluppo + test.")
    lines.append("")
    lines.append("### Fase 2 — Setup contabilità (Cat B)")
    lines.append("Predisporre OdA-ledger ricorrenti per i fornitori in cat B. Servirà:")
    lines.append("1. Decidere il \"contenitore annuale\" (un OdA da €X con righe libere mese per mese)")
    lines.append("2. Crearlo manualmente per il primo fornitore di test")
    lines.append("3. Mapparlo")
    lines.append("4. Replicare per gli altri fornitori cat B")
    lines.append("")
    lines.append("### Fase 3 — Casi complessi (Cat C)")
    lines.append("Analisi caso per caso. Tipici pattern: fornitori con righe a conti diversi (es. servizi vs materiali), fornitori con multi-contratto non ovvio, fornitori con cumulativi.")
    lines.append("")
    lines.append("### Fase 4 — Coda lunga (Cat D)")
    lines.append("Per fornitori con 1-3 fatture/anno (la maggioranza dei 116 totali) NON ha senso investire in automazione. Resta registrazione manuale via UI Odoo.")
    lines.append("")

    # ROI stimato
    a_pending_count = sum(r['n_pending'] for r in a_recs)
    b_pending_count = sum(r['n_pending'] for r in b_recs)
    c_pending_count = sum(r['n_pending'] for r in c_recs)
    d_pending_count = sum(r['n_pending'] for r in d_recs)
    total_pending = a_pending_count + b_pending_count + c_pending_count + d_pending_count

    lines.append("### Stima impatto (su fatture in attesa attuali)")
    lines.append("")
    lines.append(f"| Categoria | Fornitori | Fatture coperte | % sul totale |")
    lines.append("| --- | ---: | ---: | ---: |")
    if total_pending > 0:
        lines.append(f"| A (quick win) | {len(a_recs)} | {a_pending_count} | {a_pending_count/total_pending*100:.1f}% |")
        lines.append(f"| B (predisporre OdA) | {len(b_recs)} | {b_pending_count} | {b_pending_count/total_pending*100:.1f}% |")
        lines.append(f"| C (complessi) | {len(c_recs)} | {c_pending_count} | {c_pending_count/total_pending*100:.1f}% |")
        lines.append(f"| D (manuali) | {len(d_recs)} | {d_pending_count} | {d_pending_count/total_pending*100:.1f}% |")
    lines.append("")
    lines.append("Le fatture mappate da Trenitalia/Italo/Telecom/Wind Tre sono già coperte e non rientrano in questi numeri.")
    lines.append("")

    out_path.write_text('\n'.join(lines), encoding='utf-8')


# ============================================================
# MAIN
# ============================================================

def main():
    root = Path(__file__).parent
    timestamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    out_dir = root / 'output' / f'audit_fornitori_{timestamp}'
    out_dir.mkdir(parents=True, exist_ok=True)
    log = setup_logging(out_dir)

    log.info("=== Audit fornitori Ecotel ===")
    log.info(f"Output: {out_dir}")

    load_dotenv(root / 'config' / 'credentials.env')
    client = OdooReadOnlyClient(
        os.environ['ODOO_URL'], os.environ['ODOO_DB'],
        os.environ['ODOO_USERNAME'], os.environ['ODOO_PASSWORD']
    )
    client.connect()

    # 1. Fatture in attesa
    pending = fetch_pending_invoices(client, log)

    # 2. Parsing XML → mappa per P.IVA
    log.info("Parsing XML fatture in attesa...")
    by_vat = parse_pending_to_supplier_map(pending, log)

    # 3. Filtra esclusi
    excluded = {v: by_vat.pop(v) for v in list(by_vat) if v in EXCLUDED_VATS}
    log.info(f"Fornitori in attesa: {len(by_vat)} (esclusi {len(excluded)} mappati/retail)")

    # 4. Periodo storico
    today = dt.date.today()
    cutoff = today - dt.timedelta(days=HISTORICAL_MONTHS * 30)
    cutoff_str = cutoff.strftime('%Y-%m-%d')
    log.info(f"Storico da: {cutoff_str}")

    # 5. Analisi per fornitore
    records = []
    vats = list(by_vat.keys())
    for i, vat in enumerate(vats, 1):
        invoices = by_vat[vat]
        # Prendo nome dal primo + partner_id
        first = invoices[0]
        log.info(f"[{i}/{len(vats)}] {vat} — {first['supplier_name'][:40]} "
                 f"({len(invoices)} attesa)")
        try:
            rec = analyze_supplier(
                client, vat, first['supplier_name'], first['partner_id'],
                invoices, cutoff_str, log
            )
            records.append(rec)
        except Exception as e:
            log.error(f"Errore analisi {vat}: {e}")
            import traceback
            log.error(traceback.format_exc())

    log.info(f"\n=== Analisi completata: {len(records)} fornitori ===")

    # 6. Output
    excel_path = out_dir / 'audit_fornitori.xlsx'
    write_excel(records, excel_path)
    log.info(f"Excel: {excel_path}")

    md_path = out_dir / 'audit_fornitori.md'
    write_markdown(records, md_path, cutoff_str, today.strftime('%Y-%m-%d'))
    log.info(f"Markdown: {md_path}")

    # JSON dump per debug/uso futuro
    json_path = out_dir / 'audit_fornitori.json'
    serializable = []
    for r in records:
        d = dict(r)
        d['oda_distinct'] = [list(t) for t in d['oda_distinct']]
        d['conti_distinct'] = [list(t) for t in d['conti_distinct']]
        d['iva_distinct'] = [[list(t[0]), t[1]] for t in d['iva_distinct']]
        serializable.append(d)
    json_path.write_text(json.dumps(serializable, indent=2, default=str), encoding='utf-8')
    log.info(f"JSON: {json_path}")

    # Riepilogo finale
    by_cat = Counter(r['category'] for r in records)
    log.info("\n=== Riepilogo per categoria ===")
    for cat in ['A', 'B', 'C', 'D']:
        log.info(f"  {cat}: {by_cat.get(cat, 0)}")

    log.info("=== FINE ===")


if __name__ == '__main__':
    main()
