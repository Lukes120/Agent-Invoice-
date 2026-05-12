"""
Genera un Excel con i 19 apparati che compaiono nelle fatture Autostrade
Q1 2026 ma NON sono nel file PARCO AUTO 2026.
Il file ha le colonne pre-popolate (codice, tipo, dove appare, importo) e
le colonne da compilare (targa, veicolo, classificazione).
Da inviare a chi gestisce il parco auto / Acquisti.
"""
import os
import sys
import re
import glob
from collections import defaultdict
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from core.pdf_parser import parse_pdf_autostrade
from config.apparati_mapping import get_classificazione

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT_PDF = r"C:\Users\lranalletta\Downloads\cHECK\_extracted"

# Pattern per estrarre cc dal path: "1 - Contratto 261713569"
RE_CC = re.compile(r'Contratto\s+(\d+)')
RE_MESE = re.compile(r'(\d{2})\s*-\s*Fatture\s+(\w+)\s+(\d{4})')


def parse_path_meta(rel_path):
    cc = RE_CC.search(rel_path)
    mese = RE_MESE.search(rel_path)
    return {
        'cc': cc.group(1) if cc else '',
        'mese_n': mese.group(1) if mese else '',
        'mese_nome': mese.group(2) if mese else '',
        'anno': mese.group(3) if mese else '',
    }


def main():
    pdfs = sorted(glob.glob(os.path.join(ROOT_PDF, '**', 'Autostrade*.pdf'),
                            recursive=True))
    print(f'PDF Autostrade trovati: {len(pdfs)}')

    # struttura: missing[(tipo, id)] = {
    #   'tot_eur', 'count_pdf',
    #   'months': {YYYY-MM: tot_eur},
    #   'clienti': set(cc),
    #   'fatture': list of (numero_fattura, mese, eur),
    # }
    missing = defaultdict(lambda: {
        'tot_eur': 0.0,
        'count_pdf': 0,
        'months': defaultdict(float),
        'clienti': set(),
        'fatture': [],
        'count_movimenti': 0,
    })

    for pdf in pdfs:
        rel = pdf.replace(ROOT_PDF + os.sep, '')
        meta = parse_path_meta(rel)
        ft_match = re.search(r'(\d{12,})', os.path.basename(pdf))
        ft_num = ft_match.group(1) if ft_match else os.path.basename(pdf)
        data = parse_pdf_autostrade(pdf)
        for a in data.apparati:
            if get_classificazione(a.apparato_id):
                continue
            key = (a.tipo, a.apparato_id)
            d = missing[key]
            d['tot_eur'] += a.importo_iva_inclusa
            d['count_pdf'] += 1
            d['count_movimenti'] += a.n_movimenti
            d['clienti'].add(meta['cc'])
            month_key = f'{meta["mese_nome"]} {meta["anno"]}'
            d['months'][month_key] += a.importo_iva_inclusa
            d['fatture'].append((ft_num, month_key, a.importo_iva_inclusa))

    print(f'Apparati unici mancanti: {len(missing)}')

    # ----------------------- Excel -----------------------
    wb = Workbook()
    ws = wb.active
    ws.title = 'Apparati da censire'

    title = Font(bold=True, color='FFFFFF', size=11)
    fill_hdr = PatternFill('solid', fgColor='305496')
    fill_todo = PatternFill('solid', fgColor='FFF2CC')  # giallo per "da compilare"
    fill_alt = PatternFill('solid', fgColor='F2F2F2')
    thin = Side(border_style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        ('Tipo apparato', 14),
        ('Codice apparato (da PDF)', 24),
        ('Codice cliente (cc)', 22),
        ('# fatture in cui appare', 22),
        ('# movimenti totali Q1', 22),
        ('€ totale Q1 2026 (IVA incl.)', 28),
        ('Gennaio 2026 €', 16),
        ('Febbraio 2026 €', 16),
        ('Marzo 2026 €', 16),
        ('TARGA (DA COMPILARE)', 26),
        ('VEICOLO descrizione (DA COMPILARE)', 38),
        ('CLASSIFICAZIONE (DA COMPILARE)', 32),
        ('Note', 30),
    ]

    for i, (h, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = title
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 38

    # Data validation sulla colonna classificazione
    dv = DataValidation(
        type='list',
        formula1='"furgoni,uso_promiscuo,non assegnata / pool"',
        allow_blank=True,
    )
    dv.error = 'Valori ammessi: furgoni / uso_promiscuo / non assegnata / pool'
    dv.errorTitle = 'Valore non valido'
    dv.prompt = 'Scegli: furgoni (100%) o uso_promiscuo (70%) o non assegnata / pool'
    dv.promptTitle = 'Classificazione'
    ws.add_data_validation(dv)

    items = sorted(missing.items(), key=lambda kv: -kv[1]['tot_eur'])
    row_idx = 2
    for (tipo, aid), info in items:
        cc_list = ', '.join(sorted(info['clienti']))
        # mesi
        gen = round(info['months'].get('Gennaio 2026', 0.0), 2)
        feb = round(info['months'].get('Febbraio 2026', 0.0), 2)
        mar = round(info['months'].get('Marzo 2026', 0.0), 2)

        cells = [
            tipo,
            aid,
            cc_list,
            info['count_pdf'],
            info['count_movimenti'],
            round(info['tot_eur'], 2),
            gen if gen else '',
            feb if feb else '',
            mar if mar else '',
            '',  # targa
            '',  # veicolo
            '',  # classificazione
            '',  # note
        ]
        for col_idx, val in enumerate(cells, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border
            if col_idx in (10, 11, 12):
                c.fill = fill_todo
            elif row_idx % 2 == 0:
                c.fill = fill_alt
            if col_idx in (4, 5, 6, 7, 8, 9):
                c.alignment = Alignment(horizontal='right')
        # data validation classificazione (col 12)
        dv.add(ws.cell(row=row_idx, column=12).coordinate)
        row_idx += 1

    # totale
    ws.cell(row=row_idx, column=1, value='TOTALE').font = Font(bold=True)
    ws.cell(row=row_idx, column=4,
            value=sum(i['count_pdf'] for _, i in items)).font = Font(bold=True)
    ws.cell(row=row_idx, column=5,
            value=sum(i['count_movimenti'] for _, i in items)).font = Font(bold=True)
    ws.cell(row=row_idx, column=6,
            value=round(sum(i['tot_eur'] for _, i in items), 2)).font = Font(bold=True)
    for c in ws[row_idx]:
        c.fill = PatternFill('solid', fgColor='D9E1F2')
        c.border = border

    # freeze
    ws.freeze_panes = 'A2'

    # ----------------------- Sheet 2: come compilare -----------------------
    ws2 = wb.create_sheet('Istruzioni')
    instructions = [
        ('Cosa è questo file', ''),
        ('', 'Lista dei 19 apparati TELEPASS/VIACARD che compaiono nelle '
              'fatture Autostrade Q1 2026 (gennaio-marzo) ma NON sono '
              'presenti nel file PARCO AUTO 2026 apparati dettaglio.xlsx.'),
        ('', 'Servono per completare la mappatura apparato → veicolo → '
              'classificazione (furgoni 100% deducibili / uso promiscuo 70%).'),
        ('', ''),
        ('Cosa vi chiediamo di compilare', ''),
        ('', 'Per ciascuna riga, le 3 colonne gialle:'),
        ('  - TARGA', 'la targa del veicolo a cui è assegnato l\'apparato'),
        ('  - VEICOLO', 'descrizione (modello, es. "FIAT PANDA 1.0 HYBRID")'),
        ('  - CLASSIFICAZIONE', 'una di:'),
        ('     furgoni', '= deducibilità 100% (mezzi commerciali / strumentali)'),
        ('     uso_promiscuo', '= deducibilità 70% (auto aziendali a uso misto)'),
        ('     non assegnata / pool', '= se non è una carta su veicolo specifico'),
        ('', ''),
        ('Importi e contesto', ''),
        ('', 'Le colonne € mostrano il fatturato Q1 2026 IVA inclusa di '
              'ciascun apparato. Servono solo a darvi il senso di urgenza '
              '(quanto pesa ognuno).'),
        ('', 'Codice cliente (cc) indica i contratti Autostrade in cui '
              'l\'apparato è risultato fatturato — utile per cercarlo nei '
              'vostri sistemi.'),
        ('', ''),
        ('Quando lo restituite', ''),
        ('', 'Una volta compilato, restituite il file e provvederemo a '
              'integrarlo nel file PARCO AUTO 2026 ufficiale + nella mappa '
              'dell\'agent contabile.'),
        ('', ''),
        (f'Generato il', datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]
    for i, (k, v) in enumerate(instructions, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=bool(k))
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 90

    # ----------------------- Sheet 3: dettaglio fatture per apparato -----------------------
    ws3 = wb.create_sheet('Dettaglio fatture')
    headers3 = ['Tipo', 'Codice apparato', 'Codice cliente', 'Mese', 'Numero fattura', '€ IVA incl.']
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = title
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal='center')
    r = 2
    for (tipo, aid), info in items:
        cc_list = ', '.join(sorted(info['clienti']))
        for ft_num, mese, eur in info['fatture']:
            ws3.cell(row=r, column=1, value=tipo)
            ws3.cell(row=r, column=2, value=aid)
            ws3.cell(row=r, column=3, value=cc_list)
            ws3.cell(row=r, column=4, value=mese)
            ws3.cell(row=r, column=5, value=ft_num)
            ws3.cell(row=r, column=6, value=round(eur, 2))
            r += 1
    for col_idx, w in enumerate([12, 22, 24, 18, 24, 14], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.freeze_panes = 'A2'

    # ----------------------- Save -----------------------
    out_dir = os.path.join(ROOT, 'output')
    os.makedirs(out_dir, exist_ok=True)
    out_xlsx = os.path.join(
        out_dir,
        f'apparati_da_censire_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(out_xlsx)
    print(f'\nExcel salvato in:\n  {out_xlsx}')
    print(f'\nApparati totali: {len(items)}')
    print(f'Importo Q1 2026 da chiarire: € {sum(i["tot_eur"] for _, i in items):.2f}')


if __name__ == '__main__':
    main()
