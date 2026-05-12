"""
Generatore idempotente di config/apparati_mapping.py partendo dall'Excel
"Apparati completo TOT.xlsx" mantenuto dall'ufficio parco/contabilità.

Sorgente attesa (foglio 'Totale' o 'Apparati'):

  | Stato | Fornitore(i) | Tipo apparato | Codice apparato (da PDF) |
    Codice cliente (cc) | # fatture | # movimenti | € totale | ... |
    TARGA | VEICOLO descrizione | CLASSIFICAZIONE | Note |

Stati ammessi:
  - MAPPATO   → apparato attivo, info complete (targa+veicolo specifici)
  - CENSITO   → apparato attivo, pool/info parziali
  - DISMESSO  → restituito/disattivato (tenuto in mappa per fatture storiche)
  - GUASTA    → guasto/non più operativo (tenuto in mappa per fatture storiche)

Le righe TOTALE / vuote / con stato non ammesso vengono saltate.

Esecuzione:
    python scripts/generate_apparati_mapping.py
    python scripts/generate_apparati_mapping.py --xlsx <path>
    python scripts/generate_apparati_mapping.py --output <path>
    python scripts/generate_apparati_mapping.py --print-summary

Lo script è READ-ONLY su filesystem (eccetto il file di output).
Non tocca Odoo, non tocca DB.

Output console:
  - riepilogo entries per stato/classificazione
  - diff vs versione precedente di config/apparati_mapping.py
    (apparati nuovi, rimossi, modificati)
  - warning su consistenza chiave (codice_apparato, cc_cliente)
"""
import argparse
import os
import re
import sys
import importlib.util
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERRORE: openpyxl non installato. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "input" / "Apparati completo TOT.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "config" / "apparati_mapping.py"

STATI_ATTIVI = {"MAPPATO", "CENSITO"}
STATI_STORICI = {"DISMESSO", "GUASTA"}
STATI_AMMESSI = STATI_ATTIVI | STATI_STORICI

TARGA_PATTERN = re.compile(r'\b([A-Z]{2}\d{3}[A-Z]{2})\b')
VIACARD_PATTERN = re.compile(r'\b(\d+\.\d{6}\.\d+)\b')


def normalize_codice_apparato(raw) -> str:
    """Normalizza un codice apparato.

    Excel può consegnarlo come int (per codici interamente numerici) o come
    string. Se contiene un pattern VIACARD puntato lo restituisco così com'è;
    altrimenti normalizzo a sole cifre.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    if VIACARD_PATTERN.search(s):
        return VIACARD_PATTERN.search(s).group(1)
    no_spaces = re.sub(r'\s+', '', s)
    digit_seqs = re.findall(r'\d{6,}', no_spaces)
    if digit_seqs:
        return max(digit_seqs, key=len)
    return ""


def normalize_classificazione(raw) -> str:
    s = str(raw or "").strip().lower().replace(" ", "_")
    if s in ("uso_promiscuo", "promiscuo", "uso_promiscui"):
        return "uso_promiscuo"
    if s in ("furgoni", "furgone"):
        return "furgoni"
    return s


def normalize_stato(raw) -> str:
    return str(raw or "").strip().upper()


def extract_targa(s) -> str:
    if not s:
        return ""
    matches = TARGA_PATTERN.findall(str(s).upper())
    return matches[-1] if matches else ""


def parse_xlsx(xlsx_path: Path):
    """Legge l'Excel e ritorna (entries_attivi+storici, skipped, anomalie)."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel non trovato: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Preferisco il foglio 'Totale' (più recente). Se non c'è, uso 'Apparati'.
    sheet_name = None
    for candidate in ("Totale", "Apparati"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        raise ValueError(
            f"Foglio 'Totale' o 'Apparati' non trovato in {xlsx_path}. "
            f"Disponibili: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Foglio '{sheet_name}' vuoto")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # Mappa colonne attese -> indice
    col_map = {}
    for i, name in enumerate(header):
        n = name.lower()
        if n.startswith("stato"):
            col_map["stato"] = i
        elif n.startswith("fornitore"):
            col_map["fornitore"] = i
        elif n.startswith("tipo"):
            col_map["tipo"] = i
        elif "codice apparato" in n:
            col_map["codice"] = i
        elif "codice cliente" in n or n == "cc":
            col_map["cc"] = i
        elif n == "targa":
            col_map["targa"] = i
        elif "veicolo" in n and "descrizione" in n:
            col_map["veicolo"] = i
        elif n.startswith("classificazione"):
            col_map["classificazione"] = i
        elif n.startswith("note"):
            col_map["note"] = i

    required = {"stato", "tipo", "codice", "cc", "classificazione"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Colonne mancanti nel foglio '{sheet_name}': {missing}. "
            f"Header letto: {header}")

    entries = []
    skipped = []
    anomalie = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        stato_raw = row[col_map["stato"]] if col_map["stato"] < len(row) else None
        stato = normalize_stato(stato_raw)
        if stato not in STATI_AMMESSI:
            # Riga TOTALE, riga vuota, riga di nota: skip silenzioso
            if stato and stato not in {"", "TOTALE"} and not stato.startswith("DI CUI"):
                skipped.append({
                    "row": row_idx, "stato": stato, "motivo": "stato_non_ammesso",
                })
            continue

        codice_raw = row[col_map["codice"]]
        codice = normalize_codice_apparato(codice_raw)
        if not codice:
            anomalie.append({
                "row": row_idx, "stato": stato, "motivo": "codice_vuoto",
                "codice_raw": str(codice_raw)[:40],
            })
            continue

        cc_raw = row[col_map["cc"]]
        cc = str(cc_raw).strip() if cc_raw is not None else ""

        tipo = str(row[col_map["tipo"]] or "").strip()
        targa = ""
        if "targa" in col_map and col_map["targa"] < len(row):
            targa = extract_targa(row[col_map["targa"]])
        veicolo = ""
        if "veicolo" in col_map and col_map["veicolo"] < len(row):
            veicolo = str(row[col_map["veicolo"]] or "").strip()
        classificazione = normalize_classificazione(row[col_map["classificazione"]])
        note = ""
        if "note" in col_map and col_map["note"] < len(row):
            note = str(row[col_map["note"]] or "").strip()
        fornitori = ""
        if "fornitore" in col_map and col_map["fornitore"] < len(row):
            fornitori = str(row[col_map["fornitore"]] or "").strip()

        entries.append({
            "row": row_idx,
            "stato": stato,
            "fornitori": fornitori,
            "tipo_apparato": tipo,
            "codice_apparato": codice,
            "cc_cliente": cc,
            "targa": targa,
            "veicolo_descrizione": veicolo,
            "classificazione": classificazione,
            "note": note,
        })

    return entries, skipped, anomalie, sheet_name


def build_apparati_map(entries):
    """De-duplica per codice_apparato. Se più (codice, cc) condividono lo
    stesso codice ma cc differenti, tengo la prima entry e annoto i cc
    aggiuntivi nel campo `cc_clienti_aggiuntivi`. Logga warning.

    Se più entry hanno lo stesso (codice, cc) → tengo l'ULTIMA (riassegnazione
    storica intra-cc).
    """
    by_id = {}
    duplicati_intra_cc = []
    multi_cc = defaultdict(list)  # codice -> list(cc) tra duplicati cross-cc
    incoerenze_cls = []

    for e in entries:
        cid = e["codice_apparato"]
        if cid in by_id:
            prev = by_id[cid]
            if prev["cc_cliente"] == e["cc_cliente"]:
                duplicati_intra_cc.append({"prev": prev, "curr": e})
                # tengo l'ULTIMA (override)
                by_id[cid] = e
            else:
                # cross-cc: tengo la prima ma annoto i cc aggiuntivi
                multi_cc[cid].append(e["cc_cliente"])
                if e["classificazione"] != prev["classificazione"]:
                    incoerenze_cls.append({"codice": cid,
                                            "prev_cls": prev["classificazione"],
                                            "curr_cls": e["classificazione"],
                                            "prev_cc": prev["cc_cliente"],
                                            "curr_cc": e["cc_cliente"]})
        else:
            by_id[cid] = e

    for cid, ccs in multi_cc.items():
        by_id[cid]["cc_clienti_aggiuntivi"] = sorted(set(ccs))

    return by_id, duplicati_intra_cc, multi_cc, incoerenze_cls


def build_aliases(by_id):
    """Genera alias per supportare codici 'corti' usati nei PDF Autostrade."""
    aliases = {}
    for aid in by_id.keys():
        if len(aid) == 12 and aid.isdigit() and aid[:3] in ("490", "420"):
            short = aid[3:]
            if short not in by_id and short not in aliases:
                aliases[short] = aid
        if "." in aid:
            no_dots = aid.replace(".", "")
            if no_dots not in by_id and no_dots not in aliases:
                aliases[no_dots] = aid
    return aliases


def render_module(by_id, aliases, xlsx_source: Path, sheet_name: str,
                   skipped_count: int, dup_intra_count: int,
                   multi_cc_count: int) -> str:
    sorted_ids = sorted(by_id.keys())
    counter = defaultdict(int)
    for e in by_id.values():
        counter[(e.get("stato"), e["classificazione"])] += 1

    lines = []
    lines.append('"""')
    lines.append('Mappa apparato -> classificazione deducibilità (Ecotel).')
    lines.append('')
    lines.append('AUTO-GENERATO da scripts/generate_apparati_mapping.py')
    lines.append(f'Sorgente: {xlsx_source.name} (foglio "{sheet_name}")')
    lines.append(f'Generato: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append(f'Apparati totali: {len(by_id)}')
    lines.append(f'  - distribuzione per (stato, classificazione):')
    for (st, cls), n in sorted(counter.items()):
        lines.append(f'    {st} / {cls}: {n}')
    lines.append(f'Skipped (stato non ammesso): {skipped_count}')
    lines.append(f'Duplicati intra-cc (tenuta l\'ultima): {dup_intra_count}')
    lines.append(f'Apparati multi-cc (cross-contratto): {multi_cc_count}')
    lines.append('')
    lines.append('Stati: MAPPATO/CENSITO=attivi, DISMESSO/GUASTA=storici (mantenuti')
    lines.append('per consentire la classificazione di fatture pregresse).')
    lines.append('')
    lines.append('NON modificare a mano: rigenerare con')
    lines.append('    python scripts/generate_apparati_mapping.py')
    lines.append('"""')
    lines.append('from typing import Optional, List')
    lines.append('')
    lines.append('CLASSIFICAZIONE_USO_PROMISCUO = "uso_promiscuo"')
    lines.append('CLASSIFICAZIONE_FURGONI = "furgoni"')
    lines.append('')
    lines.append('STATI_ATTIVI = {"MAPPATO", "CENSITO"}')
    lines.append('STATI_STORICI = {"DISMESSO", "GUASTA"}')
    lines.append('')
    lines.append('# Mappa codice_apparato -> metadati')
    lines.append('APPARATI_MAP = {')
    for aid in sorted_ids:
        e = by_id[aid]
        veic = (e.get("veicolo_descrizione") or "").replace("'", "\\'")
        note = (e.get("note") or "").replace("'", "\\'")
        forn = (e.get("fornitori") or "").replace("'", "\\'")
        lines.append(f'    "{aid}": {{')
        lines.append(f'        "stato": "{e.get("stato","")}",')
        lines.append(f'        "tipo_apparato": "{e.get("tipo_apparato","")}",')
        lines.append(f'        "fornitori": \'{forn[:60]}\',')
        lines.append(f'        "cc_cliente": "{e.get("cc_cliente","")}",')
        cc_extra = e.get("cc_clienti_aggiuntivi")
        if cc_extra:
            lines.append(f'        "cc_clienti_aggiuntivi": {sorted(set(cc_extra))!r},')
        lines.append(f'        "veicolo_descrizione": \'{veic[:120]}\',')
        lines.append(f'        "targa": "{e.get("targa","")}",')
        lines.append(f'        "classificazione": "{e["classificazione"]}",')
        if note:
            lines.append(f'        "note": \'{note[:120]}\',')
        lines.append(f'    }},')
    lines.append('}')
    lines.append('')
    lines.append('# Alias: codici "corti" come appaiono nei PDF Autostrade')
    lines.append('APPARATI_ALIAS = {')
    for short, main in sorted(aliases.items()):
        lines.append(f'    "{short}": "{main}",')
    lines.append('}')
    lines.append('')
    lines.append('# Indice inverso targa -> list[apparato_id]')
    lines.append('TARGA_INDEX = {}')
    lines.append('for _aid, _e in APPARATI_MAP.items():')
    lines.append('    _t = _e.get("targa")')
    lines.append('    if _t:')
    lines.append('        TARGA_INDEX.setdefault(_t, []).append(_aid)')
    lines.append('')
    lines.append('# Indice inverso cc_cliente -> list[apparato_id]')
    lines.append('CC_INDEX = {}')
    lines.append('for _aid, _e in APPARATI_MAP.items():')
    lines.append('    _cc = _e.get("cc_cliente")')
    lines.append('    if _cc:')
    lines.append('        CC_INDEX.setdefault(_cc, []).append(_aid)')
    lines.append('    for _cc_x in _e.get("cc_clienti_aggiuntivi", []):')
    lines.append('        CC_INDEX.setdefault(_cc_x, []).append(_aid)')
    lines.append('')
    lines.append('')
    lines.append('def normalize_apparato_lookup(raw: str) -> str:')
    lines.append('    """Normalizza un codice apparato per lookup (allineato al generator)."""')
    lines.append('    import re')
    lines.append('    raw = (raw or "").strip()')
    lines.append('    if not raw:')
    lines.append('        return ""')
    lines.append('    if "VIACARD" in raw.upper():')
    lines.append('        m = re.search(r"\\b(\\d+\\.\\d{6}\\.\\d+)\\b", raw)')
    lines.append('        if m:')
    lines.append('            return m.group(1)')
    lines.append('    no_spaces = re.sub(r"\\s+", "", raw)')
    lines.append('    digit_seqs = re.findall(r"\\d{6,}", no_spaces)')
    lines.append('    if digit_seqs:')
    lines.append('        return max(digit_seqs, key=len)')
    lines.append('    return ""')
    lines.append('')
    lines.append('')
    lines.append('def _resolve_key(key: str) -> Optional[str]:')
    lines.append('    if key in APPARATI_MAP:')
    lines.append('        return key')
    lines.append('    main = APPARATI_ALIAS.get(key)')
    lines.append('    return main if main in APPARATI_MAP else None')
    lines.append('')
    lines.append('')
    lines.append('def get_classificazione(apparato_raw: str) -> Optional[str]:')
    lines.append('    """Ritorna la classificazione (\'uso_promiscuo\'|\'furgoni\') o None."""')
    lines.append('    key = normalize_apparato_lookup(apparato_raw)')
    lines.append('    resolved = _resolve_key(key)')
    lines.append('    return APPARATI_MAP[resolved]["classificazione"] if resolved else None')
    lines.append('')
    lines.append('')
    lines.append('def get_apparato_info(apparato_raw: str):')
    lines.append('    """Ritorna l\'intera entry della mappa (dict) o None."""')
    lines.append('    key = normalize_apparato_lookup(apparato_raw)')
    lines.append('    resolved = _resolve_key(key)')
    lines.append('    return APPARATI_MAP.get(resolved) if resolved else None')
    lines.append('')
    lines.append('')
    lines.append('def get_apparati_by_targa(targa: str) -> List[str]:')
    lines.append('    return list(TARGA_INDEX.get((targa or "").upper().strip(), []))')
    lines.append('')
    lines.append('')
    lines.append('def get_apparati_by_cc(cc: str) -> List[str]:')
    lines.append('    return list(CC_INDEX.get((cc or "").strip(), []))')
    lines.append('')

    return "\n".join(lines)


def load_previous_map(output_path: Path):
    """Importa la versione precedente di config/apparati_mapping.py per il diff."""
    if not output_path.exists():
        return None
    try:
        spec = importlib.util.spec_from_file_location("_prev_apparati", output_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return getattr(mod, "APPARATI_MAP", None)
    except Exception as e:
        print(f"[!] Impossibile caricare versione precedente per il diff: {e}")
        return None


def compute_diff(prev_map, new_map):
    """Confronta due APPARATI_MAP e ritorna (added, removed, modified)."""
    added = []
    removed = []
    modified = []
    for aid in sorted(set(prev_map or {}) | set(new_map or {})):
        in_prev = aid in (prev_map or {})
        in_new = aid in (new_map or {})
        if in_new and not in_prev:
            added.append((aid, new_map[aid]))
        elif in_prev and not in_new:
            removed.append((aid, prev_map[aid]))
        elif in_prev and in_new:
            prev = prev_map[aid] or {}
            curr = new_map[aid] or {}
            changes = {}
            for k in ("stato", "classificazione", "targa", "cc_cliente",
                      "veicolo_descrizione", "tipo_apparato", "note"):
                pv = prev.get(k)
                cv = curr.get(k)
                if (pv or "") != (cv or ""):
                    changes[k] = (pv, cv)
            if changes:
                modified.append((aid, changes))
    return added, removed, modified


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
        help=f"Path al file Excel sorgente (default: {DEFAULT_XLSX})")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
        help=f"Path del file Python da generare (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--print-summary", action="store_true",
        help="Stampa riepilogo dettagliato (entries, skipped, duplicati)")
    parser.add_argument("--no-diff", action="store_true",
        help="Salta il diff vs versione precedente")
    args = parser.parse_args()

    print(f"== Generator apparati mapping ==")
    print(f"  Sorgente: {args.xlsx}")
    print(f"  Output:   {args.output}")

    entries, skipped, anomalie, sheet_name = parse_xlsx(args.xlsx)
    print(f"  Foglio letto: '{sheet_name}'")
    print(f"  Righe valide (stato in {sorted(STATI_AMMESSI)}): {len(entries)}")
    if skipped:
        print(f"  Skipped (stato non ammesso): {len(skipped)}")
    if anomalie:
        print(f"  Anomalie (codice vuoto): {len(anomalie)}")

    by_id, dup_intra, multi_cc, incoerenze_cls = build_apparati_map(entries)
    aliases = build_aliases(by_id)

    if dup_intra:
        print(f"\n[!] {len(dup_intra)} duplicati intra-cc (stessa coppia codice+cc, tenuta l'ULTIMA):")
        for d in dup_intra[:10]:
            print(f"    {d['curr']['codice_apparato']} cc={d['curr']['cc_cliente']}: "
                  f"prev '{d['prev']['targa']}' -> curr '{d['curr']['targa']}'")

    if multi_cc:
        print(f"\n[i] {len(multi_cc)} apparati condivisi tra cc multipli (warning informativo):")
        for cid, ccs in list(multi_cc.items())[:10]:
            primary_cc = by_id[cid]["cc_cliente"]
            print(f"    {cid}: cc primario={primary_cc}, cc aggiuntivi={ccs}")

    if incoerenze_cls:
        print(f"\n[!!] {len(incoerenze_cls)} INCOERENZE classificazione tra cc diversi:")
        for inc in incoerenze_cls:
            print(f"    {inc['codice']}: cc {inc['prev_cc']}->{inc['prev_cls']} "
                  f"vs cc {inc['curr_cc']}->{inc['curr_cls']}")

    # Genera sorgente
    source = render_module(by_id, aliases, args.xlsx, sheet_name,
                            len(skipped), len(dup_intra), len(multi_cc))

    # Diff vs versione precedente (prima di sovrascrivere)
    if not args.no_diff:
        prev_map = load_previous_map(args.output)
        if prev_map is not None:
            added, removed, modified = compute_diff(prev_map, by_id)
            print(f"\n== Diff vs versione precedente ==")
            print(f"  + Nuovi:      {len(added)}")
            print(f"  - Rimossi:    {len(removed)}")
            print(f"  ~ Modificati: {len(modified)}")
            if args.print_summary or (added and len(added) <= 30):
                for aid, e in added[:30]:
                    print(f"    + {aid} cc={e.get('cc_cliente','')} "
                          f"targa={e.get('targa','')} {e.get('classificazione','')}")
                if len(added) > 30:
                    print(f"    ... +{len(added)-30} altri")
            if removed and (args.print_summary or len(removed) <= 30):
                for aid, e in removed[:30]:
                    print(f"    - {aid} (era {e.get('classificazione','')})")
                if len(removed) > 30:
                    print(f"    ... -{len(removed)-30} altri")
            if modified and (args.print_summary or len(modified) <= 30):
                for aid, ch in modified[:30]:
                    pieces = ", ".join(f"{k}: {p!r}->{c!r}" for k, (p, c) in ch.items())
                    print(f"    ~ {aid} {pieces}")
                if len(modified) > 30:
                    print(f"    ... ~{len(modified)-30} altri")
        else:
            print(f"\n== Diff: file precedente assente o non leggibile, skip ==")

    # Scrivi
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(source, encoding="utf-8")
    print(f"\nScritto {args.output} ({len(source)} bytes, {len(by_id)} entries, "
          f"{len(aliases)} alias)")

    # Summary classificazione
    classifica = defaultdict(int)
    stati_count = defaultdict(int)
    for e in by_id.values():
        classifica[e["classificazione"]] += 1
        stati_count[e.get("stato","")] += 1
    print(f"Classificazione: {dict(classifica)}")
    print(f"Stati:           {dict(stati_count)}")

    if args.print_summary and skipped:
        print(f"\n[!] {len(skipped)} righe skipped:")
        for s in skipped[:20]:
            print(f"    riga {s['row']}: stato='{s['stato']}' ({s['motivo']})")

    if args.print_summary and anomalie:
        print(f"\n[!] {len(anomalie)} anomalie codice vuoto:")
        for a in anomalie[:20]:
            print(f"    riga {a['row']}: stato='{a['stato']}' raw='{a['codice_raw']}'")


if __name__ == "__main__":
    main()
