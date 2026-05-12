"""
Genera report Excel sull'inventario OdA per i 7 fornitori automezzi.
Read-only su Odoo via XML-RPC.

Output: output/report_oda_automezzi_<TS>.xlsx (4 fogli):
  - Riepilogo (totali per fornitore)
  - OdA ATTIVI (con POL libere, target dell'agent)
  - OdA SATURI / storici (riferimento storico)
  - Suggerimento BULK POL per Acquisti
"""
import os
import sys
from datetime import datetime
from collections import Counter
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT))

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.odoo_client import OdooReadOnlyClient


PARTNERS = [
    ("UnipolRental", 1937),
    ("Leasys", 1638),
    ("Tecnoalt", 1305),
    ("Athlon", 1984),
    ("Arval", 1287),
    ("ALD Automotive", 1000),
    ("LeasePlan", 1415),
]


def query_oda_data():
    load_dotenv(PROJECT / "config" / "credentials.env")
    cli = OdooReadOnlyClient(
        os.environ["ODOO_URL"], os.environ["ODOO_DB"],
        os.environ["ODOO_USERNAME"], os.environ["ODOO_PASSWORD"],
    )
    cli.connect()

    rows = []
    for name, pid in PARTNERS:
        pos = cli._call(
            "purchase.order", "search_read",
            [("partner_id", "=", pid), ("company_id", "=", 1)],
            fields=["id", "name", "state", "invoice_status",
                     "amount_total", "amount_untaxed", "date_order"],
            order="date_order desc", limit=50,
        )
        for po in pos:
            lines = cli._call(
                "purchase.order.line", "search_read",
                [("order_id", "=", po["id"])],
                fields=["id", "qty_invoiced", "qty_received", "product_qty"],
            )
            n_tot = len(lines)
            n_lib = sum(
                1 for l in lines
                if (l["qty_invoiced"] or 0) == 0
                and (l["qty_received"] or 0) == 0
                and (l["product_qty"] or 0) >= 1
            )
            rows.append({
                "fornitore": name,
                "partner_id": pid,
                "oda_name": po["name"],
                "oda_id": po["id"],
                "state": po["state"],
                "invoice_status": po["invoice_status"],
                "n_pol_tot": n_tot,
                "n_pol_libere": n_lib,
                "amount_total": po["amount_total"],
                "amount_untaxed": po["amount_untaxed"],
                "date_order": (po.get("date_order") or "")[:10],
            })
    return rows


# Stili condivisi
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
TITLE_FILL = PatternFill(start_color="FF1A3A5C", end_color="FF1A3A5C", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
SECT_FONT = Font(bold=True, color="FF1A3A5C", size=11)
SECT_FILL = PatternFill(start_color="FFE0E8F0", end_color="FFE0E8F0", fill_type="solid")
ATTIVO_FILL = PatternFill(start_color="FFE8F5E8", end_color="FFE8F5E8", fill_type="solid")
SATURO_FILL = PatternFill(start_color="FFFFE0E0", end_color="FFFFE0E0", fill_type="solid")
TOT_FONT = Font(bold=True)
TOT_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")


def fill_riepilogo(ws, rows):
    ws.merge_cells("A1:G1")
    ws["A1"] = (f"Inventario OdA Automezzi — riepilogo "
                f"(generato {datetime.now().strftime('%d/%m/%Y %H:%M')})")
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Fornitore", "OdA totali", "OdA attivi",
                "OdA saturi/storici", "POL libere (su attivi)",
                "Importo OdA attivi €", "Note"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r = 4
    grand_attivi = grand_saturi = grand_libere = 0
    grand_amount = 0.0
    for fornitore, _ in PARTNERS:
        sub = [x for x in rows if x["fornitore"] == fornitore]
        attivi = [x for x in sub
                   if x["state"] == "purchase" and x["n_pol_libere"] > 0]
        saturi = [x for x in sub
                   if not (x["state"] == "purchase" and x["n_pol_libere"] > 0)]
        libere = sum(x["n_pol_libere"] for x in attivi)
        amount = sum(x["amount_total"] for x in attivi)
        nota = ""
        if not attivi and sub:
            nota = "Nessun OdA attivo - DA APRIRE"
        elif not sub:
            nota = "Nessun OdA in Odoo"

        ws.cell(row=r, column=1, value=fornitore).font = TOT_FONT
        ws.cell(row=r, column=2, value=len(sub))
        ws.cell(row=r, column=3, value=len(attivi))
        ws.cell(row=r, column=4, value=len(saturi))
        ws.cell(row=r, column=5, value=libere)
        ws.cell(row=r, column=6, value=amount).number_format = "#,##0.00"
        ws.cell(row=r, column=7, value=nota)
        if not attivi and sub:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = SATURO_FILL

        grand_attivi += len(attivi)
        grand_saturi += len(saturi)
        grand_libere += libere
        grand_amount += amount
        r += 1

    # Totale
    ws.cell(row=r, column=1, value="TOTALE").font = TOT_FONT
    ws.cell(row=r, column=1).fill = TOT_FILL
    ws.cell(row=r, column=2, value=len(rows)).font = TOT_FONT
    ws.cell(row=r, column=2).fill = TOT_FILL
    ws.cell(row=r, column=3, value=grand_attivi).font = TOT_FONT
    ws.cell(row=r, column=3).fill = TOT_FILL
    ws.cell(row=r, column=4, value=grand_saturi).font = TOT_FONT
    ws.cell(row=r, column=4).fill = TOT_FILL
    ws.cell(row=r, column=5, value=grand_libere).font = TOT_FONT
    ws.cell(row=r, column=5).fill = TOT_FILL
    cell = ws.cell(row=r, column=6, value=grand_amount)
    cell.font = TOT_FONT
    cell.fill = TOT_FILL
    cell.number_format = "#,##0.00"

    # Volume mensile atteso (calcolato da analisi Q1 2026)
    r += 3
    ws.cell(row=r, column=1, value="VOLUME MENSILE ATTESO (da fatture posted Q1 2026)").font = SECT_FONT
    ws.cell(row=r, column=1).fill = SECT_FILL
    r += 1
    vol_headers = ["Fornitore", "Fatture/mese", "Righe/fattura medie", "POL/mese attese"]
    for i, h in enumerate(vol_headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r += 1
    volumi = [
        ("Leasys", 6, 5, 30),
        ("Tecnoalt", 8, 1.5, 12),
        ("UnipolRental", 4, 2.5, 10),
        ("ALD Automotive", 2, 2, 4),
        ("Athlon", 2, 2, 4),
        ("Arval", 1, 2, 2),
        ("LeasePlan", 0, 0, 0),
    ]
    tot_pol_mese = 0
    for forn, fatt, righe, pol in volumi:
        ws.cell(row=r, column=1, value=forn)
        ws.cell(row=r, column=2, value=fatt)
        ws.cell(row=r, column=3, value=righe)
        ws.cell(row=r, column=4, value=pol)
        tot_pol_mese += pol
        r += 1
    ws.cell(row=r, column=1, value="TOTALE atteso").font = TOT_FONT
    ws.cell(row=r, column=1).fill = TOT_FILL
    ws.cell(row=r, column=4, value=tot_pol_mese).font = TOT_FONT
    ws.cell(row=r, column=4).fill = TOT_FILL
    r += 1
    if grand_libere > 0:
        mesi_copertura = grand_libere / tot_pol_mese if tot_pol_mese else 0
        ws.cell(row=r, column=1,
                  value=f"COPERTURA STIMATA: {grand_libere} libere / "
                        f"{tot_pol_mese} POL/mese ≈ {mesi_copertura:.1f} mesi"
                  ).font = TOT_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    # Larghezze
    widths = [22, 12, 14, 18, 22, 22, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def fill_oda_table(ws, rows, only_active=True):
    title = "OdA ATTIVI (rilevanti per l'agent)" if only_active else "OdA SATURI / storici"
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Fornitore", "OdA", "Stato", "Invoice status",
                "Data ordine", "POL totali", "POL libere", "Importo €"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL

    r = 4
    for fornitore, _ in PARTNERS:
        sub = [x for x in rows if x["fornitore"] == fornitore]
        if only_active:
            sub = [x for x in sub
                    if x["state"] == "purchase" and x["n_pol_libere"] > 0]
        else:
            sub = [x for x in sub
                    if not (x["state"] == "purchase" and x["n_pol_libere"] > 0)]
        sub.sort(key=lambda x: x["date_order"], reverse=True)
        for x in sub:
            ws.cell(row=r, column=1, value=x["fornitore"])
            ws.cell(row=r, column=2, value=x["oda_name"])
            ws.cell(row=r, column=3, value=x["state"])
            ws.cell(row=r, column=4, value=x["invoice_status"])
            ws.cell(row=r, column=5, value=x["date_order"])
            ws.cell(row=r, column=6, value=x["n_pol_tot"])
            ws.cell(row=r, column=7, value=x["n_pol_libere"])
            ws.cell(row=r, column=8, value=x["amount_total"]).number_format = "#,##0.00"
            fill = ATTIVO_FILL if only_active else SATURO_FILL
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = fill
            r += 1

    widths = [20, 14, 12, 14, 12, 12, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def fill_suggerimento_bulk(ws, rows):
    ws.merge_cells("A1:E1")
    ws["A1"] = "Suggerimento Acquisti — preparazione POL libere bulk"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    r = 3
    ws.cell(row=r, column=1, value="ISTRUZIONI GENERALI").font = SECT_FONT
    ws.cell(row=r, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    note = [
        "L'agent consume-POL multi-line richiede POL libere pre-create da Acquisti su ogni OdA.",
        "Le POL libere vengono RISCRITTE COMPLETAMENTE dall'agent al consumo (name, price, tax, conto).",
        "Pattern raccomandato: name='TEST', price_unit=€1, tax_id=11, qty=1, 1 POL/mese.",
        "Quando l'agent finisce le POL libere, segnala: 'POL libere insufficienti su {oda}'.",
        "L'utente puo' rilanciare i bulk dopo che Acquisti ha aggiunto altre POL.",
    ]
    for n in note:
        ws.cell(row=r, column=1, value=("• " + n))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    r += 1
    # Tabella consigli per OdA attivo
    ws.cell(row=r, column=1, value="CONSIGLI PER OdA ATTIVO").font = SECT_FONT
    ws.cell(row=r, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers = ["Fornitore", "OdA", "POL libere oggi", "POL/mese atteso", "Azione consigliata"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r += 1

    consigli = []
    pol_mese_per_fornitore = {
        "Leasys": 30, "Tecnoalt": 12, "UnipolRental": 10,
        "ALD Automotive": 4, "Athlon": 4, "Arval": 2, "LeasePlan": 0,
    }
    for fornitore, _ in PARTNERS:
        attivi = [x for x in rows
                   if x["fornitore"] == fornitore
                   and x["state"] == "purchase"
                   and x["n_pol_libere"] > 0]
        pol_mese = pol_mese_per_fornitore.get(fornitore, 0)
        if not attivi:
            consigli.append((fornitore, "(nessun OdA attivo)", 0, pol_mese,
                             f"DA APRIRE nuovo OdA + ~{pol_mese*9} POL TEST per 9 mesi" if pol_mese else
                             "Nessuna fattura attesa, aprire al primo arrivo"))
            continue
        for x in attivi:
            mesi = x["n_pol_libere"] / max(pol_mese / max(len(attivi), 1), 1) if pol_mese else 999
            if mesi >= 5:
                az = f"OK (~{mesi:.0f} mesi copertura)"
            elif mesi >= 2:
                az = f"OK breve termine (~{mesi:.1f} mesi). Ricaricare entro Q3 2026"
            elif mesi >= 1:
                az = f"⚠ Ricaricare presto (~{mesi:.1f} mesi)"
            else:
                az = f"⚠⚠ URGENTE: aggiungere ~{int(pol_mese*6)} POL TEST"
            consigli.append((fornitore, x["oda_name"], x["n_pol_libere"], pol_mese, az))

    for forn, oda, libere, pol_mese, az in consigli:
        ws.cell(row=r, column=1, value=forn)
        ws.cell(row=r, column=2, value=oda)
        ws.cell(row=r, column=3, value=libere)
        ws.cell(row=r, column=4, value=pol_mese)
        ws.cell(row=r, column=5, value=az)
        if "URGENTE" in az or "DA APRIRE" in az:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = SATURO_FILL
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="PATTERN POL TEST GENERIC (consigliato per la maggior parte degli OdA)").font = SECT_FONT
    ws.cell(row=r, column=1).fill = SECT_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    pattern_lines = [
        ("Campo", "Valore", "Nota"),
        ("name", "TEST", "L'agent lo riscrive con voce + targa + cc"),
        ("price_unit", "1.00", "L'agent lo riscrive con importo riga XML"),
        ("tax_id", "11 (22% S)", "L'agent lo riscrive con tax giusto per fornitore (6/11/47/54/73)"),
        ("product_qty", "1", "Quantita standard"),
        ("date_planned", "ultimo del mese", "L'agent lo aggiorna a invoice_date"),
        ("product_id", "noleggio (categoria standard)", "L'agent eredita"),
        ("account_analytic_id", "(vuoto)", "L'agent non lo tocca"),
    ]
    for line in pattern_lines:
        for i, val in enumerate(line, start=1):
            c = ws.cell(row=r, column=i, value=val)
            if line == pattern_lines[0]:
                c.font = HDR_FONT
                c.fill = HDR_FILL
        r += 1

    widths = [22, 16, 14, 16, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def main():
    print("Connessione Odoo...")
    rows = query_oda_data()
    print(f"Recuperati {len(rows)} OdA totali per i 7 fornitori")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Riepilogo"
    fill_riepilogo(ws1, rows)

    ws2 = wb.create_sheet("OdA ATTIVI")
    fill_oda_table(ws2, rows, only_active=True)

    ws3 = wb.create_sheet("OdA saturi (storici)")
    fill_oda_table(ws3, rows, only_active=False)

    ws4 = wb.create_sheet("Suggerimento BULK")
    fill_suggerimento_bulk(ws4, rows)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = PROJECT / "output" / f"report_oda_automezzi_{ts}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Generato: {out}")
    print(f"Dimensione: {out.stat().st_size:,} bytes")
    print(f"Fogli: {wb.sheetnames}")


if __name__ == "__main__":
    main()
