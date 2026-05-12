"""
Genera un Excel "preview" che simula le bozze account.move che l'agent
costruirebbe per le 3 fatture Autostrade in input/, applicando la nuova
mappa apparati e la stessa logica del writer create_bozza_autostrade.

Niente scritture su Odoo: lavora dai PDF locali + interroga in sola lettura
P03718 per riportare partner_id, currency_id, conti, residuo righe libere.

Uso:
    python scripts/preview_bozze_autostrade.py
    python scripts/preview_bozze_autostrade.py --output <path.xlsx>
"""
import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import pdfplumber
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from config.apparati_mapping import (APPARATI_MAP, get_classificazione,
                                       get_apparato_info)
from config.rules import MAPPATURA_FORNITORI_FISSI, resolve_mapping_entry
from core.pdf_parser import parse_pdf_autostrade, calcola_split_furgoni_promiscuo
from core.odoo_client import OdooReadOnlyClient


PIVA_AUTOSTRADE = 'IT07516911000'
DEFAULT_OUTPUT = (PROJECT_ROOT / 'output' /
                   f'preview_bozze_autostrade_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')


def extract_pdf_header(pdf_path: Path) -> dict:
    """Estrae cc, numero, data, imponibile dalla prima pagina del PDF."""
    with pdfplumber.open(str(pdf_path)) as pdf:
        text = pdf.pages[0].extract_text() or ''

    out = {'pdf_path': str(pdf_path), 'cc': None, 'numero': None,
            'data': None, 'imponibile_xml': None, 'totale_iva_incl': None}

    # cc cliente (riga successiva a "CODICE CLIENTE NUMERO VIAGGI ...")
    m = re.search(r'CODICE\s+CLIENTE.*?\n\s*(\d{6,12})', text, re.IGNORECASE)
    if m:
        out['cc'] = m.group(1)

    # numero + data fattura
    # Formato: "000000002740429D del 30-04-2026" oppure "del 30-apr-2026"
    m = re.search(r'(0+\d+D)\s+del\s+(\d{1,2}-(?:[a-z]+|\d{1,2})-\d{4})',
                    text, re.IGNORECASE)
    if m:
        out['numero'] = m.group(1)
        out['data'] = _parse_it_date(m.group(2))

    # Totale imponibile (alla fine del riepilogo IVA)
    m = re.search(r'Totale\s+imponibile\s*[€€]?\s*([\d.,]+)', text, re.IGNORECASE)
    if m:
        out['imponibile_xml'] = _to_float(m.group(1))

    # Totale fattura (riga "TOTALE FATTURA" o "Totale Documento")
    m = re.search(r'(?:TOTALE\s+FATTURA|Totale\s+Documento)\s*[€€]?\s*([\d.,]+)', text, re.IGNORECASE)
    if m:
        out['totale_iva_incl'] = _to_float(m.group(1))

    return out


def _to_float(s: str) -> float:
    """Converte '1.234,56' o '724,10' in float."""
    if not s:
        return 0.0
    return float(s.replace('.', '').replace(',', '.'))


def _parse_it_date(s: str) -> str:
    """'30-apr-2026' o '30-04-2026' -> '2026-04-30' (ISO)."""
    months = {'gen': '01', 'feb': '02', 'mar': '03', 'apr': '04',
              'mag': '05', 'giu': '06', 'lug': '07', 'ago': '08',
              'set': '09', 'ott': '10', 'nov': '11', 'dic': '12'}
    s = s.lower()
    # Formato numerico
    m = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', s)
    if m:
        d, mn, y = m.groups()
        return f'{y}-{int(mn):02d}-{int(d):02d}'
    # Formato testuale
    m = re.match(r'(\d{1,2})-([a-z]+)-(\d{4})', s)
    if m:
        d, mon, y = m.groups()
        mon_n = months.get(mon[:3], '01')
        return f'{y}-{mon_n}-{int(d):02d}'
    return s


def end_of_month(date_iso: str) -> str:
    if not date_iso or len(date_iso) < 7:
        return date_iso
    y, m, _ = date_iso.split('-')
    last_day = {'01': 31, '02': 28, '03': 31, '04': 30, '05': 31, '06': 30,
                 '07': 31, '08': 31, '09': 30, '10': 31, '11': 30, '12': 31}[m]
    if m == '02' and int(y) % 4 == 0:
        last_day = 29
    return f'{y}-{m}-{last_day:02d}'


def query_oda_state(oda_name: str = 'P03718'):
    """Read-only Odoo: stato attuale dell'OdA + tutte le POL + fatture linked."""
    load_dotenv(PROJECT_ROOT / 'config' / 'credentials.env')
    cli = OdooReadOnlyClient(
        os.environ['ODOO_URL'], os.environ['ODOO_DB'],
        os.environ['ODOO_USERNAME'], os.environ['ODOO_PASSWORD'])
    cli.connect()
    pos = cli._call('purchase.order', 'search_read',
        [('name', '=', oda_name), ('company_id', '=', 1)],
        fields=['id', 'name', 'state', 'partner_id', 'currency_id',
                 'amount_total', 'amount_untaxed', 'invoice_status',
                 'invoice_ids', 'date_order'],
        limit=1)
    if not pos:
        return None
    po = pos[0]
    lines = cli._call('purchase.order.line', 'search_read',
        [('order_id', '=', po['id'])],
        fields=['id', 'qty_invoiced', 'qty_received', 'product_qty',
                 'price_unit', 'price_subtotal', 'name', 'product_id',
                 'taxes_id'])
    libere = sum(1 for l in lines
                  if l['qty_invoiced'] == 0 and l['qty_received'] == 0
                  and l['product_qty'] >= 1)
    po['n_lines_total'] = len(lines)
    po['n_lines_libere'] = libere
    po['lines'] = lines

    # Linked bills tramite invoice_ids della PO
    po['linked_bills'] = []
    if po.get('invoice_ids'):
        bills = cli._call('account.move', 'search_read',
            [('id', 'in', po['invoice_ids'])],
            fields=['id', 'name', 'state', 'invoice_date', 'amount_total',
                     'amount_untaxed', 'ref', 'invoice_origin'])
        po['linked_bills'] = bills
    return po


def simulate_one(pdf_path: Path, oda_state: dict) -> dict:
    """Per un PDF, costruisce il dizionario 'preview' delle 2 move_line."""
    header = extract_pdf_header(pdf_path)
    cc = header['cc']
    if not cc:
        return {'pdf_path': str(pdf_path), 'error': 'cc non estratto da PDF'}

    # Parse apparati
    pdf_data = parse_pdf_autostrade(str(pdf_path))

    # Resolve mapping (xml_data fittizio con cc)
    base_entry = MAPPATURA_FORNITORI_FISSI.get(PIVA_AUTOSTRADE)
    fake_xml = type('XmlStub', (), {})()
    fake_xml.id_documento = None
    fake_xml.codice_cliente = cc
    fake_xml.dati_contratto = []
    fake_xml.numero_ordine_acquisto = None
    fake_xml.data_ordine_acquisto = None
    mapping = resolve_mapping_entry(base_entry, fake_xml)
    if not mapping:
        return {'pdf_path': str(pdf_path), 'cc': cc,
                'error': f'cc {cc} non risolto in mapping'}

    cc_type = mapping.get('cc_type')
    oda_name = mapping.get('oda_fisso')

    imponibile_xml = header['imponibile_xml']
    if not imponibile_xml and pdf_data.totale_fattura_iva_inclusa:
        imponibile_xml = round(pdf_data.totale_fattura_iva_inclusa / 1.22, 2)

    split = calcola_split_furgoni_promiscuo(
        pdf_data, imponibile_xml or 0, get_classificazione)

    # Compose move_line preview (mirror del writer Autostrade)
    tax_id = (mapping.get('taxes_id') or [11])[0]
    conto_furgoni = mapping.get('conto_furgoni_id', 368)
    conto_promiscuo = mapping.get('conto_promiscuo_id', 1124)
    move_lines = []
    if cc_type == 'ecotel_main':
        move_lines.append({
            'name': f"{oda_name}: PEDAGGI AUTOSTRADALI | "
                    f"Codice cliente: {cc} | furgoni",
            'account_id_label': '420160 (id 368) — Pedaggi 100%',
            'account_id': conto_furgoni,
            'price_unit': split['imponibile_furgoni'],
            'quantity': 1,
            'tax_id': tax_id,
            'tax_label': 'IVA 22% S (id 11)',
        })
        move_lines.append({
            'name': f"{oda_name}: PEDAGGI AUTOSTRADALI | "
                    f"Codice cliente: {cc} | uso promiscuo",
            'account_id_label': '420840 (id 1124) — Pedaggi 70%',
            'account_id': conto_promiscuo,
            'price_unit': split['imponibile_promiscuo'],
            'quantity': 1,
            'tax_id': tax_id,
            'tax_label': 'IVA 22% S (id 11)',
        })

    return {
        'pdf_path': str(pdf_path),
        'header': header,
        'cc': cc,
        'cc_type': cc_type,
        'oda_name': oda_name,
        'imponibile_xml': imponibile_xml,
        'split': split,
        'move_lines': move_lines,
        'mapping_taxes': mapping.get('taxes_id'),
        'partner_id': oda_state['partner_id'][0] if oda_state and oda_state.get('partner_id') else None,
        'currency_id': oda_state['currency_id'][0] if oda_state and oda_state.get('currency_id') else None,
        'date_contabile': '(create_date attachment SDI — disponibile a runtime)',
        'date_iva_competenza': end_of_month(header.get('data') or ''),
    }


def match_pol_for_invoice(lines, cc: str, classificazione: str,
                            excluded_ids=None):
    """Cerca POL libere matchando (cc + cls). Mirror della logica del writer:
    priorità a POL cc-specifica (descrizione SOLO quel cc), poi POL jolly
    (descrizione multi-cc).

    Allineato con OdooWriter._find_pol_autostrade_match (core/odoo_writer.py).
    """
    excluded = excluded_ids or set()
    cls_token = classificazione.replace('_', ' ')
    cc_specific = []
    jolly = []
    for ln in lines:
        if ln['id'] in excluded:
            continue
        if ((ln.get('qty_invoiced') or 0) != 0
            or (ln.get('qty_received') or 0) != 0
            or (ln.get('product_qty') or 0) < 1):
            continue
        desc = (ln.get('name') or '').lower()
        if cc not in desc or cls_token not in desc:
            continue
        ccs_in_desc = re.findall(r'\b\d{8,12}\b', ln.get('name') or '')
        if len(set(ccs_in_desc)) <= 1:
            cc_specific.append(ln)
        else:
            jolly.append(ln)
    cc_specific.sort(key=lambda l: l['id'])
    jolly.sort(key=lambda l: l['id'])
    return cc_specific + jolly


def assign_pol_simulation(oda_state, previews):
    """Per ogni fattura, prova a determinare quali 2 POL libere
    (furgoni + uso_promiscuo) verrebbero consumate dal cc/classificazione.

    Ritorna: dict pol_id -> {'fattura_idx', 'classificazione', 'old_price', 'new_price'}
    e lista warning per fatture con candidati 0 o multipli.
    """
    if not oda_state:
        return {}, []
    lines = oda_state.get('lines', [])
    assignments = {}  # pol_id -> assign info
    warnings = []
    used = set()
    for idx, p in enumerate(previews, start=1):
        if 'error' in p and 'header' not in p:
            continue
        cc = p.get('cc', '')
        split = p.get('split', {})
        for cls, new_price in [('furgoni', split.get('imponibile_furgoni', 0)),
                                 ('uso_promiscuo', split.get('imponibile_promiscuo', 0))]:
            cands = match_pol_for_invoice(lines, cc, cls, excluded_ids=used)
            if not cands:
                warnings.append(
                    f'Fattura #{idx} ({cc}, {cls}): NESSUNA POL libera con cc + cls trovata')
                continue
            chosen = cands[0]
            if len(cands) > 1:
                warnings.append(
                    f'Fattura #{idx} ({cc}, {cls}): {len(cands)} POL candidate, '
                    f'scelgo POL {chosen["id"]} (FIFO). Altre: '
                    f'{[c["id"] for c in cands[1:]]}')
            cls_label = 'furgoni' if cls == 'furgoni' else 'uso promiscuo'
            new_name = (f"PEDAGGI AUTOSTRADALI\n"
                         f"Codice cliente: {cc}\n"
                         f"{cls_label}")
            assignments[chosen['id']] = {
                'fattura_idx': idx,
                'numero': p['header'].get('numero', ''),
                'cc': cc,
                'classificazione': cls,
                'old_price': chosen['price_unit'],
                'new_price': round(new_price, 2),
                'new_name': new_name,
            }
            used.add(chosen['id'])
    return assignments, warnings


def fill_oda_state(ws, oda_state, previews):
    """Foglio dedicato — snapshot OdA P03718 prima/dopo le 3 bozze."""
    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
    title_fill = PatternFill(start_color='FF1A3A5C', end_color='FF1A3A5C', fill_type='solid')
    sect_fill = PatternFill(start_color='FFE0E8F0', end_color='FFE0E8F0', fill_type='solid')
    sect_font = Font(bold=True, color='FF1A3A5C')
    bold = Font(bold=True)
    free_fill = PatternFill(start_color='FFE8F5E8', end_color='FFE8F5E8', fill_type='solid')
    used_fill = PatternFill(start_color='FFFAFAFA', end_color='FFFAFAFA', fill_type='solid')
    info_fill = PatternFill(start_color='FFFFF4E0', end_color='FFFFF4E0', fill_type='solid')
    sim_fill = PatternFill(start_color='FFFFD0D0', end_color='FFFFD0D0', fill_type='solid')
    sim_font = Font(bold=True, color='FFB00000')

    ws.merge_cells('A1:H1')
    ws['A1'] = 'STATO OdA P03718 — prima/dopo le 3 bozze Autostrade'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    if not oda_state:
        ws['A3'] = 'OdA non disponibile (Odoo non raggiungibile)'
        return

    # Calcola assegnazioni POL simulate
    assignments, sim_warnings = assign_pol_simulation(oda_state, previews)

    # Modello scrittura
    box = ws['A3']
    box.value = ('SIMULAZIONE consume-POL: il writer Autostrade dal 07/05/2026 '
                  'usa il pattern Trenitalia/Italo (consume-POL). Questa pagina '
                  'mostra esattamente cosa l\'agent scriverebbe sull\'OdA P03718. '
                  'Le 6 righe ROSSE sono le POL libere che verrebbero consumate '
                  '(1 furgoni + 1 promiscuo per ognuna delle 3 fatture, matchate '
                  'per cc + classificazione nella descrizione POL). Le colonne '
                  '"old -> new price" mostrano il delta. Le 6 righe diventeranno '
                  'qty_received=1, qty_received_manual=1, price_unit=<imponibile>, '
                  'name aggiornato col numero fattura. Le altre POL libere (29) '
                  'restano invariate per le fatture future.')
    box.alignment = Alignment(wrap_text=True, vertical='top')
    box.fill = info_fill
    ws.merge_cells('A3:K3')
    ws.row_dimensions[3].height = 80

    # Header OdA
    rows_h = [
        ('Nome OdA', oda_state['name']),
        ('Stato', f'{oda_state["state"]} ({oda_state["invoice_status"]})'),
        ('Partner', oda_state['partner_id'][1] if oda_state.get('partner_id') else ''),
        ('Data ordine', oda_state.get('date_order', '')),
        ('Importo totale', f'€ {oda_state["amount_total"]:,.2f}'),
        ('Imponibile (untaxed)', f'€ {oda_state["amount_untaxed"]:,.2f}'),
        ('Righe POL totali', oda_state['n_lines_total']),
        ('Righe POL libere (qty_inv=0 AND qty_rec=0 AND qty>=1)', oda_state['n_lines_libere']),
        ('Fatture già collegate (Linked Bills)', len(oda_state.get('linked_bills', []))),
    ]
    r = 5
    ws.cell(row=r, column=1, value='HEADER OdA').font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for label, val in rows_h:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1

    # Tutte le POL con stato
    n_assigned = len(assignments)
    ws.cell(row=r, column=1, value=(f'TUTTE LE POL ({oda_state["n_lines_total"]}) '
                                      f'— rosso = SAREBBE CONSUMATA da una delle 3 nuove '
                                      f'fatture ({n_assigned}), verde = libera, '
                                      f'grigio = già fatturata')).font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    headers = ['#', 'POL id', 'Stato', 'qty', 'qty_invoiced', 'qty_received',
                'price_unit €', 'Descrizione (primi 80 char)',
                'Assegnata a fattura #', 'Vecchio price €', 'Nuovo price €',
                'Nuovo name (POL aggiornata)']
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = bold
        c.fill = sect_fill
    r += 1

    for idx, ln in enumerate(sorted(oda_state['lines'], key=lambda x: x['id']), start=1):
        is_free = (ln['qty_invoiced'] == 0 and ln['qty_received'] == 0
                    and ln['product_qty'] >= 1)
        assign = assignments.get(ln['id'])
        if assign:
            stato = 'CONSUMATA (sim)'
        elif is_free:
            stato = 'LIBERA'
        else:
            stato = 'consumata'
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=ln['id'])
        ws.cell(row=r, column=3, value=stato)
        ws.cell(row=r, column=4, value=ln['product_qty'])
        ws.cell(row=r, column=5, value=ln['qty_invoiced'])
        ws.cell(row=r, column=6, value=ln['qty_received'])
        ws.cell(row=r, column=7, value=ln['price_unit']).number_format = '#,##0.00'
        desc = (ln.get('name') or '').replace('\n', ' | ')
        ws.cell(row=r, column=8, value=desc[:80])
        if assign:
            ws.cell(row=r, column=9,
                value=f"#{assign['fattura_idx']} {assign['numero']} ({assign['cc']} {assign['classificazione']})")
            ws.cell(row=r, column=10, value=assign['old_price']).number_format = '#,##0.00'
            ws.cell(row=r, column=11, value=assign['new_price']).number_format = '#,##0.00'
            new_name_cell = ws.cell(row=r, column=12, value=assign.get('new_name', ''))
            new_name_cell.alignment = Alignment(wrap_text=True, vertical='top')
        # Coloring
        if assign:
            for col in range(1, 13):
                ws.cell(row=r, column=col).fill = sim_fill
                if col in (3, 9, 10, 11, 12):
                    ws.cell(row=r, column=col).font = sim_font
            # altezza riga per leggere il nuovo name multi-line
            ws.row_dimensions[r].height = 45
        else:
            fill = free_fill if is_free else used_fill
            for col in range(1, 13):
                ws.cell(row=r, column=col).fill = fill
        r += 1

    # Warning su matching POL (multipli o assenti)
    if sim_warnings:
        r += 1
        ws.cell(row=r, column=1, value='AVVISI MATCHING POL').font = sect_font
        ws.cell(row=r, column=1).fill = info_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        r += 1
        for w in sim_warnings:
            cell = ws.cell(row=r, column=1, value=w)
            cell.fill = info_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
            r += 1

    r += 1

    # Effetto delle 3 bozze
    ws.cell(row=r, column=1, value='EFFETTO DELLE 3 BOZZE — DELTA SU OdA').font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    n_assigned = len(assignments)
    delta_amount = sum(
        (assignments[k]['new_price'] - assignments[k]['old_price'])
        for k in assignments)
    delta_rows = [
        ('Righe POL aggiornate (rosse)', f'{n_assigned} su 6 attese (2 per ognuna delle 3 fatture)'),
        ('Righe libere prima -> dopo', f'{oda_state["n_lines_libere"]} -> {oda_state["n_lines_libere"] - n_assigned}'),
        ('amount_total OdA prima', f'€ {oda_state["amount_total"]:,.2f}'),
        ('amount_total OdA dopo', f'€ {oda_state["amount_total"] + delta_amount:,.2f} (delta {delta_amount:+,.2f})'),
        ('Linked Bills', f'{len(oda_state.get("linked_bills", []))} -> {len(oda_state.get("linked_bills", [])) + len([p for p in previews if "error" not in p or "header" in p])} (+3)'),
        ('Move line -> POL', '6 collegamenti via purchase_line_id (1 furgoni + 1 promiscuo per fattura)'),
    ]
    for label, val in delta_rows:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1

    # Le 3 bozze che si aggancerebbero
    ws.cell(row=r, column=1, value='NUOVE FATTURE COLLEGATE (via invoice_origin = "P03718")').font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    headers_b = ['#', 'Numero', 'Data', 'cc', 'Imponibile €',
                  'Furgoni €', 'Promiscuo €', 'Totale fattura €']
    for i, h in enumerate(headers_b):
        c = ws.cell(row=r, column=i+1, value=h)
        c.font = bold
        c.fill = sect_fill
    r += 1
    for i, p in enumerate(previews, start=1):
        if 'error' in p and 'header' not in p:
            continue
        h = p.get('header', {})
        split = p.get('split', {})
        imp_f = split.get('imponibile_furgoni', 0)
        imp_p = split.get('imponibile_promiscuo', 0)
        tot_imp = imp_f + imp_p
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=h.get('numero', ''))
        ws.cell(row=r, column=3, value=h.get('data', ''))
        ws.cell(row=r, column=4, value=p.get('cc', ''))
        ws.cell(row=r, column=5, value=p.get('imponibile_xml') or 0).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=imp_f).number_format = '#,##0.00'
        ws.cell(row=r, column=7, value=imp_p).number_format = '#,##0.00'
        ws.cell(row=r, column=8, value=tot_imp * 1.22).number_format = '#,##0.00'
        r += 1

    r += 1

    # Linked bills già esistenti
    bills = oda_state.get('linked_bills', [])
    if bills:
        ws.cell(row=r, column=1, value=f'LINKED BILLS GIÀ PRESENTI SU P03718 ({len(bills)})').font = sect_font
        ws.cell(row=r, column=1).fill = sect_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        headers_lb = ['#', 'move_id', 'Name move', 'Stato', 'Data', 'Ref', 'Imponibile €', 'Totale €']
        for i, h in enumerate(headers_lb):
            c = ws.cell(row=r, column=i+1, value=h)
            c.font = bold
            c.fill = sect_fill
        r += 1
        for i, b in enumerate(sorted(bills, key=lambda x: x.get('invoice_date') or '', reverse=True), start=1):
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=b['id'])
            ws.cell(row=r, column=3, value=b.get('name', ''))
            ws.cell(row=r, column=4, value=b.get('state', ''))
            ws.cell(row=r, column=5, value=b.get('invoice_date', ''))
            ws.cell(row=r, column=6, value=b.get('ref', ''))
            ws.cell(row=r, column=7, value=b.get('amount_untaxed', 0)).number_format = '#,##0.00'
            ws.cell(row=r, column=8, value=b.get('amount_total', 0)).number_format = '#,##0.00'
            r += 1

    # Layout
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 38
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 36


def fill_summary(ws, previews, oda_state):
    """Foglio 1 — Riepilogo."""
    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
    title_fill = PatternFill(start_color='FF1A3A5C', end_color='FF1A3A5C', fill_type='solid')
    bold = Font(bold=True)
    align_c = Alignment(horizontal='center', vertical='center')
    align_l = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.merge_cells('A1:H1')
    ws['A1'] = 'PREVIEW BOZZE AUTOSTRADE — fatture Aprile 2026'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = align_c

    ws['A2'] = f'Generato: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'] = 'NB: nessuna scrittura su Odoo. Solo simulazione su file locali.'

    if oda_state:
        ws['A5'] = f'OdA P03718: stato={oda_state["state"]} - invoice_status={oda_state["invoice_status"]}'
        ws['A6'] = f'Righe POL totali: {oda_state["n_lines_total"]}, libere: {oda_state["n_lines_libere"]}'
        ws['A7'] = f'Importo totale: € {oda_state["amount_total"]:,.2f} (untaxed € {oda_state["amount_untaxed"]:,.2f})'

    headers = ['#', 'PDF', 'Numero fattura', 'Data', 'Cc cliente', 'OdA',
                'Imponibile XML €', 'Tot apparati IVA-incl €']
    row = 9
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = Font(bold=True, color='FFFFFFFF')
        c.fill = title_fill
        c.alignment = align_c
    row += 1
    for i, p in enumerate(previews, start=1):
        if 'error' in p and 'header' not in p:
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=Path(p['pdf_path']).name)
            ws.cell(row=row, column=3, value=f"ERRORE: {p['error']}")
            row += 1
            continue
        h = p.get('header', {})
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=Path(p['pdf_path']).name)
        ws.cell(row=row, column=3, value=h.get('numero', ''))
        ws.cell(row=row, column=4, value=h.get('data', ''))
        ws.cell(row=row, column=5, value=p.get('cc', ''))
        ws.cell(row=row, column=6, value=p.get('oda_name', ''))
        ws.cell(row=row, column=7, value=p.get('imponibile_xml') or 0).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=p.get('split', {}).get('totale_iva_inclusa_pdf', 0)).number_format = '#,##0.00'
        row += 1

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 22


def fill_detail(ws, idx: int, p: dict):
    """Foglio dettaglio per singola fattura."""
    title_font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    title_fill = PatternFill(start_color='FF1A3A5C', end_color='FF1A3A5C', fill_type='solid')
    sect_fill = PatternFill(start_color='FFE0E8F0', end_color='FFE0E8F0', fill_type='solid')
    sect_font = Font(bold=True, color='FF1A3A5C')
    bold = Font(bold=True)
    warn_fill = PatternFill(start_color='FFFFF4E0', end_color='FFFFF4E0', fill_type='solid')

    h = p.get('header', {})

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Fattura #{idx}: {h.get('numero','?')} cc={p.get('cc','?')}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center')

    if 'error' in p and 'header' not in p:
        ws['A3'] = f"ERRORE: {p['error']}"
        ws['A3'].font = Font(bold=True, color='FFCC0000')
        return

    # Header dati fattura
    rows_h = [
        ('PDF sorgente', Path(p['pdf_path']).name),
        ('Numero', h.get('numero', '')),
        ('Data fattura', h.get('data', '')),
        ('Codice cliente', p.get('cc', '')),
        ('cc_type', p.get('cc_type', '')),
        ('OdA mappato', p.get('oda_name', '')),
        ('Imponibile XML', f"€ {p.get('imponibile_xml') or 0:,.2f}"),
        ('Totale apparati IVA-incl (PDF)', f"€ {p['split']['totale_iva_inclusa_pdf']:,.2f}"),
    ]
    r = 3
    for label, val in rows_h:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1

    # Dettaglio apparati
    ws.cell(row=r, column=1, value='DETTAGLIO APPARATI (estratti dal PDF)').font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    r += 1
    headers_app = ['Tipo', 'Codice apparato', '# movimenti', '€ IVA-incl',
                    'Classificazione', 'Targa', 'Veicolo']
    for i, hh in enumerate(headers_app):
        c = ws.cell(row=r, column=i+1, value=hh)
        c.font = bold
        c.fill = sect_fill
    r += 1

    split = p['split']
    all_app = (
        [(a, 'furgoni') for a in split['apparati_furgoni']] +
        [(a, 'uso_promiscuo') for a in split['apparati_promiscuo']] +
        [(a, 'NON MAPPATO') for a in split['apparati_non_mappati']]
    )
    for app, cls in all_app:
        info = get_apparato_info(app.apparato_id) or {}
        ws.cell(row=r, column=1, value=app.tipo)
        ws.cell(row=r, column=2, value=app.apparato_id)
        ws.cell(row=r, column=3, value=app.n_movimenti)
        ws.cell(row=r, column=4, value=app.importo_iva_inclusa).number_format = '#,##0.00'
        cell_cls = ws.cell(row=r, column=5, value=cls)
        if cls == 'NON MAPPATO':
            cell_cls.fill = warn_fill
            cell_cls.font = Font(bold=True, color='FFAA0000')
        ws.cell(row=r, column=6, value=info.get('targa', ''))
        ws.cell(row=r, column=7, value=info.get('veicolo_descrizione', ''))
        r += 1

    r += 1

    # Move lines preview
    ws.cell(row=r, column=1, value='RIGHE BOZZA (account.move) CHE L\'AGENT CREEREBBE').font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    r += 1
    move_headers = ['#', 'Descrizione', 'Conto contabile', 'Quantità',
                     'Prezzo unitario €', 'IVA', 'Subtotale €']
    for i, hh in enumerate(move_headers):
        c = ws.cell(row=r, column=i+1, value=hh)
        c.font = bold
        c.fill = sect_fill
    r += 1
    tot = 0.0
    for i, ml in enumerate(p.get('move_lines', []), start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=ml['name'])
        ws.cell(row=r, column=3, value=ml['account_id_label'])
        ws.cell(row=r, column=4, value=ml['quantity'])
        ws.cell(row=r, column=5, value=ml['price_unit']).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=ml['tax_label'])
        sub = ml['quantity'] * ml['price_unit']
        ws.cell(row=r, column=7, value=sub).number_format = '#,##0.00'
        tot += sub
        r += 1
    ws.cell(row=r, column=6, value='Totale imponibile').font = bold
    ws.cell(row=r, column=7, value=tot).number_format = '#,##0.00'
    ws.cell(row=r, column=7).font = bold
    r += 1
    ws.cell(row=r, column=6, value='IVA 22%').font = bold
    iva_calc = tot * 0.22
    ws.cell(row=r, column=7, value=iva_calc).number_format = '#,##0.00'
    r += 1
    ws.cell(row=r, column=6, value='Totale fattura').font = bold
    ws.cell(row=r, column=7, value=tot + iva_calc).number_format = '#,##0.00'
    ws.cell(row=r, column=7).font = bold

    r += 2

    # Metadati move
    ws.cell(row=r, column=1, value='METADATI BOZZA').font = sect_font
    ws.cell(row=r, column=1).fill = sect_fill
    r += 1
    meta = [
        ('move_type', 'in_invoice'),
        ('partner_id Odoo', p.get('partner_id')),
        ('currency_id', p.get('currency_id')),
        ('invoice_origin', p.get('oda_name')),
        ('ref', h.get('numero', '')),
        ('invoice_date', h.get('data', '')),
        ('date (contabile)', p.get('date_contabile')),
        ('l10n_it_vat_settlement_date (competenza IVA)', p.get('date_iva_competenza')),
        ('journal_id', '2 (acquisti Ecotel)'),
        ('company_id', '1 (Ecotel Italia)'),
    ]
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=str(val) if val is not None else '')
        r += 1

    r += 1

    # Warnings split
    if split.get('warnings'):
        ws.cell(row=r, column=1, value='AVVISI / WARNINGS').font = sect_font
        ws.cell(row=r, column=1).fill = warn_fill
        r += 1
        for w in split['warnings']:
            cell = ws.cell(row=r, column=1, value=w)
            cell.fill = warn_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            r += 1

    # Layout
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--input-dir', type=Path,
                          default=PROJECT_ROOT / 'input',
                          help='Cartella con i PDF Autostrade')
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT,
                          help='Path Excel output')
    args = parser.parse_args()

    pdfs = sorted(args.input_dir.glob('ft *.pdf'))
    if not pdfs:
        print(f'Nessun PDF "ft *.pdf" in {args.input_dir}', file=sys.stderr)
        sys.exit(1)
    print(f'Trovati {len(pdfs)} PDF')

    print('Connessione Odoo (read-only) per stato P03718...')
    try:
        oda_state = query_oda_state('P03718')
        print(f'  P03718: {oda_state["n_lines_libere"]} righe libere su {oda_state["n_lines_total"]}')
    except Exception as e:
        print(f'  [warn] Odoo non raggiungibile: {e}')
        oda_state = None

    previews = []
    for pdf in pdfs:
        print(f'\n{pdf.name}')
        try:
            p = simulate_one(pdf, oda_state)
            previews.append(p)
            if 'error' in p and 'header' not in p:
                print(f'  ERROR: {p["error"]}')
                continue
            split = p['split']
            print(f'  cc={p["cc"]} oda={p["oda_name"]} imponibile_xml=€{p["imponibile_xml"]:,.2f}')
            print(f'  apparati: {len(split["apparati_furgoni"])} furgoni + '
                  f'{len(split["apparati_promiscuo"])} promiscuo + '
                  f'{len(split["apparati_non_mappati"])} non mappati')
            print(f'  split: furgoni € {split["imponibile_furgoni"]:,.2f}, '
                  f'promiscuo € {split["imponibile_promiscuo"]:,.2f}')
            if split['apparati_non_mappati']:
                for a in split['apparati_non_mappati']:
                    print(f'    [!] non mappato: {a.tipo} {a.apparato_id} '
                          f'(€{a.importo_iva_inclusa:.2f})')
        except Exception as e:
            import traceback
            traceback.print_exc()
            previews.append({'pdf_path': str(pdf), 'error': str(e)})

    # Build Excel
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Riepilogo'
    fill_summary(ws, previews, oda_state)

    if oda_state:
        ws_oda = wb.create_sheet('Stato OdA P03718')
        fill_oda_state(ws_oda, oda_state, previews)

    for i, p in enumerate(previews, start=1):
        h = p.get('header', {})
        num = (h.get('numero') or 'sconosciuto')
        # Excel non ammette ? * [ ] : / \ nei nomi foglio
        safe_num = re.sub(r'[\?\*\[\]\:/\\]', '', num)
        sheet_name = f"Fattura {i} - {safe_num[-8:]}"[:31]
        ws_d = wb.create_sheet(sheet_name)
        fill_detail(ws_d, i, p)

    wb.save(args.output)
    print(f'\nGenerato: {args.output}')
    print(f'Dimensione: {args.output.stat().st_size:,} bytes')


if __name__ == '__main__':
    main()
