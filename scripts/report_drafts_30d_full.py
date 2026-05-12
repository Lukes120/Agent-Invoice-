"""
Report REALTIME Odoo - ultimi 30 giorni - account.move (in_invoice + in_refund).

1. Vista totale: chi ha creato bozze (qualunque utente)
2. Per ogni creatore: stato attuale (draft/posted/cancel) + chi ha posted
3. Zoom agent: solo move con fatturapa_attachment_in_id collegato
4. Export Excel multi-sheet in output/

Read-only.
"""
import os
import sys
from datetime import datetime, timedelta
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(ROOT, 'config', 'credentials.env'))

from core.odoo_client import OdooReadOnlyClient

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import base64
import re


def fmt_pct(num, den):
    return (100.0 * num / den) if den else 0.0


_TD_RE = re.compile(rb'<TipoDocumento>\s*(TD\d{2})\s*</TipoDocumento>')


def extract_tipo_documento(xml_bytes):
    """Cerca il primo <TipoDocumento>TDxx</TipoDocumento> nell'XML."""
    if not xml_bytes:
        return None
    m = _TD_RE.search(xml_bytes)
    return m.group(1).decode('ascii') if m else None


def main():
    cli = OdooReadOnlyClient(
        url=os.getenv('ODOO_URL'),
        db=os.getenv('ODOO_DB'),
        username=os.getenv('ODOO_USERNAME'),
        password=os.getenv('ODOO_PASSWORD'),
    )
    cli.connect()
    print(f'Connesso. Mio uid={cli.uid}')

    # campo fatturaPA disponibile?
    fields_info = cli._call('account.move', 'fields_get', [], attributes=['string'])
    fatturapa_field = 'fatturapa_attachment_in_id' if 'fatturapa_attachment_in_id' in fields_info else None
    print(f'Campo fatturaPA: {fatturapa_field}')

    cutoff = '2026-04-21 00:00:00'
    print(f'Cutoff create_date >= {cutoff} (fino a oggi)\n')

    # ----------------------------------------------------
    # 1) Tutti i move passivi creati negli ultimi 30 giorni
    # ----------------------------------------------------
    domain = [
        ('create_date', '>=', cutoff),
        ('move_type', 'in', ['in_invoice', 'in_refund']),
    ]
    fields = ['id', 'name', 'state', 'move_type', 'invoice_date', 'date',
              'amount_total', 'partner_id', 'invoice_origin',
              'company_id',
              'create_uid', 'create_date', 'write_uid', 'write_date']
    if fatturapa_field:
        fields.append(fatturapa_field)

    moves = cli._call('account.move', 'search_read', domain,
                      fields=fields, order='create_date asc')
    print(f'Totale move passivi (in_invoice + in_refund) ultimi 30gg: {len(moves)}')

    # ----------------------------------------------------
    # 2) Aggregazioni
    # ----------------------------------------------------
    # totali per creator
    by_creator = defaultdict(lambda: Counter())          # creator_name -> Counter(state)
    by_creator_td = defaultdict(lambda: defaultdict(Counter))  # creator -> td -> Counter(state)
    posted_by_writer_per_creator = defaultdict(Counter)  # creator -> Counter(write_uid_name)

    # zoom agent
    agent_moves = []
    for m in moves:
        creator = (m.get('create_uid') or [None, '?'])[1]
        td = 'TD04 (NC)' if m.get('move_type') == 'in_refund' else 'TD01/24/25 (FT)'
        st = m.get('state')
        by_creator[creator][st] += 1
        by_creator_td[creator][td][st] += 1
        if st == 'posted':
            wu = (m.get('write_uid') or [None, '?'])[1]
            posted_by_writer_per_creator[creator][wu] += 1
        if fatturapa_field and m.get(fatturapa_field):
            agent_moves.append(m)

    # ----------------------------------------------------
    # 2b) Per i move dell'agent estraggo il TipoDocumento dall'XML
    #     così possiamo splittare TD01 vs TD24 vs TD25
    # ----------------------------------------------------
    print(f'\nEstrazione TipoDocumento da XML per {len(agent_moves)} move agent...')
    att_ids = list({m[fatturapa_field][0] for m in agent_moves
                    if fatturapa_field and m.get(fatturapa_field)})
    att_data = {}
    if att_ids:
        rows = cli._call('fatturapa.attachment.in', 'read', att_ids,
                         fields=['id', 'datas'])
        for r in rows:
            att_data[r['id']] = r.get('datas')

    move_td_real = {}  # move_id -> TipoDocumento (TD01/04/24/25/...)
    for m in agent_moves:
        att_id = m[fatturapa_field][0]
        b64 = att_data.get(att_id)
        td_real = None
        if b64:
            try:
                xml_bytes = base64.b64decode(b64)
                td_real = extract_tipo_documento(xml_bytes)
            except Exception:
                td_real = None
        if not td_real:
            td_real = 'TD04' if m.get('move_type') == 'in_refund' else 'TD01'
        move_td_real[m['id']] = td_real
    print('Estrazione completata.\n')

    # ordino i creator per totale desc
    creators_sorted = sorted(by_creator.items(),
                             key=lambda kv: -sum(kv[1].values()))

    # ----------------------------------------------------
    # 3) Stampa report
    # ----------------------------------------------------
    print()
    print('=' * 78)
    print(f'TOTALE MOVE CREATI ULTIMI 30 GIORNI: {len(moves)}')
    print('  (in_invoice + in_refund, qualunque utente, qualunque flusso)')
    print('=' * 78)
    print()
    print(f'{"Creatore":<35} {"Tot":>5}  {"Draft":>6} {"Posted":>7} {"Cancel":>7}')
    print('-' * 78)
    for creator, states in creators_sorted:
        tot = sum(states.values())
        d = states.get('draft', 0)
        p = states.get('posted', 0)
        c = states.get('cancel', 0)
        print(f'{creator:<35} {tot:>5}  {d:>6} {p:>7} {c:>7}')
    print('-' * 78)
    tot_all = len(moves)
    tot_draft = sum(s.get('draft', 0) for s in by_creator.values())
    tot_posted = sum(s.get('posted', 0) for s in by_creator.values())
    tot_cancel = sum(s.get('cancel', 0) for s in by_creator.values())
    print(f'{"TOTALE":<35} {tot_all:>5}  {tot_draft:>6} {tot_posted:>7} {tot_cancel:>7}')
    print(f'                                              ({fmt_pct(tot_draft,tot_all):.1f}%)  ({fmt_pct(tot_posted,tot_all):.1f}%)  ({fmt_pct(tot_cancel,tot_all):.1f}%)')

    print()
    print('=' * 78)
    print('DETTAGLIO PER CREATORE (TD01 fatture vs TD04 note credito)')
    print('=' * 78)
    for creator, _ in creators_sorted:
        sub = by_creator_td[creator]
        print(f'\n{creator}')
        for td in sorted(sub.keys()):
            s = sub[td]
            tot = sum(s.values())
            d = s.get('draft', 0)
            p = s.get('posted', 0)
            c = s.get('cancel', 0)
            print(f'  {td:6s}  totale={tot:4d}  draft={d:4d}  posted={p:4d}  cancel={c:4d}')

        # chi ha posted le bozze creati da questo creator
        writers = posted_by_writer_per_creator[creator]
        if writers:
            tot_p = sum(writers.values())
            print(f'  posted da:')
            for wu, cnt in writers.most_common():
                print(f'    {wu:<35} {cnt:>4}  ({fmt_pct(cnt,tot_p):.1f}%)')

    # ----------------------------------------------------
    # 4) Zoom agent (solo move con fatturapa)
    # ----------------------------------------------------
    print()
    print('=' * 78)
    print(f'ZOOM AGENT - solo move con fatturaPA collegata: {len(agent_moves)}')
    print('=' * 78)
    ag_state = Counter()
    ag_td = defaultdict(Counter)
    ag_writer = Counter()
    for m in agent_moves:
        st = m.get('state')
        td = move_td_real.get(m['id']) or ('TD04' if m.get('move_type') == 'in_refund' else 'TD01')
        ag_state[st] += 1
        ag_td[td][st] += 1
        if st == 'posted':
            wu = (m.get('write_uid') or [None, '?'])[1]
            ag_writer[wu] += 1
    print()
    print(f'Totale: {len(agent_moves)}')
    for st, cnt in ag_state.most_common():
        print(f'  {st:10s}: {cnt:4d}  ({fmt_pct(cnt,len(agent_moves)):.1f}%)')
    print()
    print('  Split per TipoDocumento (TD01/TD04/TD24/TD25...):')
    for td in sorted(ag_td.keys()):
        s = ag_td[td]
        tot = sum(s.values())
        d = s.get('draft', 0)
        p = s.get('posted', 0)
        c = s.get('cancel', 0)
        print(f'  {td:6s}  totale={tot:4d}  draft={d:4d}  posted={p:4d}  cancel={c:4d}')
    print()
    print('Posted dell\'agent confermati da:')
    tot_p = sum(ag_writer.values())
    for wu, cnt in ag_writer.most_common():
        print(f'  {wu:<35} {cnt:>4}  ({fmt_pct(cnt,tot_p):.1f}%)')

    # ----------------------------------------------------
    # 5) Export Excel
    # ----------------------------------------------------
    out_dir = os.path.join(ROOT, 'output')
    os.makedirs(out_dir, exist_ok=True)
    out_xlsx = os.path.join(out_dir, f'report_drafts_30d_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

    wb = Workbook()

    # --- Sheet 1: Riepilogo per creatore
    ws = wb.active
    ws.title = 'Riepilogo creatori'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='305496')

    ws.append(['Creatore', 'Totale', 'Draft', 'Posted', 'Cancel',
               'FT (TD01/24/25) tot', 'FT draft', 'FT posted',
               'NC (TD04) tot', 'NC draft', 'NC posted'])
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    KEY_FT = 'TD01/24/25 (FT)'
    KEY_NC = 'TD04 (NC)'
    for creator, states in creators_sorted:
        sub = by_creator_td[creator]
        ft = sub.get(KEY_FT, Counter())
        nc = sub.get(KEY_NC, Counter())
        ws.append([
            creator,
            sum(states.values()),
            states.get('draft', 0),
            states.get('posted', 0),
            states.get('cancel', 0),
            sum(ft.values()),
            ft.get('draft', 0),
            ft.get('posted', 0),
            sum(nc.values()),
            nc.get('draft', 0),
            nc.get('posted', 0),
        ])
    # totale
    ws.append([
        'TOTALE',
        tot_all, tot_draft, tot_posted, tot_cancel,
        sum(sum(by_creator_td[c].get(KEY_FT, Counter()).values()) for c, _ in creators_sorted),
        sum(by_creator_td[c].get(KEY_FT, Counter()).get('draft', 0) for c, _ in creators_sorted),
        sum(by_creator_td[c].get(KEY_FT, Counter()).get('posted', 0) for c, _ in creators_sorted),
        sum(sum(by_creator_td[c].get(KEY_NC, Counter()).values()) for c, _ in creators_sorted),
        sum(by_creator_td[c].get(KEY_NC, Counter()).get('draft', 0) for c, _ in creators_sorted),
        sum(by_creator_td[c].get(KEY_NC, Counter()).get('posted', 0) for c, _ in creators_sorted),
    ])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='D9E1F2')

    # --- Sheet 2: Posted by writer per creatore
    ws2 = wb.create_sheet('Chi ha posted')
    ws2.append(['Creatore bozza', 'Posted da', 'Numero', '% sui posted del creatore'])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
    for creator, _ in creators_sorted:
        writers = posted_by_writer_per_creator[creator]
        tot_p = sum(writers.values())
        if not writers:
            ws2.append([creator, '(nessuno ancora posted)', 0, 0])
            continue
        for wu, cnt in writers.most_common():
            ws2.append([creator, wu, cnt, round(fmt_pct(cnt, tot_p), 1)])

    # --- Sheet 3: Dettaglio agent
    ws3 = wb.create_sheet('Dettaglio agent')
    ws3.append(['ID', 'Numero', 'Move type', 'TipoDoc XML', 'Partner', 'Importo',
                'Data fattura', 'Stato', 'Creato da', 'Creato il', 'Posted da', 'Posted il',
                'OdA (origin)'])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
    for m in agent_moves:
        td_real = move_td_real.get(m['id'], '')
        partner = (m.get('partner_id') or [None, ''])[1]
        creator = (m.get('create_uid') or [None, ''])[1]
        wu = (m.get('write_uid') or [None, ''])[1] if m.get('state') == 'posted' else ''
        wd = m.get('write_date') if m.get('state') == 'posted' else ''
        ws3.append([
            m['id'], m.get('name') or '', m.get('move_type'), td_real,
            partner, m.get('amount_total'),
            m.get('invoice_date'), m.get('state'),
            creator, m.get('create_date'), wu, wd,
            m.get('invoice_origin') or '',
        ])

    # --- Sheet 4: Tutti i move (raw)
    ws4 = wb.create_sheet('Tutti i move')
    ws4.append(['ID', 'Numero', 'Tipo', 'Partner', 'Importo', 'Stato',
                'Creato da', 'Creato il', 'Posted/last writer', 'Last write',
                'OdA', 'FatturaPA?'])
    for c in ws4[1]:
        c.font = header_font
        c.fill = header_fill
    for m in moves:
        partner = (m.get('partner_id') or [None, ''])[1]
        creator = (m.get('create_uid') or [None, ''])[1]
        wu = (m.get('write_uid') or [None, ''])[1]
        ws4.append([
            m['id'], m.get('name') or '', m.get('move_type'),
            partner, m.get('amount_total'), m.get('state'),
            creator, m.get('create_date'),
            wu, m.get('write_date'),
            m.get('invoice_origin') or '',
            'SI' if (fatturapa_field and m.get(fatturapa_field)) else 'no',
        ])

    # column widths
    for ws_i in [ws, ws2, ws3, ws4]:
        for col_idx, col_cells in enumerate(ws_i.columns, 1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws_i.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 45)

    wb.save(out_xlsx)
    print(f'\nExcel salvato in: {out_xlsx}')


if __name__ == '__main__':
    main()
