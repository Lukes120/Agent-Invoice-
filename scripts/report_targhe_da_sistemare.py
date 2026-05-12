"""Report Excel delle targhe da sistemare in Parco Auto.

Scansiona ultimi 3 mesi di fatture sui 7 fornitori automezzi.
Per ogni riga XML con targa, classifica:
  - voce e cls secondo l'agent oggi (PARCO_BY_TARGA / PARCO_BY_CONTRATTO)
  - cls deducibile dal name della POL pre-pianificata su OdA (fonte Acquisti)

Anomalie segnalate (3 fogli):
  1. Targhe mancanti nel Parco Auto (cls_source = default_unknown)
  2. Discordanze: PARCO dice X, name POL dice Y -> presumibile errore PARCO
  3. Targhe nuove non viste (compaiono in XML ma non sono in PARCO_BY_TARGA)

Output: output/targhe_da_sistemare_<YYYYMMDD_HHMMSS>.xlsx
"""
import os
import sys
import io
import re
import logging
from datetime import datetime, timedelta
from collections import defaultdict
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
logging.basicConfig(level=logging.WARNING, format='%(message)s')
load_dotenv('config/credentials.env')

from core.odoo_rw_client import OdooReadWriteClient
from core.matcher import InvoiceMatcher
from core.fatturapa_analyzer import FatturaPAAnalyzer
from core.odoo_writer import OdooWriter
from core.keyword_rules import classify_line_by_keyword
from core.fatturapa_parser import parse_from_base64
from config.rules import (
    TOLLERANZA_PERCENTUALE, TOLLERANZA_ASSOLUTA, TOLLERANZA_TOTALE_FATTURA,
    MATCH_IMPLICITO_ATTIVO, TOLLERANZA_MATCH_IMPLICITO_PERCENT,
    TOLLERANZA_MATCH_IMPLICITO_ASSOLUTA,
    TOLLERANZA_MATCH_IMPLICITO_LARGA_ASSOLUTA,
    TOLLERANZA_MATCH_IMPLICITO_LARGA_PERCENT,
    MATCH_PARZIALE_ATTIVO, SUGGERIMENTI_ATTIVI,
    MAPPATURA_FORNITORI_FISSI_ATTIVA, MAPPATURA_FORNITORI_FISSI,
    MAPPATURA_AUTOMEZZI,
)
from config.parco_auto_mapping import (
    PARCO_BY_TARGA, get_classificazione_by_targa,
    get_classificazione_by_contratto,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

VATS = [
    ('IT06714021000', 'Leasys'),
    ('IT01924961004', 'ALD'),
    ('IT04911190488', 'Arval'),
    ('IT03740811207', 'UnipolRental'),
    ('IT05580391000', 'Tecnoalt'),
    ('IT10641441000', 'Athlon'),
    ('IT02615080963', 'LeasePlan'),
]

CUTOFF = (datetime.now() - timedelta(days=92)).strftime('%Y-%m-%d')


def fetch_attachments(client, vat, label):
    """Scarica fatture in batch piccoli (evita timeout XML-RPC)."""
    import time
    out = []
    offset = 0
    batch = 40
    while True:
        try:
            res = client._call(
                'fatturapa.attachment.in', 'search_read',
                [('xml_supplier_id.vat', '=', vat),
                 ('create_date', '>=', CUTOFF)],
                fields=['id', 'name', 'datas', 'invoices_total',
                        'invoices_date', 'registered', 'xml_supplier_id'],
                limit=batch, offset=offset, order='create_date desc')
        except Exception as e:
            print(f"    !! errore batch {offset}: {e}; retry")
            time.sleep(1)
            try:
                res = client._call(
                    'fatturapa.attachment.in', 'search_read',
                    [('xml_supplier_id.vat', '=', vat),
                     ('create_date', '>=', CUTOFF)],
                    fields=['id', 'name', 'datas', 'invoices_total',
                            'invoices_date', 'registered', 'xml_supplier_id'],
                    limit=batch, offset=offset, order='create_date desc')
            except Exception as e2:
                print(f"    !! retry fallito: {e2}; skip rest")
                break
        if not res:
            break
        for r in res:
            r['_vat'] = vat
            r['_label'] = label
        out.extend(res)
        if len(res) < batch:
            break
        offset += batch
    return out


def main():
    client = OdooReadWriteClient(
        os.getenv('ODOO_URL'), os.getenv('ODOO_DB'),
        os.getenv('ODOO_USERNAME'), os.getenv('ODOO_PASSWORD'))
    client.connect()
    matcher = InvoiceMatcher(tol_percent=TOLLERANZA_PERCENTUALE,
                             tol_absolute=TOLLERANZA_ASSOLUTA,
                             tol_total=TOLLERANZA_TOTALE_FATTURA,
                             keyword_classifier=classify_line_by_keyword)
    analyzer = FatturaPAAnalyzer(
        client, matcher, TOLLERANZA_TOTALE_FATTURA,
        implicit_match_enabled=MATCH_IMPLICITO_ATTIVO,
        implicit_match_tolerance_percent=TOLLERANZA_MATCH_IMPLICITO_PERCENT,
        implicit_match_tolerance_absolute=TOLLERANZA_MATCH_IMPLICITO_ASSOLUTA,
        implicit_match_loose_tolerance_absolute=TOLLERANZA_MATCH_IMPLICITO_LARGA_ASSOLUTA,
        implicit_match_loose_tolerance_percent=TOLLERANZA_MATCH_IMPLICITO_LARGA_PERCENT,
        partial_match_enabled=MATCH_PARZIALE_ATTIVO,
        suggestions_enabled=SUGGERIMENTI_ATTIVI,
        supplier_mapping_enabled=MAPPATURA_FORNITORI_FISSI_ATTIVA,
        supplier_mapping=MAPPATURA_FORNITORI_FISSI)
    writer = OdooWriter(client, dry_run=True)

    # Hook log writer per intercettare i log DRY_RUN con dettaglio POL
    log_buf = io.StringIO()
    log_handler = logging.StreamHandler(log_buf)
    log_handler.setLevel(logging.INFO)
    log_handler.setFormatter(logging.Formatter('%(message)s'))
    ow_logger = logging.getLogger('core.odoo_writer')
    ow_logger.addHandler(log_handler)
    ow_logger.setLevel(logging.INFO)

    # Anomalie raccolte
    mancanti = []      # cls_source=default_unknown
    discordanze = []   # cls_PARCO vs cls_POL diversi
    nuove_targhe = defaultdict(lambda: {'count':0, 'tot':0.0, 'first_fattura':'',
                                          'first_date':'', 'fornitore':'',
                                          'sample_desc':''})

    print(f"Scansione fatture ultimi 90gg (>= {CUTOFF})...")
    for vat, label in VATS:
        atts = fetch_attachments(client, vat, label)
        print(f"  {label}: {len(atts)} fatture")
        for a in atts:
            try:
                analysis = analyzer.analyze(a)
            except Exception as e:
                continue
            if analysis.classification != 'MAPPATURA_AUTOMEZZI':
                continue
            mapping_entry = MAPPATURA_AUTOMEZZI.get(vat) or {}
            if not mapping_entry:
                continue
            # Simula dry-run e cattura log per estrarre POL pre-pianificata
            log_buf.truncate(0)
            log_buf.seek(0)
            try:
                result = writer.create_bozza_automezzi(analysis, mapping_entry)
            except Exception as e:
                continue
            if not result.success:
                continue
            log_text = log_buf.getvalue()
            # Parse log per estrarre (po_line_id, voce, cls, targa) consumate
            pol_lines = re.findall(
                r'POL (\d+) on \S+ \((\w+)/(\w+), EUR([0-9.\-]+)->EUR([0-9.\-]+),'
                r' tax(\d+), acc(\d+), targa=([^)]*)\)', log_text)
            # Re-leggi le POL consumate per ottenere il name (Old name dalla
            # POL pre-modifica).
            consumed_pol_ids = [int(p[0]) for p in pol_lines]
            if not consumed_pol_ids:
                continue
            try:
                pols = client._call('purchase.order.line', 'read',
                    consumed_pol_ids, fields=['id', 'name'])
            except Exception:
                continue
            pol_name_map = {p['id']: (p.get('name') or '') for p in pols}

            # Per ogni riga XML, ricostruisci routing
            for r in (analysis.xml_data.righe or []):
                desc = r.descrizione or ''
                voce_xml = OdooWriter._classify_voce_automezzi_full(r)
                targa = OdooWriter._extract_targa_automezzi(r) or ''
                cls_parco, src_parco = writer._resolve_classificazione_veicolo(
                    targa, '', mapping_entry=mapping_entry)
                importo = float(r.prezzo_totale or 0)

                # Targhe nuove: presenti in XML ma non in PARCO_BY_TARGA
                if targa and targa not in PARCO_BY_TARGA:
                    nt = nuove_targhe[targa]
                    nt['count'] += 1
                    nt['tot'] += importo
                    nt['fornitore'] = label
                    if not nt['first_fattura']:
                        nt['first_fattura'] = analysis.xml_data.numero or a['name']
                        nt['first_date'] = a.get('invoices_date') or ''
                        nt['sample_desc'] = desc[:80]

                # MANCANTI: cls_source default_unknown
                if src_parco in ('default_unknown', 'default_no_mapping'):
                    mancanti.append({
                        'fornitore': label,
                        'fattura': analysis.xml_data.numero or a['name'],
                        'data': a.get('invoices_date') or '',
                        'riga': r.numero_linea,
                        'targa': targa,
                        'voce': voce_xml,
                        'desc': desc[:120],
                        'importo': importo,
                        'cls_attuale_agent': cls_parco,
                        'src_attuale': src_parco,
                    })

            # DISCORDANZE: per ogni POL consumata, confronta cls da log
            # (cls_attuale_agent) con cls dedotta dal name POL.
            for (polid, voce_log, cls_log, oldp, newp, tax, acc, targa_log) in pol_lines:
                pol_name = pol_name_map.get(int(polid), '')
                cls_from_pol = OdooWriter._classify_cls_from_pol_name(pol_name)
                if cls_from_pol and cls_log and cls_from_pol != cls_log:
                    discordanze.append({
                        'fornitore': label,
                        'fattura': analysis.xml_data.numero or a['name'],
                        'data': a.get('invoices_date') or '',
                        'targa': targa_log or '(nessuna)',
                        'voce': voce_log,
                        'cls_agent': cls_log,
                        'cls_da_pol_name': cls_from_pol,
                        'pol_name': pol_name[:160],
                        'importo': float(newp),
                    })

    # === Scrittura Excel ===
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(out_dir, f'targhe_da_sistemare_{ts}.xlsx')

    wb = Workbook()
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='305496')
    WARN_FILL = PatternFill('solid', fgColor='FFE699')

    def write_sheet(ws, headers, rows, freeze='A2'):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        # Auto-width approssimato
        for col_idx, h in enumerate(headers, 1):
            max_len = max([len(str(h))] + [len(str(r[col_idx - 1])) for r in rows[:200]] or [10])
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        ws.freeze_panes = freeze

    # Sheet 1: Riepilogo
    ws_sum = wb.active
    ws_sum.title = 'Riepilogo'
    ws_sum['A1'] = 'Report Targhe da Sistemare (ultimi 90gg)'
    ws_sum['A1'].font = Font(bold=True, size=14)
    ws_sum['A2'] = f'Generato: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    ws_sum['A3'] = f'Cutoff: create_date >= {CUTOFF}'
    ws_sum['A5'] = 'Anomalia'
    ws_sum['B5'] = 'Conteggio'
    ws_sum['C5'] = 'Importo totale'
    for c in ('A5', 'B5', 'C5'):
        ws_sum[c].font = HEADER_FONT
        ws_sum[c].fill = HEADER_FILL
    ws_sum['A6'] = 'Targhe mancanti nel Parco Auto'
    ws_sum['B6'] = len(mancanti)
    ws_sum['C6'] = sum(m['importo'] for m in mancanti)
    ws_sum['A7'] = 'Discordanze PARCO vs name POL'
    ws_sum['B7'] = len(discordanze)
    ws_sum['C7'] = sum(d['importo'] for d in discordanze)
    ws_sum['A8'] = 'Targhe nuove non in PARCO_BY_TARGA'
    ws_sum['B8'] = len(nuove_targhe)
    ws_sum['C8'] = sum(v['tot'] for v in nuove_targhe.values())
    ws_sum.column_dimensions['A'].width = 40
    ws_sum.column_dimensions['B'].width = 14
    ws_sum.column_dimensions['C'].width = 18

    # Sheet 2: Mancanti
    ws_m = wb.create_sheet('Mancanti')
    mancanti_sorted = sorted(mancanti, key=lambda x: (x['fornitore'], x['data'], x['fattura']))
    write_sheet(ws_m,
        ['Fornitore', 'Fattura', 'Data', 'Riga', 'Targa', 'Voce', 'Descrizione XML',
         'Importo', 'Cls fallback agent', 'Source'],
        [(m['fornitore'], m['fattura'], m['data'], m['riga'], m['targa'],
          m['voce'], m['desc'], m['importo'], m['cls_attuale_agent'], m['src_attuale'])
         for m in mancanti_sorted])

    # Sheet 3: Discordanze
    ws_d = wb.create_sheet('Discordanze')
    discordanze_sorted = sorted(discordanze, key=lambda x: (x['fornitore'], x['data']))
    write_sheet(ws_d,
        ['Fornitore', 'Fattura', 'Data', 'Targa', 'Voce',
         'Cls agent (PARCO)', 'Cls da name POL', 'POL name', 'Importo'],
        [(d['fornitore'], d['fattura'], d['data'], d['targa'], d['voce'],
          d['cls_agent'], d['cls_da_pol_name'], d['pol_name'], d['importo'])
         for d in discordanze_sorted])

    # Sheet 4: Targhe nuove
    ws_n = wb.create_sheet('Nuove targhe')
    nuove_list = sorted(nuove_targhe.items(), key=lambda kv: (kv[1]['fornitore'], -kv[1]['count']))
    write_sheet(ws_n,
        ['Targa', 'Fornitore', 'Prima fattura', 'Prima data', 'N. occorrenze',
         'Importo totale', 'Esempio descrizione XML'],
        [(targa, v['fornitore'], v['first_fattura'], v['first_date'],
          v['count'], v['tot'], v['sample_desc'])
         for targa, v in nuove_list])

    wb.save(out_path)
    print()
    print(f"Excel salvato: {out_path}")
    print(f"  Mancanti        : {len(mancanti)}")
    print(f"  Discordanze     : {len(discordanze)}")
    print(f"  Targhe nuove    : {len(nuove_targhe)}")


if __name__ == '__main__':
    main()
