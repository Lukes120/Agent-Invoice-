"""
Genera un report Excel sintetico degli apparati Telepass-network
basandosi su config/apparati_mapping.py.

Uso: python scripts/report_apparati.py
Output: output/report_apparati_<TS>.xlsx
"""
import sys
from datetime import datetime
from collections import Counter, defaultdict
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config.apparati_mapping import APPARATI_MAP


def aggregate(apparati):
    total = len(apparati)
    by_tipo = Counter()
    by_cls = Counter()
    by_stato = Counter()
    by_cc = Counter()
    tipo_x_cls = defaultdict(Counter)
    cc_x_cls = defaultdict(Counter)
    cc_x_tipo = defaultdict(Counter)
    for aid, info in apparati.items():
        tipo = (info.get("tipo_apparato") or "N/D").upper()
        cls = info.get("classificazione") or "N/D"
        stato = info.get("stato") or "N/D"
        cc = info.get("cc_cliente") or "N/D"
        by_tipo[tipo] += 1
        by_cls[cls] += 1
        by_stato[stato] += 1
        by_cc[cc] += 1
        tipo_x_cls[tipo][cls] += 1
        cc_x_cls[cc][cls] += 1
        cc_x_tipo[cc][tipo] += 1
    return dict(
        total=total, by_tipo=by_tipo, by_cls=by_cls, by_stato=by_stato,
        by_cc=by_cc, tipo_x_cls=tipo_x_cls, cc_x_cls=cc_x_cls,
        cc_x_tipo=cc_x_tipo,
    )


# Stili
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
TITLE_FILL = PatternFill(start_color="FF1A3A5C", end_color="FF1A3A5C", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
SECT_FONT = Font(bold=True, color="FF1A3A5C", size=11)
SECT_FILL = PatternFill(start_color="FFE0E8F0", end_color="FFE0E8F0", fill_type="solid")
TOT_FONT = Font(bold=True)
TOT_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")


def riepilogo(ws, agg):
    ws.merge_cells("A1:D1")
    ws["A1"] = (f"Apparati Telepass-network — riepilogo "
                f"(generato {datetime.now().strftime('%d/%m/%Y %H:%M')})")
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    r = 3
    cell = ws.cell(row=r, column=1, value=f"TOTALE APPARATI: {agg['total']}")
    cell.font = TOT_FONT
    cell.fill = TOT_FILL
    r += 2

    # Per tipo
    cell = ws.cell(row=r, column=1, value="Per Tipo Apparato")
    cell.font = SECT_FONT
    cell.fill = SECT_FILL
    r += 1
    for i, h in enumerate(["Tipo", "Numero", "%"], start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r += 1
    for tipo, n in sorted(agg["by_tipo"].items()):
        ws.cell(row=r, column=1, value=tipo)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=n / agg["total"]).number_format = "0.0%"
        r += 1

    r += 1
    # Per classificazione
    cell = ws.cell(row=r, column=1, value="Per Classificazione (deducibilita fiscale)")
    cell.font = SECT_FONT
    cell.fill = SECT_FILL
    r += 1
    for i, h in enumerate(["Classificazione", "Numero", "%"], start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r += 1
    for cls, n in sorted(agg["by_cls"].items(), key=lambda x: -x[1]):
        suffix = ""
        if cls == "furgoni":
            suffix = " (100%)"
        elif cls == "uso_promiscuo":
            suffix = " (70%)"
        ws.cell(row=r, column=1, value=cls + suffix)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=n / agg["total"]).number_format = "0.0%"
        r += 1

    r += 1
    # Per stato
    cell = ws.cell(row=r, column=1, value="Per Stato")
    cell.font = SECT_FONT
    cell.fill = SECT_FILL
    r += 1
    for i, h in enumerate(["Stato", "Numero"], start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r += 1
    for stato, n in sorted(agg["by_stato"].items()):
        ws.cell(row=r, column=1, value=stato)
        ws.cell(row=r, column=2, value=n)
        r += 1

    r += 1
    # Per cc cliente
    cell = ws.cell(row=r, column=1, value="Per Codice Cliente Ecotel")
    cell.font = SECT_FONT
    cell.fill = SECT_FILL
    r += 1
    for i, h in enumerate(["cc cliente", "Numero", "%"], start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    r += 1
    for cc, n in sorted(agg["by_cc"].items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=cc)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=n / agg["total"]).number_format = "0.0%"
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12


def crosstab(ws, title, row_labels_dict, col_labels, get_count, totals_row, totals_col):
    """Genera una crosstab in foglio. row_labels_dict ha keys da iterare, get_count(row, col)."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_labels) + 2)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    r = 3
    c0 = ws.cell(row=r, column=1, value="")
    c0.font = HDR_FONT
    c0.fill = HDR_FILL
    for i, lbl in enumerate(col_labels, start=2):
        c = ws.cell(row=r, column=i, value=lbl)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    c = ws.cell(row=r, column=2 + len(col_labels), value="TOT")
    c.font = HDR_FONT
    c.fill = HDR_FILL
    r += 1

    for row_key in sorted(row_labels_dict.keys()):
        ws.cell(row=r, column=1, value=row_key).font = TOT_FONT
        rsum = 0
        for i, col in enumerate(col_labels, start=2):
            n = get_count(row_key, col)
            if n:
                ws.cell(row=r, column=i, value=n)
            rsum += n
        c = ws.cell(row=r, column=2 + len(col_labels), value=rsum)
        c.font = TOT_FONT
        r += 1

    # Riga totale
    cell = ws.cell(row=r, column=1, value="TOTALE")
    cell.font = TOT_FONT
    cell.fill = TOT_FILL
    for i, col in enumerate(col_labels, start=2):
        c = ws.cell(row=r, column=i, value=totals_col.get(col, 0))
        c.font = TOT_FONT
        c.fill = TOT_FILL
    c = ws.cell(row=r, column=2 + len(col_labels), value=totals_row)
    c.font = TOT_FONT
    c.fill = TOT_FILL

    ws.column_dimensions["A"].width = 18
    for i in range(2, 2 + len(col_labels) + 1):
        ws.column_dimensions[chr(64 + i)].width = 16


def lista_completa(ws, apparati):
    ws.merge_cells("A1:G1")
    ws["A1"] = "Lista completa apparati"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Codice apparato", "Tipo", "cc cliente", "Targa",
               "Classificazione", "Stato", "Veicolo descrizione"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL

    r = 4
    for aid in sorted(apparati.keys()):
        info = apparati[aid]
        ws.cell(row=r, column=1, value=aid)
        ws.cell(row=r, column=2, value=info.get("tipo_apparato", ""))
        ws.cell(row=r, column=3, value=info.get("cc_cliente", ""))
        ws.cell(row=r, column=4, value=info.get("targa", ""))
        ws.cell(row=r, column=5, value=info.get("classificazione", ""))
        ws.cell(row=r, column=6, value=info.get("stato", ""))
        ws.cell(row=r, column=7, value=(info.get("veicolo_descrizione") or "")[:60])
        r += 1

    widths = [16, 10, 14, 12, 16, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def main():
    agg = aggregate(APPARATI_MAP)
    wb = Workbook()

    # Foglio 1
    ws = wb.active
    ws.title = "Riepilogo"
    riepilogo(ws, agg)

    # Foglio 2: Tipo x Classificazione
    cls_keys = sorted(agg["by_cls"].keys())
    ws2 = wb.create_sheet("Tipo per Classificazione")
    crosstab(
        ws2,
        title="Tipo apparato per Classificazione",
        row_labels_dict=agg["tipo_x_cls"],
        col_labels=cls_keys,
        get_count=lambda r, c: agg["tipo_x_cls"][r][c],
        totals_row=agg["total"],
        totals_col=agg["by_cls"],
    )

    # Foglio 3: cc x Classificazione
    ws3 = wb.create_sheet("cc per Classificazione")
    crosstab(
        ws3,
        title="Codice Cliente per Classificazione",
        row_labels_dict=agg["cc_x_cls"],
        col_labels=cls_keys,
        get_count=lambda r, c: agg["cc_x_cls"][r][c],
        totals_row=agg["total"],
        totals_col=agg["by_cls"],
    )

    # Foglio 4: cc x Tipo
    tipo_keys = sorted(agg["by_tipo"].keys())
    ws4 = wb.create_sheet("cc per Tipo")
    crosstab(
        ws4,
        title="Codice Cliente per Tipo Apparato",
        row_labels_dict=agg["cc_x_tipo"],
        col_labels=tipo_keys,
        get_count=lambda r, c: agg["cc_x_tipo"][r][c],
        totals_row=agg["total"],
        totals_col=agg["by_tipo"],
    )

    # Foglio 5: lista completa
    ws5 = wb.create_sheet("Lista completa")
    lista_completa(ws5, APPARATI_MAP)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = PROJECT / "output" / f"report_apparati_{ts}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Generato: {out}")
    print(f"Dimensione: {out.stat().st_size:,} bytes")
    print(f"Fogli: {wb.sheetnames}")
    print()
    print(f"Totale apparati: {agg['total']}")
    print(f"Per tipo: {dict(agg['by_tipo'])}")
    print(f"Per classificazione: {dict(agg['by_cls'])}")
    print(f"Per stato: {dict(agg['by_stato'])}")
    print(f"Per cc cliente: {dict(agg['by_cc'])}")


if __name__ == "__main__":
    main()
