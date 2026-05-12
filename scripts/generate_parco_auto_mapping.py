"""
Generator idempotente di config/parco_auto_mapping.py partendo da
input/Parco Auto.xlsx.

File sorgente atteso (foglio "Parco Auto"):
  | MODELLO | Targa | FORNITORE | POOL/USO PROMISCUO | Numero CONTRATTO |
    DATA INIZIO | DATA FINE | CANONE | [STATO] | [CLASSIFICAZIONE FISCALE] |

Le ultime 2 colonne sono OPZIONALI (forward-compat):
  - STATO: ATTIVO / DISMESSO (se manca: dedotto da DATA FINE vs oggi)
  - CLASSIFICAZIONE FISCALE: POOL / uso_promiscuo / super_lusso
    (se manca: dedotta da col D, "POOL" testuale -> POOL,
    altrimenti -> uso_promiscuo)

Genera config/parco_auto_mapping.py con:
  - PARCO_BY_TARGA: dict[targa, info]
  - PARCO_BY_CONTRATTO: dict[numero_contratto, info]
  - Helper: get_classificazione_by_targa, get_info_by_contratto

Uso:
    python scripts/generate_parco_auto_mapping.py
    python scripts/generate_parco_auto_mapping.py --print-summary

Lo script è READ-ONLY su filesystem (eccetto il file di output).
"""
import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRORE: openpyxl non installato (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "input" / "Parco Auto.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "config" / "parco_auto_mapping.py"

CLASSIFICAZIONI_AMMESSE = {"POOL", "uso_promiscuo", "super_lusso"}
STATI_AMMESSI = {"ATTIVO", "DISMESSO"}

TARGA_PATTERN = re.compile(r"\b([A-Z]{2}\d{3}[A-Z]{2})\b")


def normalize_targa(raw):
    if not raw:
        return ""
    s = str(raw).upper().strip()
    m = TARGA_PATTERN.search(s)
    return m.group(1) if m else s


def normalize_contratto(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    return s


def normalize_classificazione(raw, fallback_pool_promiscuo=None):
    """Normalizza 'classificazione fiscale' a POOL/uso_promiscuo/super_lusso.

    Se raw non valorizzato e fallback_pool_promiscuo fornito (è il valore
    della colonna D 'POOL/USO PROMISCUO'), deduce: testo 'POOL' -> POOL,
    altrimenti uso_promiscuo (presunzione: nome persona = veicolo
    promiscuo).
    """
    s = (str(raw) if raw is not None else "").strip()
    if not s and fallback_pool_promiscuo is not None:
        f = str(fallback_pool_promiscuo).strip().upper()
        if f == "POOL":
            return "POOL"
        if f:  # nome persona
            return "uso_promiscuo"
        return "uso_promiscuo"  # default conservativo
    s_lower = s.lower().replace(" ", "_")
    if s_lower in ("pool", "furgoni", "furgone"):
        return "POOL"
    if s_lower in ("uso_promiscuo", "promiscuo", "uso_promiscui"):
        return "uso_promiscuo"
    if s_lower in ("super_lusso", "superlusso", "super-lusso", "lusso"):
        return "super_lusso"
    # fallback su valore originale (segnala anomalia)
    return s or "uso_promiscuo"


def normalize_stato(raw, data_fine=None):
    s = (str(raw) if raw is not None else "").strip().upper()
    if s in STATI_AMMESSI:
        return s
    # Se manca, deduce da data_fine
    if data_fine is None:
        return "ATTIVO"
    try:
        if isinstance(data_fine, datetime):
            df = data_fine.date()
        elif isinstance(data_fine, date):
            df = data_fine
        else:
            # Prova a parsare stringa
            df_str = str(data_fine)
            if "-" in df_str:
                df = datetime.strptime(df_str[:10], "%Y-%m-%d").date()
            else:
                return "ATTIVO"
        return "ATTIVO" if df > date.today() else "DISMESSO"
    except Exception:
        return "ATTIVO"


def parse_xlsx(xlsx_path):
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel non trovato: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = None
    for cand in ("Parco Auto", "Foglio1", "Sheet1"):
        if cand in wb.sheetnames:
            sheet_name = cand
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel vuoto")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # Mappa colonne note a indice
    col = {}
    for i, name in enumerate(header):
        n = name.lower()
        if n.startswith("modello"):
            col["modello"] = i
        elif n == "targa":
            col["targa"] = i
        elif n.startswith("fornitore"):
            col["fornitore"] = i
        elif "pool" in n and "promiscuo" in n:
            col["pool_promiscuo"] = i
        elif "numero contratto" in n or "contratto" in n and "numero" in n:
            col["contratto"] = i
        elif "data inizio" in n:
            col["data_inizio"] = i
        elif "data fine" in n:
            col["data_fine"] = i
        elif n.startswith("canone"):
            col["canone"] = i
        elif n == "stato":
            col["stato"] = i
        elif "classificazione" in n and "fiscale" in n:
            col["classificazione_fiscale"] = i

    required = {"targa", "fornitore", "contratto"}
    missing = required - set(col.keys())
    if missing:
        raise ValueError(
            f"Colonne mancanti nel foglio '{sheet_name}': {missing}. "
            f"Header letto: {header}"
        )

    has_stato_col = "stato" in col
    has_cls_col = "classificazione_fiscale" in col

    entries = []
    skipped = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        targa = normalize_targa(row[col["targa"]] if col["targa"] < len(row) else None)
        fornitore = (str(row[col["fornitore"]] or "").strip()
                      if col["fornitore"] < len(row) else "")
        contratto = normalize_contratto(
            row[col["contratto"]] if col["contratto"] < len(row) else None
        )
        if not targa and not contratto:
            skipped.append({"row": row_idx, "motivo": "no_targa_no_contratto"})
            continue
        modello = ""
        if "modello" in col and col["modello"] < len(row):
            modello = str(row[col["modello"]] or "").strip()
        pool_promiscuo = ""
        if "pool_promiscuo" in col and col["pool_promiscuo"] < len(row):
            pool_promiscuo = str(row[col["pool_promiscuo"]] or "").strip()
        data_inizio = (row[col["data_inizio"]]
                        if "data_inizio" in col and col["data_inizio"] < len(row)
                        else None)
        data_fine = (row[col["data_fine"]]
                      if "data_fine" in col and col["data_fine"] < len(row)
                      else None)
        canone = (row[col["canone"]]
                   if "canone" in col and col["canone"] < len(row) else None)
        try:
            canone = float(canone) if canone is not None and str(canone).strip() else None
        except (ValueError, TypeError):
            canone = None

        # Stato
        stato_raw = (row[col["stato"]]
                      if has_stato_col and col["stato"] < len(row) else None)
        stato = normalize_stato(stato_raw, data_fine)

        # Classificazione
        cls_raw = (row[col["classificazione_fiscale"]]
                    if has_cls_col and col["classificazione_fiscale"] < len(row)
                    else None)
        classificazione = normalize_classificazione(cls_raw, fallback_pool_promiscuo=pool_promiscuo)

        # Assegnatario / pool: se col D = "POOL" -> "POOL", altrimenti il nome
        assegnatario = pool_promiscuo if pool_promiscuo.upper() != "POOL" else "POOL"

        entries.append({
            "row": row_idx,
            "targa": targa,
            "modello": modello,
            "fornitore": fornitore,
            "assegnatario": assegnatario,
            "contratto": contratto,
            "data_inizio": _to_iso_date(data_inizio),
            "data_fine": _to_iso_date(data_fine),
            "canone": canone,
            "stato": stato,
            "classificazione": classificazione,
        })
    return entries, skipped, sheet_name, has_stato_col, has_cls_col


def _to_iso_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return ""
    return s[:10] if len(s) >= 10 else s


def build_indexes(entries):
    """Costruisce gli indici per targa e contratto.

    Per targa: se più contratti per stessa targa (es. veicolo cambiato
    contratto a metà anno), tiene l'ATTIVO più recente.
    Per contratto: ogni contratto è univoco; se duplicato, log warning.
    """
    by_targa = {}
    by_contratto = {}
    duplicati_targa = []
    duplicati_contratto = []

    # Ordino per stato (ATTIVO prima) e poi data_inizio desc
    sorted_entries = sorted(
        entries,
        key=lambda e: (
            0 if e["stato"] == "ATTIVO" else 1,
            -(int(e["data_inizio"][:4]) if e["data_inizio"][:4].isdigit() else 0),
            -(int(e["data_inizio"][5:7]) if len(e["data_inizio"]) >= 7
              and e["data_inizio"][5:7].isdigit() else 0),
        ),
    )

    for e in sorted_entries:
        t = e["targa"]
        c = e["contratto"]
        if t and t not in by_targa:
            by_targa[t] = e
        elif t:
            duplicati_targa.append((t, by_targa[t], e))
        if c and c not in by_contratto:
            by_contratto[c] = e
        elif c:
            duplicati_contratto.append((c, by_contratto[c], e))

    return by_targa, by_contratto, duplicati_targa, duplicati_contratto


def render_module(by_targa, by_contratto, sheet_name, xlsx_path,
                   has_stato_col, has_cls_col, n_skipped, n_dup_t, n_dup_c):
    sorted_targhe = sorted(by_targa.keys())
    sorted_contratti = sorted(by_contratto.keys())

    by_cls = Counter()
    by_stato = Counter()
    by_fornitore = Counter()
    for e in by_targa.values():
        by_cls[e["classificazione"]] += 1
        by_stato[e["stato"]] += 1
        by_fornitore[e["fornitore"]] += 1

    L = []
    L.append('"""')
    L.append('Mappa Parco Auto -> classificazione fiscale veicolo (Ecotel).')
    L.append('')
    L.append('AUTO-GENERATO da scripts/generate_parco_auto_mapping.py')
    L.append(f'Sorgente: {xlsx_path.name} (foglio "{sheet_name}")')
    L.append(f'Generato: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    L.append(f'Targhe distinte: {len(by_targa)}')
    L.append(f'Contratti distinti: {len(by_contratto)}')
    L.append(f'Colonna STATO presente: {has_stato_col} '
             f'(altrimenti dedotto da DATA FINE)')
    L.append(f'Colonna CLASSIFICAZIONE FISCALE presente: {has_cls_col} '
             f'(altrimenti dedotta da col D)')
    L.append(f'Distribuzione classificazione: {dict(by_cls)}')
    L.append(f'Distribuzione stato: {dict(by_stato)}')
    L.append(f'Distribuzione fornitore: {dict(by_fornitore)}')
    L.append(f'Skipped: {n_skipped} | Duplicati targa: {n_dup_t} | '
             f'Duplicati contratto: {n_dup_c}')
    L.append('')
    L.append('NON modificare a mano: rigenerare con')
    L.append('    python scripts/generate_parco_auto_mapping.py')
    L.append('"""')
    L.append('from typing import Optional, Dict')
    L.append('')
    L.append('# Costanti classificazione')
    L.append('CLASSIFICAZIONE_POOL = "POOL"')
    L.append('CLASSIFICAZIONE_USO_PROMISCUO = "uso_promiscuo"')
    L.append('CLASSIFICAZIONE_SUPER_LUSSO = "super_lusso"')
    L.append('')
    L.append('STATI_ATTIVI = {"ATTIVO"}')
    L.append('STATI_STORICI = {"DISMESSO"}')
    L.append('')

    # PARCO_BY_TARGA
    L.append('# Mappa per TARGA -> info veicolo (1 entry per targa, attivo prevale)')
    L.append('PARCO_BY_TARGA: Dict[str, dict] = {')
    for t in sorted_targhe:
        e = by_targa[t]
        modello = (e["modello"] or "").replace("'", "\\'")
        forn = (e["fornitore"] or "").replace("'", "\\'")
        ass = (e["assegnatario"] or "").replace("'", "\\'")
        L.append(f'    "{t}": {{')
        L.append(f'        "targa": "{t}",')
        L.append(f'        "modello": \'{modello[:80]}\',')
        L.append(f'        "fornitore": \'{forn[:40]}\',')
        L.append(f'        "assegnatario": \'{ass[:40]}\',')
        L.append(f'        "contratto": "{e["contratto"]}",')
        L.append(f'        "classificazione": "{e["classificazione"]}",')
        L.append(f'        "stato": "{e["stato"]}",')
        L.append(f'        "data_inizio": "{e["data_inizio"]}",')
        L.append(f'        "data_fine": "{e["data_fine"]}",')
        if e["canone"] is not None:
            L.append(f'        "canone": {e["canone"]},')
        L.append(f'    }},')
    L.append('}')
    L.append('')

    # PARCO_BY_CONTRATTO
    L.append('# Mappa per NUMERO CONTRATTO -> info veicolo')
    L.append('# (Tecnoalt/ALD usano DatiContratto.IdDocumento, UnipolRental via desc)')
    L.append('PARCO_BY_CONTRATTO: Dict[str, dict] = {')
    for c in sorted_contratti:
        e = by_contratto[c]
        modello = (e["modello"] or "").replace("'", "\\'")
        forn = (e["fornitore"] or "").replace("'", "\\'")
        ass = (e["assegnatario"] or "").replace("'", "\\'")
        L.append(f'    "{c}": {{')
        L.append(f'        "targa": "{e["targa"]}",')
        L.append(f'        "modello": \'{modello[:80]}\',')
        L.append(f'        "fornitore": \'{forn[:40]}\',')
        L.append(f'        "assegnatario": \'{ass[:40]}\',')
        L.append(f'        "contratto": "{c}",')
        L.append(f'        "classificazione": "{e["classificazione"]}",')
        L.append(f'        "stato": "{e["stato"]}",')
        L.append(f'        "data_inizio": "{e["data_inizio"]}",')
        L.append(f'        "data_fine": "{e["data_fine"]}",')
        if e["canone"] is not None:
            L.append(f'        "canone": {e["canone"]},')
        L.append(f'    }},')
    L.append('}')
    L.append('')

    # Helper
    L.append('')
    L.append('def get_info_by_targa(targa: str) -> Optional[dict]:')
    L.append('    """Ritorna info veicolo dato la targa (uppercase normalizzata)."""')
    L.append('    if not targa:')
    L.append('        return None')
    L.append('    return PARCO_BY_TARGA.get(targa.upper().strip())')
    L.append('')
    L.append('')
    L.append('def get_info_by_contratto(num: str) -> Optional[dict]:')
    L.append('    """Ritorna info veicolo dato il numero contratto."""')
    L.append('    if not num:')
    L.append('        return None')
    L.append('    return PARCO_BY_CONTRATTO.get(str(num).strip())')
    L.append('')
    L.append('')
    L.append('def get_classificazione_by_targa(targa: str) -> Optional[str]:')
    L.append('    """Ritorna classificazione fiscale (POOL/uso_promiscuo/super_lusso) o None."""')
    L.append('    info = get_info_by_targa(targa)')
    L.append('    return info["classificazione"] if info else None')
    L.append('')
    L.append('')
    L.append('def get_classificazione_by_contratto(num: str) -> Optional[str]:')
    L.append('    """Ritorna classificazione fiscale dato il numero contratto."""')
    L.append('    info = get_info_by_contratto(num)')
    L.append('    return info["classificazione"] if info else None')
    L.append('')

    return "\n".join(L)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"Path file Excel sorgente (default: {DEFAULT_XLSX})")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help=f"Path file Python output (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--print-summary", action="store_true",
                        help="Stampa riepilogo dettagliato")
    args = parser.parse_args()

    print(f"== Generator parco_auto_mapping ==")
    print(f"  Sorgente: {args.xlsx}")
    print(f"  Output:   {args.output}")

    entries, skipped, sheet_name, has_stato_col, has_cls_col = parse_xlsx(args.xlsx)
    print(f"  Foglio letto: '{sheet_name}'")
    print(f"  Colonna STATO: {'presente' if has_stato_col else 'ASSENTE (deduco da data_fine)'}")
    print(f"  Colonna CLASSIFICAZIONE FISCALE: "
          f"{'presente' if has_cls_col else 'ASSENTE (deduco da col D)'}")
    print(f"  Righe valide: {len(entries)}")
    print(f"  Skipped: {len(skipped)}")

    by_targa, by_contratto, dup_t, dup_c = build_indexes(entries)
    print(f"  Targhe distinte: {len(by_targa)}")
    print(f"  Contratti distinti: {len(by_contratto)}")
    if dup_t:
        print(f"  [!] Duplicati targa (tenuta entry attiva/recente): {len(dup_t)}")
        for t, prev, curr in dup_t[:5]:
            print(f"    {t}: prev contratto={prev['contratto']} stato={prev['stato']} "
                  f"-> curr contratto={curr['contratto']} stato={curr['stato']}")
    if dup_c:
        print(f"  [!] Duplicati contratto: {len(dup_c)}")

    by_cls = Counter()
    by_stato = Counter()
    by_fornitore = Counter()
    for e in by_targa.values():
        by_cls[e["classificazione"]] += 1
        by_stato[e["stato"]] += 1
        by_fornitore[e["fornitore"]] += 1
    print(f"  Distribuzione classificazione: {dict(by_cls)}")
    print(f"  Distribuzione stato: {dict(by_stato)}")
    print(f"  Distribuzione fornitore: {dict(by_fornitore)}")

    src = render_module(by_targa, by_contratto, sheet_name, args.xlsx,
                         has_stato_col, has_cls_col,
                         len(skipped), len(dup_t), len(dup_c))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(src, encoding="utf-8")
    print(f"\nScritto {args.output} ({len(src):,} bytes)")

    if args.print_summary and skipped:
        print(f"\nRighe skipped ({len(skipped)}):")
        for s in skipped[:20]:
            print(f"  {s}")


if __name__ == "__main__":
    main()
