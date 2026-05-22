"""
Generator idempotente di config/carte_carburante_mapping.py partendo da
input/carte_uta.xlsx.

File sorgente atteso (foglio "CARTE UTA"):
    | numero_carta | targa | nickname | classificazione_fiscale | stato |

Solo le righe con stato='ATTIVO' vengono incluse nel mapping.

Genera config/carte_carburante_mapping.py con:
  - CARTE_UTA_BY_NUMERO: dict[numero_carta_str, info]
  - Helper: get_carta_uta(numero), get_classificazione_carta_uta(numero)

Uso:
    python scripts/generate_carte_carburante_mapping.py
    python scripts/generate_carte_carburante_mapping.py --print-summary

Lo script è READ-ONLY su filesystem (eccetto il file di output).
"""
import argparse
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRORE: openpyxl non installato (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "input" / "carte_uta.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "config" / "carte_carburante_mapping.py"

CLASSIFICAZIONI_AMMESSE = {"POOL", "uso_promiscuo", "super_lusso", "SERVIZIO"}
STATI_AMMESSI = {"ATTIVO", "DISMESSO"}


def normalize_numero_carta(raw):
    """Numero carta UTA: testo grezzo, senza conversione numerica.
    Lo trattiamo sempre come stringa per evitare problemi di precisione
    su numeri >15 cifre (le carte UTA hanno 17 cifre)."""
    if raw is None:
        return ""
    s = str(raw).strip()
    # Rimuovo eventuali decimali se Excel l'ha letto come float
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalize_targa(raw):
    if not raw:
        return ""
    return str(raw).upper().strip()


def normalize_nickname(raw):
    if not raw:
        return ""
    return str(raw).strip()


def normalize_classificazione(raw):
    if not raw:
        return ""
    return str(raw).strip()


def normalize_stato(raw):
    if not raw:
        return "ATTIVO"
    s = str(raw).strip().upper()
    return s if s in STATI_AMMESSI else "ATTIVO"


def parse_xlsx(xlsx_path):
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel non trovato: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = None
    for cand in ("CARTE UTA", "Carte UTA", "Foglio1", "Sheet1"):
        if cand in wb.sheetnames:
            sheet_name = cand
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel vuoto")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # Mappa header -> indice
    col = {}
    for i, name in enumerate(header):
        n = name.lower()
        if n == "numero_carta":
            col["numero_carta"] = i
        elif n == "targa":
            col["targa"] = i
        elif n == "nickname":
            col["nickname"] = i
        elif "classificazione" in n:
            col["classificazione"] = i
        elif n == "stato":
            col["stato"] = i

    required = {"numero_carta", "classificazione"}
    missing = required - set(col.keys())
    if missing:
        raise ValueError(
            f"Colonne obbligatorie mancanti nel foglio '{sheet_name}': {missing}. "
            f"Header letto: {header}"
        )

    entries = []
    skipped = []
    cls_anomale = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        numero = normalize_numero_carta(
            row[col["numero_carta"]] if col["numero_carta"] < len(row) else None
        )
        if not numero:
            skipped.append({"row": row_idx, "motivo": "numero_carta vuoto"})
            continue

        targa = (normalize_targa(row[col["targa"]])
                  if "targa" in col and col["targa"] < len(row) else "")
        nickname = (normalize_nickname(row[col["nickname"]])
                     if "nickname" in col and col["nickname"] < len(row) else "")
        classif = normalize_classificazione(
            row[col["classificazione"]] if col["classificazione"] < len(row) else None
        )
        stato = (normalize_stato(row[col["stato"]])
                  if "stato" in col and col["stato"] < len(row) else "ATTIVO")

        if not classif:
            skipped.append({
                "row": row_idx, "numero": numero,
                "motivo": "classificazione_fiscale vuota"
            })
            continue
        if classif not in CLASSIFICAZIONI_AMMESSE:
            cls_anomale.append({
                "row": row_idx, "numero": numero, "valore": classif
            })
            # La aggiungo comunque, il writer ricadrà su fallback
        if stato != "ATTIVO":
            skipped.append({
                "row": row_idx, "numero": numero,
                "motivo": f"stato={stato} (non ATTIVO)"
            })
            continue

        entries.append({
            "row": row_idx,
            "numero_carta": numero,
            "targa": targa,
            "nickname": nickname,
            "classificazione": classif,
            "stato": stato,
        })
    return entries, skipped, cls_anomale, sheet_name


def build_index(entries):
    """Indice principale: numero_carta -> info. Logga duplicati."""
    by_numero = {}
    duplicati = []
    for e in entries:
        n = e["numero_carta"]
        if n in by_numero:
            duplicati.append((n, by_numero[n]["row"], e["row"]))
        else:
            by_numero[n] = e
    return by_numero, duplicati


def render_module(by_numero, sheet_name, xlsx_path, n_skipped, n_dup, n_cls_anomale):
    sorted_nums = sorted(by_numero.keys())

    by_cls = Counter()
    by_nick = Counter()
    for e in by_numero.values():
        by_cls[e["classificazione"]] += 1
        if e["nickname"]:
            kind = e["nickname"].split()[0].upper() if " " in e["nickname"] else "ALTRO"
            by_nick[kind] += 1
        else:
            by_nick["(senza_nickname)"] += 1

    L = []
    L.append('"""')
    L.append('Mappa carte carburante UTA -> classificazione fiscale veicolo (Ecotel).')
    L.append('')
    L.append('AUTO-GENERATO da scripts/generate_carte_carburante_mapping.py')
    L.append(f'Sorgente: {xlsx_path.name} (foglio "{sheet_name}")')
    L.append(f'Generato: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    L.append(f'Carte attive: {len(by_numero)}')
    L.append(f'Skipped (non-ATTIVO o vuoti): {n_skipped} | '
             f'Duplicati: {n_dup} | Classificazioni anomale: {n_cls_anomale}')
    L.append(f'Distribuzione classificazione: {dict(by_cls)}')
    L.append(f'Distribuzione nickname: {dict(by_nick)}')
    L.append('')
    L.append('NON modificare a mano: rigenerare con')
    L.append('    python scripts/generate_carte_carburante_mapping.py')
    L.append('"""')
    L.append('from typing import Optional, Dict')
    L.append('')
    L.append('# Costanti classificazione')
    L.append('CLASSIFICAZIONE_POOL = "POOL"')
    L.append('CLASSIFICAZIONE_USO_PROMISCUO = "uso_promiscuo"')
    L.append('CLASSIFICAZIONE_SUPER_LUSSO = "super_lusso"')
    L.append('CLASSIFICAZIONE_SERVIZIO = "SERVIZIO"')
    L.append('')
    L.append('# Mappa per NUMERO CARTA UTA -> info veicolo/classe')
    L.append('# Chiave: stringa numero_carta (RiferimentoAmministrazione XML)')
    L.append('CARTE_UTA_BY_NUMERO: Dict[str, dict] = {')
    for n in sorted_nums:
        e = by_numero[n]
        targa = (e["targa"] or "").replace("'", "\\'")
        nick = (e["nickname"] or "").replace("'", "\\'")
        L.append(f'    "{n}": {{')
        L.append(f'        "numero_carta": "{n}",')
        L.append(f'        "targa": \'{targa}\',')
        L.append(f'        "nickname": \'{nick}\',')
        L.append(f'        "classificazione": "{e["classificazione"]}",')
        L.append(f'        "stato": "{e["stato"]}",')
        L.append(f'    }},')
    L.append('}')
    L.append('')
    L.append('')
    L.append('def get_carta_uta(numero: str) -> Optional[dict]:')
    L.append('    """Ritorna info carta UTA dato il numero (stringa, normalizzata)."""')
    L.append('    if numero is None:')
    L.append('        return None')
    L.append('    return CARTE_UTA_BY_NUMERO.get(str(numero).strip())')
    L.append('')
    L.append('')
    L.append('def get_classificazione_carta_uta(numero: str) -> Optional[str]:')
    L.append('    """Ritorna classificazione fiscale (POOL/uso_promiscuo/super_lusso/SERVIZIO)')
    L.append('    o None se carta non in mappa (riga da censire)."""')
    L.append('    info = get_carta_uta(numero)')
    L.append('    return info["classificazione"] if info else None')
    L.append('')

    return "\n".join(L)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"Path Excel sorgente (default: {DEFAULT_XLSX})")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help=f"Path Python output (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--print-summary", action="store_true",
                        help="Stampa riepilogo dettagliato")
    args = parser.parse_args()

    print(f"== Generator carte_carburante_mapping ==")
    print(f"  Sorgente: {args.xlsx}")
    print(f"  Output:   {args.output}")

    entries, skipped, cls_anomale, sheet_name = parse_xlsx(args.xlsx)
    print(f"  Foglio letto: '{sheet_name}'")
    print(f"  Righe ATTIVE valide: {len(entries)}")
    print(f"  Skipped: {len(skipped)}")

    by_numero, dup = build_index(entries)
    print(f"  Carte distinte: {len(by_numero)}")
    if dup:
        print(f"  [!] Duplicati numero_carta: {len(dup)}")
        for n, r1, r2 in dup[:5]:
            print(f"    {n}: prima a riga {r1}, ripetuta a riga {r2}")

    if cls_anomale:
        print(f"  [!] Classificazioni anomale (non in {CLASSIFICAZIONI_AMMESSE}): {len(cls_anomale)}")
        for x in cls_anomale[:5]:
            print(f"    riga {x['row']}: numero={x['numero']} valore={x['valore']!r}")

    by_cls = Counter()
    for e in by_numero.values():
        by_cls[e["classificazione"]] += 1
    print(f"  Distribuzione classificazione: {dict(by_cls)}")

    src = render_module(by_numero, sheet_name, args.xlsx,
                         len(skipped), len(dup), len(cls_anomale))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(src, encoding="utf-8")
    print(f"\nScritto {args.output} ({len(src):,} bytes)")

    if args.print_summary and skipped:
        print(f"\nRighe skipped ({len(skipped)}):")
        for s in skipped[:20]:
            print(f"  {s}")


if __name__ == "__main__":
    main()
