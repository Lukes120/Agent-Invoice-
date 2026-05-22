"""
Report giornaliero presenza fatture in "e-fatture in ingresso" (Ecotel, no autofatture)
dal 2026-01-01 a oggi.

Per ogni giorno X:
- A "arrivate"       = N. attachment con create_date in giorno X
- B "registrate"     = N. attachment registrati in giorno X
                       (= create_date della account.move collegata = giorno X)
- C "stock fine giorno X" = cumulato arrivate(<=X) − cumulato registrate(<=X)
                            Include il backlog ereditato pre-2026.

Differenza rispetto alla v1: la "data di registrazione" non è più stimata col
write_date dell'attachment (che è facilmente alterato da bulk update), ma è la
create_date della account.move collegata via fatturapa_attachment_in_id, che è
immutabile e quindi storicamente affidabile.

Mapping attachment↔move: matching via tripletta (partner_id, ref, amount_total)
perché lettura diretta del Many2one fatturapa_attachment_in_id innesta il bug
Odoo "column tipo_documento does not exist".

Output:
- output/efatture_giornaliero.xlsx con tabella + grafico
- output/xml.txt con log testuale
"""
import sys, os
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(ROOT / 'config' / 'credentials.env')

from core.odoo_client import OdooReadOnlyClient
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_TXT = ROOT / 'output' / 'xml.txt'
OUT_XLSX = ROOT / 'output' / 'efatture_giornaliero.xlsx'
OUT_TXT.parent.mkdir(parents=True, exist_ok=True)

_logf = open(OUT_TXT, 'w', encoding='utf-8')
_stdout = sys.stdout
class _Tee:
    def __init__(self, *s): self.s = s
    def write(self, d):
        for x in self.s:
            try: x.write(d)
            except: pass
    def flush(self):
        for x in self.s:
            try: x.flush()
            except: pass
sys.stdout = _Tee(_stdout, _logf)


client = OdooReadOnlyClient(
    url=os.environ['ODOO_URL'], db=os.environ['ODOO_DB'],
    username=os.environ['ODOO_USERNAME'], password=os.environ['ODOO_PASSWORD'])
client.connect()

start_date = date(2026, 1, 1)
today = date.today()

print("=" * 100)
print(f"REPORT GIORNALIERO E-FATTURE IN INGRESSO — Ecotel, no autofatture")
print(f"Periodo: {start_date} → {today}")
print("=" * 100)

# === 1. Carico attachment Ecotel (tutto lo storico) ===
print("\n[1/3] Fetch attachment storici...")
attachments = client._call('fatturapa.attachment.in', 'search_read',
    [('company_id', '=', 1),
     ('is_self_invoice', '=', False)],
    fields=['id', 'create_date', 'registered',
            'invoices_date', 'invoices_total', 'xml_supplier_id'],
    order='create_date asc',
    limit=100000)
print(f"  Attachments: {len(attachments)}")

# === 2. Carico move con fatturapa_attachment_in_id valorizzato ===
print("\n[2/3] Fetch account.move collegate a fatturapa...")
moves = client._call('account.move', 'search_read',
    [('fatturapa_attachment_in_id', '!=', False),
     ('company_id', '=', 1)],
    fields=['id', 'create_date', 'invoice_date', 'partner_id', 'amount_total'],
    order='create_date asc',
    limit=100000)
print(f"  Moves: {len(moves)}")

# === 3. Mapping attachment ↔ move via tripletta (partner, invoice_date, amount_total) ===
print("\n[3/3] Matching attachment ↔ move (partner + invoice_date + amount)...")

def _parse_invoices_date(s):
    """invoices_date sull'attachment è una stringa, di solito 'dd/mm/YYYY' o ISO."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

# Indicizzo le move per chiave (partner_id, invoice_date, amount_total)
moves_by_key = defaultdict(list)
for m in moves:
    p = m.get('partner_id')
    pid = p[0] if isinstance(p, list) and p else None
    inv_date = m.get('invoice_date')
    if isinstance(inv_date, str):
        try:
            inv_date = datetime.strptime(inv_date[:10], '%Y-%m-%d').date()
        except ValueError:
            inv_date = None
    amt = round(float(m.get('amount_total') or 0), 2)
    if pid and inv_date:
        moves_by_key[(pid, inv_date, amt)].append(m)

# Per ogni attachment registered=True trovo la move corrispondente
unmatched_registered = 0
registration_date_by_att = {}
used_move_ids = set()
for a in attachments:
    if not a.get('registered'):
        continue
    sup = a.get('xml_supplier_id')
    sup_id = sup[0] if isinstance(sup, list) and sup else None
    inv_d = _parse_invoices_date(a.get('invoices_date'))
    tot = round(float(a.get('invoices_total') or 0), 2)
    candidates = moves_by_key.get((sup_id, inv_d, tot), [])
    # filtro le move già usate per altri attachment (evita doppi match)
    free_cands = [m for m in candidates if m['id'] not in used_move_ids]
    if free_cands:
        m = min(free_cands, key=lambda x: x.get('create_date') or '')
        used_move_ids.add(m['id'])
        try:
            rd = datetime.strptime((m.get('create_date') or '')[:10], '%Y-%m-%d').date()
            registration_date_by_att[a['id']] = rd
        except ValueError:
            unmatched_registered += 1
    else:
        unmatched_registered += 1

print(f"  Match riusciti:   {len(registration_date_by_att)}")
print(f"  Senza match:      {unmatched_registered} (registered=True ma nessuna move agganciata via tripletta)")

# Per gli unmatched, uso fallback: write_date dell'attachment.
# Ma per consistenza con la metrica B, li conteggio come "registrati il giorno dell'arrivo"
# (non ne abbiamo la data precisa).

# === Aggregazione per giorno ===
arrived_day = defaultdict(int)
registered_day = defaultdict(int)
oldest = None

for a in attachments:
    cd_raw = a.get('create_date') or ''
    try:
        cd = datetime.strptime(cd_raw[:10], '%Y-%m-%d').date()
    except ValueError:
        continue
    if oldest is None or cd < oldest:
        oldest = cd
    arrived_day[cd] += 1
    if a.get('registered'):
        rd = registration_date_by_att.get(a['id'])
        if rd is None:
            # fallback: stesso giorno arrivo (mancanza match = assunzione conservativa)
            rd = cd
        registered_day[rd] += 1

print(f"\nRecord più vecchio: {oldest}")

# Serie cumulata da oldest a today
all_days = []
cur = oldest
while cur <= today:
    all_days.append(cur)
    cur += timedelta(days=1)

stock_cum = 0
stock_by_day = {}
for d in all_days:
    stock_cum += arrived_day.get(d, 0)
    stock_cum -= registered_day.get(d, 0)
    stock_by_day[d] = stock_cum

# Ritaglio sul periodo richiesto
rows = []
cur = start_date
while cur <= today:
    rows.append({
        'giorno': cur,
        'arrivate': arrived_day.get(cur, 0),
        'registrate': registered_day.get(cur, 0),
        'stock_fine_giorno': stock_by_day.get(cur, 0),
    })
    cur += timedelta(days=1)

total_arrived = sum(r['arrivate'] for r in rows)
total_registered = sum(r['registrate'] for r in rows)
stock_oggi = rows[-1]['stock_fine_giorno']
stock_iniziale = rows[0]['stock_fine_giorno']

print(f"\nGiorni nel periodo:    {len(rows)}")
print(f"Stock al {start_date}: {stock_iniziale} (residui pre-2026)")
print(f"Arrivate nel periodo:  {total_arrived}")
print(f"Registrate nel periodo:{total_registered}")
print(f"Stock oggi ({today}):  {stock_oggi}")

# === XLSX ===
print(f"\nGenerazione XLSX...")
wb = Workbook()
ws = wb.active
ws.title = "Report giornaliero"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
title_font = Font(bold=True, size=14)
small_bold = Font(bold=True)
border = Border(left=Side(style='thin', color='BBBBBB'),
                right=Side(style='thin', color='BBBBBB'),
                top=Side(style='thin', color='BBBBBB'),
                bottom=Side(style='thin', color='BBBBBB'))
center = Alignment(horizontal='center')
right = Alignment(horizontal='right')

ws['A1'] = "E-fatture in ingresso — Ecotel (no autofatture)"
ws['A1'].font = title_font
ws.merge_cells('A1:E1')
ws['A2'] = f"Periodo: {start_date} → {today}  |  Generato: {datetime.now():%Y-%m-%d %H:%M}"
ws['A2'].font = Font(italic=True, color="666666")
ws.merge_cells('A2:E2')

ws['G1'] = "Sintesi"
ws['G1'].font = title_font
summary = [
    (f"Stock iniziale (fine {start_date})", stock_iniziale),
    ("Arrivate nel periodo", total_arrived),
    ("Registrate nel periodo", total_registered),
    (f"Stock OGGI ({today})", stock_oggi),
    ("Delta stock nel periodo", stock_oggi - stock_iniziale),
    ("Match attachment↔move non riusciti", unmatched_registered),
]
for i, (lbl, val) in enumerate(summary, start=2):
    ws.cell(row=i, column=7, value=lbl).font = small_bold
    ws.cell(row=i, column=8, value=val)

HDR_ROW = 5
headers = ["Giorno", "Arrivate", "Registrate", "Stock fine giorno"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

max_stock = max(r['stock_fine_giorno'] for r in rows) if rows else 0
for i, r in enumerate(rows, start=HDR_ROW + 1):
    ws.cell(row=i, column=1, value=r['giorno']).number_format = 'YYYY-MM-DD'
    ws.cell(row=i, column=2, value=r['arrivate']).alignment = right
    ws.cell(row=i, column=3, value=r['registrate']).alignment = right
    sc = ws.cell(row=i, column=4, value=r['stock_fine_giorno'])
    sc.alignment = right
    if max_stock > 0:
        if r['stock_fine_giorno'] >= max_stock * 0.9:
            sc.fill = PatternFill("solid", fgColor="F4B084")
        elif r['stock_fine_giorno'] >= max_stock * 0.5:
            sc.fill = PatternFill("solid", fgColor="FCE4D6")
    for c in (1, 2, 3, 4):
        ws.cell(row=i, column=c).border = border

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 20
ws.column_dimensions['G'].width = 36
ws.column_dimensions['H'].width = 12
ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

# Grafico
chart = LineChart()
chart.title = "E-fatture — flusso giornaliero e stock cumulato"
chart.style = 12
chart.y_axis.title = "Numero fatture"
chart.x_axis.title = "Giorno"
chart.height = 11
chart.width = 26

DATA_START = HDR_ROW + 1
DATA_END = HDR_ROW + len(rows)
data_ref = Reference(ws, min_col=2, min_row=HDR_ROW, max_col=4, max_row=DATA_END)
cats_ref = Reference(ws, min_col=1, min_row=DATA_START, max_row=DATA_END)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
colors = ["305496", "70AD47", "C00000"]
for i, s in enumerate(chart.series):
    s.graphicalProperties.line.solidFill = colors[i]
    s.graphicalProperties.line.width = 18000
ws.add_chart(chart, "G10")

wb.save(OUT_XLSX)
print(f"  XLSX salvato: {OUT_XLSX}")
print(f"\nReport testuale: {OUT_TXT}")

sys.stdout = _stdout
_logf.close()
