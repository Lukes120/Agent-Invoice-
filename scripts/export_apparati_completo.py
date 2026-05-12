"""
Genera un Excel con TUTTI gli apparati visti nelle fatture del network
Telepass Q1 2026 (gennaio-marzo) — Autostrade + Apcoa — + Aprile 2026
(PDF in input/ scaricati dal portale Autostrade).

I PDF Telepass sono canoni aggregati senza lista apparati, quindi non
producono righe in questo report (segnalato in console).

Le righe MAPPATE hanno targa, veicolo e classificazione gia' compilate
(verde). Le righe DA CENSIRE hanno quelle 3 colonne in giallo, vuote.
"""
import os
import sys
import re
import glob
from collections import defaultdict
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from core.pdf_parser import parse_pdf_autostrade
from config.apparati_mapping import get_classificazione, get_apparato_info

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT_PDF = r"C:\Users\lranalletta\Downloads\cHECK\_extracted"
INPUT_DIR = os.path.join(ROOT, 'input')

RE_CC = re.compile(r'Contratto\s+(\d+)')
RE_MESE = re.compile(r'(\d{2})\s*-\s*Fatture\s+(\w+)\s+(\d{4})')
# CC dentro al PDF (per i file in input/ che non hanno path-mese)
RE_CC_PDF = re.compile(r'CODICE\s+CLIENTE[^0-9]*(\d{6,})', re.IGNORECASE)

# fornitori che hanno la lista apparati nel PDF + parser compatibile
# (lo stesso parser_pdf_autostrade gestisce anche Apcoa grazie al regex
# esteso "NUMERO SOSTE PARCHEGGI" in core/pdf_parser.py)
PDF_PATTERNS = {
    'Autostrade': 'Autostrade*.pdf',
    'Apcoa': 'Apcoa*.pdf',
}
# fornitori senza lista apparati (canoni aggregati): solo segnalati
PDF_NO_APPARATI = ['Telepass*.pdf']


def parse_path_meta(rel_path):
    cc = RE_CC.search(rel_path)
    mese = RE_MESE.search(rel_path)
    return {
        'cc': cc.group(1) if cc else '',
        'mese_nome': mese.group(2) if mese else '',
        'anno': mese.group(3) if mese else '',
    }


def estrai_cc_da_pdf(pdf_path):
    """Per PDF senza path-mese (input/): cerca CODICE CLIENTE NNN nel testo."""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text() or ''
        m = RE_CC_PDF.search(text)
        return m.group(1) if m else ''
    except Exception:
        return ''


def main():
    # 1. Conta i PDF senza apparati (Telepass) — solo per nota informativa
    n_telepass = 0
    for pat in PDF_NO_APPARATI:
        n_telepass += len(glob.glob(os.path.join(ROOT_PDF, '**', pat),
                                     recursive=True))
    print(f'PDF Telepass canone (no apparati nel PDF): {n_telepass} — saltati')

    # 2. raccolgo apparati da Autostrade e Apcoa
    seen = defaultdict(lambda: {
        'tot_eur': 0.0,
        'count_pdf': 0,
        'count_movimenti': 0,
        'months': defaultdict(float),
        'clienti': set(),
        'fornitori': set(),
    })
    n_pdfs = {}
    for fornitore, pat in PDF_PATTERNS.items():
        pdfs = sorted(glob.glob(os.path.join(ROOT_PDF, '**', pat),
                                recursive=True))
        n_pdfs[fornitore] = len(pdfs)
        print(f'PDF {fornitore} (Q1 cHECK) trovati: {len(pdfs)}')
        for pdf in pdfs:
            rel = pdf.replace(ROOT_PDF + os.sep, '')
            meta = parse_path_meta(rel)
            data = parse_pdf_autostrade(pdf)
            for a in data.apparati:
                key = (a.tipo, a.apparato_id)
                d = seen[key]
                d['tot_eur'] += a.importo_iva_inclusa
                d['count_pdf'] += 1
                d['count_movimenti'] += a.n_movimenti
                d['clienti'].add(meta['cc'])
                d['fornitori'].add(fornitore)
                month_key = f'{meta["mese_nome"]} {meta["anno"]}'
                d['months'][month_key] += a.importo_iva_inclusa

    # 2b. PDF in input/ (Aprile 2026) — sono Autostrade dal portale
    input_pdfs = sorted(glob.glob(os.path.join(INPUT_DIR, '*.pdf')))
    n_pdfs['Autostrade Aprile (input)'] = len(input_pdfs)
    print(f'PDF Autostrade aprile in input/ trovati: {len(input_pdfs)}')
    for pdf in input_pdfs:
        cc = estrai_cc_da_pdf(pdf)
        data = parse_pdf_autostrade(pdf)
        for a in data.apparati:
            key = (a.tipo, a.apparato_id)
            d = seen[key]
            d['tot_eur'] += a.importo_iva_inclusa
            d['count_pdf'] += 1
            d['count_movimenti'] += a.n_movimenti
            if cc:
                d['clienti'].add(cc)
            d['fornitori'].add('Autostrade')
            d['months']['Aprile 2026'] += a.importo_iva_inclusa

    print(f'\nApparati unici visti nel network (Q1 + Aprile): {len(seen)}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Apparati'

    title = Font(bold=True, color='FFFFFF', size=11)
    fill_hdr = PatternFill('solid', fgColor='305496')
    fill_todo = PatternFill('solid', fgColor='FFF2CC')
    fill_ok = PatternFill('solid', fgColor='E2EFDA')
    thin = Side(border_style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        ('Stato', 14),
        ('Fornitore(i)', 18),
        ('Tipo apparato', 14),
        ('Codice apparato (da PDF)', 24),
        ('Codice cliente (cc)', 22),
        ('# fatture in cui appare', 22),
        ('# movimenti totali (Q1+Apr)', 24),
        ('€ totale Q1+Aprile 2026 (IVA incl.)', 32),
        ('Gennaio 2026 €', 16),
        ('Febbraio 2026 €', 16),
        ('Marzo 2026 €', 16),
        ('Aprile 2026 €', 16),
        ('TARGA', 16),
        ('VEICOLO descrizione', 38),
        ('CLASSIFICAZIONE', 22),
        ('Note', 30),
    ]
    for i, (h, w) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = title
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 38

    dv = DataValidation(
        type='list',
        formula1='"furgoni,uso_promiscuo,non assegnata / pool"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)

    items = sorted(seen.items(), key=lambda kv: -kv[1]['tot_eur'])
    row_idx = 2
    n_mapped = 0
    n_todo = 0
    eur_mapped = 0.0
    eur_todo = 0.0

    for (tipo, aid), info in items:
        meta_map = get_apparato_info(aid)
        is_mapped = meta_map is not None

        if is_mapped:
            stato = 'MAPPATO'
            targa = meta_map.get('targa', '')
            veicolo = meta_map.get('veicolo_descrizione', '')
            classificazione = meta_map.get('classificazione', '')
            n_mapped += 1
            eur_mapped += info['tot_eur']
        else:
            stato = 'DA CENSIRE'
            targa = ''
            veicolo = ''
            classificazione = ''
            n_todo += 1
            eur_todo += info['tot_eur']

        gen = round(info['months'].get('Gennaio 2026', 0.0), 2)
        feb = round(info['months'].get('Febbraio 2026', 0.0), 2)
        mar = round(info['months'].get('Marzo 2026', 0.0), 2)
        apr = round(info['months'].get('Aprile 2026', 0.0), 2)

        cells = [
            stato,
            ', '.join(sorted(info['fornitori'])),
            tipo, aid,
            ', '.join(sorted(info['clienti'])),
            info['count_pdf'],
            info['count_movimenti'],
            round(info['tot_eur'], 2),
            gen if gen else '',
            feb if feb else '',
            mar if mar else '',
            apr if apr else '',
            targa, veicolo, classificazione, '',
        ]
        for col_idx, val in enumerate(cells, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border
            if col_idx in (6, 7, 8, 9, 10, 11, 12):
                c.alignment = Alignment(horizontal='right')
            if col_idx in (13, 14, 15):
                c.fill = fill_ok if is_mapped else fill_todo

        ws.cell(row=row_idx, column=1).font = Font(
            bold=True, color='006100' if is_mapped else '9C5700')
        ws.cell(row=row_idx, column=1).fill = (
            fill_ok if is_mapped else fill_todo)

        dv.add(ws.cell(row=row_idx, column=15).coordinate)
        row_idx += 1

    tot_eur = round(sum(i['tot_eur'] for _, i in items), 2)
    tot_mov = sum(i['count_movimenti'] for _, i in items)
    eur_mapped = round(eur_mapped, 2)
    eur_todo = round(eur_todo, 2)

    ws.cell(row=row_idx, column=1, value='TOTALE').font = Font(bold=True)
    ws.cell(row=row_idx, column=7, value=tot_mov).font = Font(bold=True)
    ws.cell(row=row_idx, column=8, value=tot_eur).font = Font(bold=True)
    for c in ws[row_idx]:
        c.fill = PatternFill('solid', fgColor='D9E1F2')
        c.border = border
    row_idx += 1

    ws.cell(row=row_idx, column=1, value='   di cui MAPPATI').font = Font(italic=True)
    ws.cell(row=row_idx, column=3, value=n_mapped)
    ws.cell(row=row_idx, column=8, value=eur_mapped)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value='   di cui DA CENSIRE').font = Font(italic=True)
    ws.cell(row=row_idx, column=3, value=n_todo)
    ws.cell(row=row_idx, column=8, value=eur_todo)
    row_idx += 2

    # nota Telepass
    ws.cell(row=row_idx, column=1,
            value=f'Nota: {n_telepass} fatture Telepass canone aggregato '
                  '(no lista apparati nel PDF) NON incluse in questo report.'
           ).font = Font(italic=True, color='808080')
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=16)

    ws.freeze_panes = 'E2'

    out_dir = os.path.join(ROOT, 'output')
    os.makedirs(out_dir, exist_ok=True)
    out_xlsx = os.path.join(
        out_dir,
        f'apparati_completo_network_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(out_xlsx)

    print()
    print(f'PDF Autostrade processati: {n_pdfs.get("Autostrade",0)}')
    print(f'PDF Apcoa processati: {n_pdfs.get("Apcoa",0)}')
    print(f'PDF Telepass canone (no apparati): {n_telepass} — saltati')
    print()
    print(f'Apparati unici totali: {len(items)}')
    print(f'  MAPPATI: {n_mapped} (€ {eur_mapped:.2f})')
    print(f'  DA CENSIRE: {n_todo} (€ {eur_todo:.2f})')
    print(f'\nExcel salvato in:\n  {out_xlsx}')


if __name__ == '__main__':
    main()
